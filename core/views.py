from calendar import monthrange
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import csv
import re
from zipfile import ZipFile, ZIP_DEFLATED
from django.http import HttpResponse
import uuid
import hashlib
import qrcode

from django.urls import reverse

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.platypus import Image

from usuarios.utils import validar_permiso_o_redirigir, tiene_permiso
from usuarios.multiempresa import (
    es_admin_master,
    obtener_empresa_usuario,
    obtener_empresa_activa,
    filtrar_empresa_directa,
    filtrar_empresa_funcionario,
    validar_objeto_empresa_funcionario,
    validar_objeto_empresa_directa,
)

from datetime import datetime
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render, redirect
from django.utils import timezone

from .forms import MarcacionManualForm
from .models import Asistencia

from .forms import (
    ConfiguracionGeneralForm,
    DeudaForm,
    DiaLibreForm,
    EmpresaForm,
    FuncionarioForm,
    LiquidacionForm,
    MarcacionForm,
    PermisoLicenciaForm,
    SucursalForm,
    TurnoForm,
    VacacionForm,
    ComunicacionLaboralForm,
    PlanillaBancariaForm,
    BancoHorasOtorgarForm,
    DocumentoFuncionarioForm,
    HistorialLaboralFuncionarioForm,
    ConductaFuncionarioForm,
    HistorialSalarialFuncionarioForm,
    SuscripcionSistemaForm,
    PagoSuscripcionSistemaForm,
    AjusteManualLiquidacionForm,
    DiaristaForm,
    AsistenciaDiaristaForm,
    PagoDiaristaForm,
)



def _documento_busqueda_q(campo_cedula, campo_tipo, termino):
    limpio = re.sub(r"\D", "", termino or "")
    q_obj = Q(**{f"{campo_cedula}__icontains": termino}) | Q(**{f"{campo_tipo}__icontains": termino})
    if limpio and limpio != termino:
        q_obj |= Q(**{f"{campo_cedula}__icontains": limpio})
    return q_obj

from .liquidacion_utils import (
    calcular_liquidacion_funcionario,
    construir_alertas_proteccion_liquidacion,
)
from .models import (
    Asistencia,
    ConfiguracionGeneral,
    Deuda,
    DiaLibre,
    Empresa,
    Funcionario,
    HistorialAccion,
    Liquidacion,
    AjusteManualLiquidacion,
    NominaMensual,
    AguinaldoAnual,
    CierreNomina,
    PermisoLicencia,
    Sucursal,
    Turno,
    Vacacion,
    PlanillaSemanalFuncionario,
    ComunicacionLaboral,
    PlanillaBancaria,
    BancoHorasMovimiento,
    DocumentoFirmado,
    DocumentoFuncionario,
    HistorialLaboralFuncionario,
    ConductaFuncionario,
    HistorialSalarialFuncionario,
    SuscripcionSistema,
    PagoSuscripcionSistema,
    Diarista,
    AsistenciaDiarista,
    PagoDiarista,
)


def intro_clockin(request):
    if request.user.is_authenticated:
        if es_admin_master(request.user) and not obtener_empresa_activa(request):
            return redirect("panel_global_empresas")
        return redirect("dashboard")

    return render(request, "core/intro.html")


def _bloquear_si_no_admin_master(request):
    admin_master = es_admin_master(request.user)

    if not admin_master:
        messages.error(request, "Este módulo solo está disponible para administración global.")
        return redirect("dashboard")

    return None

@login_required
def marcacion_manual(request):
    permiso = validar_permiso_o_redirigir(request, "asistencia", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = MarcacionManualForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            funcionario = form.cleaned_data["funcionario"]
            tipo = form.cleaned_data["tipo"]
            fecha = form.cleaned_data["fecha"]
            hora = form.cleaned_data["hora"]
            motivo = form.cleaned_data["motivo"]

            if not admin_master:
                if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes registrar asistencia manual para otra empresa.")
                    return redirect("asistencia_marcar")

            fecha_hora_manual = timezone.make_aware(
                datetime.combine(fecha, hora)
            )

            if funcionario.turno and funcionario.turno.hora_salida <= funcionario.turno.hora_entrada:
                if tipo in ["salida", "salida_almuerzo", "regreso_almuerzo"] and hora < funcionario.turno.hora_entrada:
                    fecha_hora_manual = timezone.make_aware(
                        datetime.combine(fecha + timezone.timedelta(days=1), hora)
                    )

            asistencia, creada = Asistencia.objects.get_or_create(
                funcionario=funcionario,
                fecha=fecha,
            )

            if tipo in ["salida_almuerzo", "regreso_almuerzo"] and (
                not funcionario.turno or not funcionario.turno.usa_almuerzo
            ):
                messages.error(request, "Este funcionario no tiene un turno con almuerzo configurado.")
                return redirect("marcacion_manual")

            if tipo == "entrada":
                asistencia.hora_entrada = fecha_hora_manual
                asistencia.calcular_atraso()

            elif tipo == "salida_almuerzo":
                asistencia.hora_salida_almuerzo = fecha_hora_manual

            elif tipo == "regreso_almuerzo":
                asistencia.hora_regreso_almuerzo = fecha_hora_manual

            elif tipo == "salida":
                asistencia.hora_salida = fecha_hora_manual

            asistencia.origen_marcacion = "manual"
            asistencia.marcado_manual_por = request.user
            asistencia.motivo_marcacion_manual = motivo
            asistencia.fecha_hora_real_sistema = timezone.now()

            if tipo == "entrada":
                if asistencia.llego_tarde:
                    asistencia.observacion = f"Entrada manual. Llegó con {asistencia.minutos_atraso} minuto(s) de atraso."
                else:
                    asistencia.observacion = "Entrada manual registrada en horario."

            elif tipo == "salida_almuerzo":
                asistencia.observacion = "Salida a almuerzo manual registrada."

            elif tipo == "regreso_almuerzo":
                asistencia.observacion = "Regreso de almuerzo manual registrado."

            else:
                asistencia.observacion = "Salida final manual registrada."

            asistencia.save()

            registrar_historial(
                request,
                "Asistencia",
                "Marcación manual",
                f"Marcación manual de {tipo} para {funcionario.nombre_completo}. "
                f"Hora registrada: {fecha_hora_manual.strftime('%d/%m/%Y %H:%M:%S')}. "
                f"Operador: {request.user}. Motivo: {motivo}"
            )

            messages.success(
                request,
                f"Marcación manual registrada correctamente para {funcionario.nombre_completo}."
            )
            return redirect("asistencia_marcar")
    else:
        form = MarcacionManualForm()

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "asistencias/marcacion_manual.html", {
        "form": form
    })


def registrar_historial(request, modulo, accion, descripcion):
    empresa = obtener_empresa_activa(request)

    HistorialAccion.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        empresa=empresa,
        modulo=modulo,
        accion=accion,
        descripcion=descripcion,
    )

def obtener_empresa_documento(funcionario=None, empresa=None):
    if empresa:
        return empresa
    if funcionario and getattr(funcionario, "sucursal_rel", None):
        return funcionario.sucursal_rel.empresa
    return None


def construir_encabezado_empresa_pdf(empresa, titulo):
    elementos = []

    datos_empresa = []
    logo_elemento = ""

    if empresa and getattr(empresa, "logo", None):
        try:
            if empresa.logo and empresa.logo.path:
                logo_elemento = Image(empresa.logo.path, width=28 * mm, height=28 * mm)
        except Exception:
            logo_elemento = ""

    nombre = empresa.nombre if empresa else "ClockIn"
    ruc = getattr(empresa, "ruc", "") if empresa else ""
    direccion = getattr(empresa, "direccion", "") if empresa else ""
    telefono = getattr(empresa, "telefono", "") if empresa else ""
    email = getattr(empresa, "email", "") if empresa else ""

    texto_empresa = f"""
    <b>{nombre}</b><br/>
    RUC: {ruc or "-"}<br/>
    Dirección: {direccion or "-"}<br/>
    Teléfono: {telefono or "-"}<br/>
    Email: {email or "-"}
    """

    styles = getSampleStyleSheet()
    empresa_style = ParagraphStyle(
        name="EmpresaHeader",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#111827"),
    )

    titulo_style = ParagraphStyle(
        name="TituloDocumentoEmpresa",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=17,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
    )

    datos_empresa.append([
        logo_elemento,
        Paragraph(texto_empresa, empresa_style),
        Paragraph(f"<b>{titulo}</b>", titulo_style),
    ])

    tabla = Table(datos_empresa, colWidths=[32 * mm, 70 * mm, 68 * mm])
    tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 10))
    return elementos


def agregar_texto_legal_empresa_pdf(elementos, empresa):
    texto = getattr(empresa, "texto_legal_pdf", "") if empresa else ""

    if not texto:
        return

    styles = getSampleStyleSheet()
    legal_style = ParagraphStyle(
        name="TextoLegalEmpresa",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#374151"),
    )

    elementos.append(Spacer(1, 8))
    elementos.append(Paragraph(f"<b>Observación legal / empresarial:</b><br/>{texto}", legal_style))

def generar_codigo_documento(tipo_documento):
    anio = timezone.localdate().year
    tipo = str(tipo_documento or "DOC").upper()[:8]
    corto = uuid.uuid4().hex[:10].upper()
    return f"CLK-{anio}-{tipo}-{corto}"


def crear_documento_firmado(
    request,
    empresa,
    tipo_documento,
    documento_id=None,
    funcionario=None,
    titulo="",
):
    codigo = generar_codigo_documento(tipo_documento)

    base_hash = f"{empresa.id if empresa else ''}|{tipo_documento}|{documento_id}|{funcionario.id if funcionario else ''}|{timezone.now().isoformat()}|{codigo}"
    hash_documento = hashlib.sha256(base_hash.encode("utf-8")).hexdigest()

    documento = DocumentoFirmado.objects.create(
        empresa=empresa,
        codigo=codigo,
        tipo_documento=tipo_documento,
        documento_id=documento_id,
        funcionario=funcionario,
        titulo=titulo or tipo_documento,
        hash_documento=hash_documento,
        firmado_por_nombre=getattr(empresa, "nombre_gerente", "") or "",
        firmado_por_cargo=getattr(empresa, "cargo_gerente", "") or "Gerente General",
        emitido_por=request.user if request.user.is_authenticated else None,
        valido=True,
    )

    return documento


def agregar_firma_qr_documento_pdf(
    elementos,
    request,
    empresa,
    tipo_documento,
    documento_id=None,
    funcionario=None,
    titulo="Documento ClockIn",
):
    documento = crear_documento_firmado(
        request=request,
        empresa=empresa,
        tipo_documento=tipo_documento,
        documento_id=documento_id,
        funcionario=funcionario,
        titulo=titulo,
    )

    styles = getSampleStyleSheet()

    firma_style = ParagraphStyle(
        name="FirmaDigitalInterna",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
    )

    codigo_style = ParagraphStyle(
        name="CodigoVerificacionClockIn",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=7,
        leading=9,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
    )

    url_verificacion = request.build_absolute_uri(
        reverse("verificar_documento", args=[documento.codigo])
    )

    qr = qrcode.make(url_verificacion)
    qr_buffer = BytesIO()
    qr.save(qr_buffer, format="PNG")
    qr_buffer.seek(0)

    qr_img = Image(qr_buffer, width=28 * mm, height=28 * mm)

    firma_img = Spacer(1, 18 * mm)

    if empresa and getattr(empresa, "firma_gerente", None):
        try:
            if empresa.firma_gerente and empresa.firma_gerente.path:
                firma_img = Image(
                    empresa.firma_gerente.path,
                    width=45 * mm,
                    height=18 * mm
                )
        except Exception:
            firma_img = Spacer(1, 18 * mm)

    nombre = documento.firmado_por_nombre or "Firma autorizada"
    cargo = documento.firmado_por_cargo or "Gerencia"

    bloque_firma = [
        firma_img,
        Paragraph(f"<b>{nombre}</b><br/>{cargo}", firma_style),
        Paragraph("Firma interna automatizada", codigo_style),
    ]

    bloque_qr = [
        qr_img,
        Paragraph(f"<b>Código:</b> {documento.codigo}", codigo_style),
        Paragraph(f"Hash: {documento.hash_documento[:18]}...", codigo_style),
        Paragraph("Escanee el QR para verificar este documento.", codigo_style),
    ]

    tabla = Table(
        [[bloque_firma, bloque_qr]],
        colWidths=[85 * mm, 85 * mm]
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e5e7eb")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(Spacer(1, 12))
    elementos.append(tabla)

    return documento

def verificar_documento(request, codigo):
    documento = get_object_or_404(
        DocumentoFirmado.objects.select_related(
            "empresa",
            "funcionario",
            "emitido_por",
        ),
        codigo=codigo
    )

    return render(request, "core/verificar_documento.html", {
        "documento": documento,
    })

def funcionario_tiene_dia_libre(funcionario, fecha=None):
    fecha = fecha or timezone.localdate()
    dia_semana = fecha.weekday()

    return DiaLibre.objects.filter(
        funcionario=funcionario,
        activo=True,
        dia_semana=dia_semana,
    ).filter(
        Q(fecha_inicio__isnull=True) | Q(fecha_inicio__lte=fecha),
        Q(fecha_fin__isnull=True) | Q(fecha_fin__gte=fecha),
    ).exists()

def obtener_fecha_operativa_asistencia(funcionario, ahora=None):
    ahora = ahora or timezone.localtime()
    hoy = ahora.date()

    if not funcionario.turno:
        return hoy

    turno = funcionario.turno

    # Turno normal
    if turno.hora_salida > turno.hora_entrada:
        return hoy

    # Turno nocturno: sale al día siguiente
    ayer = hoy - timezone.timedelta(days=1)

    asistencia_ayer = Asistencia.objects.filter(
        funcionario=funcionario,
        fecha=ayer,
        hora_entrada__isnull=False,
        hora_salida__isnull=True,
    ).first()

    if asistencia_ayer:
        return ayer

    return hoy


def contar_dias_libres_mes(funcionario, mes, anio):
    total = 0
    dias_mes = monthrange(anio, mes)[1]

    for dia in range(1, dias_mes + 1):
        fecha = date(anio, mes, dia)
        if funcionario_tiene_dia_libre(funcionario, fecha):
            total += 1

    return total


def calcular_icl_funcionario_mes(funcionario, mes, anio):
    dias_mes = monthrange(anio, mes)[1]
    total_dias_laborales_estimados = sum(
        1 for dia in range(1, dias_mes + 1)
        if date(anio, mes, dia).weekday() != 6
    )

    dias_libres_mes = contar_dias_libres_mes(funcionario, mes, anio)

    asistencias = Asistencia.objects.filter(
        funcionario=funcionario,
        fecha__year=anio,
        fecha__month=mes,
        hora_entrada__isnull=False,
    )

    asistencias_count = asistencias.count()
    atrasos_count = asistencias.filter(llego_tarde=True).count()

    permisos_aprobados = PermisoLicencia.objects.filter(
        funcionario=funcionario,
        estado=PermisoLicencia.Estados.APROBADO,
        fecha_desde__year=anio,
        fecha_desde__month=mes,
    ).count()

    vacaciones_aprobadas = Vacacion.objects.filter(
        funcionario=funcionario,
        estado=Vacacion.Estados.APROBADO,
        fecha_desde__year=anio,
        fecha_desde__month=mes,
    ).count()

    total_dias_laborales_reales = max(total_dias_laborales_estimados - dias_libres_mes, 0)
    ausencias_estimadas = max(total_dias_laborales_reales - asistencias_count, 0)
    ausencias_justificadas = permisos_aprobados + vacaciones_aprobadas
    ausencias_no_justificadas = max(ausencias_estimadas - ausencias_justificadas, 0)

    icl = 100 - (atrasos_count * 2) - (ausencias_no_justificadas * 5)
    icl = max(0, min(100, icl))

    return {
        "icl": icl,
        "asistencias": asistencias_count,
        "atrasos": atrasos_count,
        "ausencias_estimadas": ausencias_estimadas,
        "ausencias_no_justificadas": ausencias_no_justificadas,
        "dias_libres_mes": dias_libres_mes,
        "total_dias_laborales_reales": total_dias_laborales_reales,
    }


def icl_activo_para_funcionario(funcionario):
    empresa = getattr(getattr(funcionario, "sucursal_rel", None), "empresa", None)
    if empresa is None:
        return True
    return bool(getattr(empresa, "icl_activo", True))


def calcular_bono_pagable_por_icl(bono_base, icl, icl_activo=True):
    bono_base = Decimal(bono_base or 0).quantize(Decimal("0.01"))
    if not icl_activo:
        return bono_base
    return (bono_base * Decimal(icl or 0) / Decimal("100")).quantize(Decimal("0.01"))


def generar_nomina_funcionario(funcionario, mes, anio):
    resumen_icl = calcular_icl_funcionario_mes(funcionario, mes, anio)
    icl_activo = icl_activo_para_funcionario(funcionario)

    salario_base = funcionario.salario_base_aplicable
    bono_base = funcionario.bono_aplicable
    bono_icl = Decimal("0.00") if funcionario.usa_salario_diferenciado else calcular_bono_pagable_por_icl(bono_base, resumen_icl["icl"], icl_activo)
    salario_bruto = funcionario.salario_bruto_aplicable if funcionario.usa_salario_diferenciado else (salario_base + bono_icl).quantize(Decimal("0.01"))
    descuento_ips = funcionario.descuento_ips
    descuento_deudas = funcionario.descuento_deudas_mes

    salario_neto = salario_bruto - descuento_ips - descuento_deudas
    if salario_neto < 0:
        salario_neto = Decimal("0.00")
    salario_neto = salario_neto.quantize(Decimal("0.01"))

    defaults = {
        "salario_base": salario_base,
        "bono_base": bono_base,
        "bono_icl": bono_icl,
        "salario_bruto": salario_bruto,
        "descuento_ips": descuento_ips,
        "descuento_deudas": descuento_deudas,
        "salario_neto": salario_neto,
        "modalidad_cobro": funcionario.modalidad_cobro,
        "banco": funcionario.banco,
        "tipo_cuenta": funcionario.tipo_cuenta,
        "numero_cuenta": funcionario.numero_cuenta,
    }

    nomina, creada = NominaMensual.objects.update_or_create(
        funcionario=funcionario,
        mes=mes,
        anio=anio,
        defaults=defaults,
    )
    return nomina


@login_required
def dashboard(request):
    if getattr(request.user, "rol", "") == "funcionario":
        return redirect("portal_dashboard")

    if es_admin_master(request.user) and not obtener_empresa_activa(request):
        return redirect("panel_global_empresas")

    hoy = timezone.localdate()

    perm_funcionarios = tiene_permiso(request.user, "funcionarios", "puede_ver")
    perm_asistencia = tiene_permiso(request.user, "asistencia", "puede_ver")
    perm_deudas = tiene_permiso(request.user, "deudas", "puede_ver")
    perm_nomina = tiene_permiso(request.user, "nomina", "puede_ver")
    perm_icl = tiene_permiso(request.user, "icl", "puede_ver")
    perm_reportes = tiene_permiso(request.user, "reportes", "puede_ver")
    perm_dias_libres = tiene_permiso(request.user, "dias_libres", "puede_ver")
    perm_liquidacion = tiene_permiso(request.user, "liquidacion", "puede_ver")
    perm_aguinaldo = tiene_permiso(request.user, "aguinaldo", "puede_ver")
    perm_planilla_bancaria = tiene_permiso(request.user, "planilla_bancaria", "puede_ver")
    perm_banco_horas = tiene_permiso(request.user, "banco_horas", "puede_ver")

    empresa_usuario = obtener_empresa_activa(request)
    sucursales_dashboard = Sucursal.objects.none()
    sucursal_seleccionada = None
    sucursal_id = request.GET.get("sucursal", "").strip()

    if empresa_usuario:
        sucursales_dashboard = Sucursal.objects.filter(
            empresa=empresa_usuario,
            activo=True,
        ).order_by("nombre")
        if sucursal_id:
            sucursal_seleccionada = sucursales_dashboard.filter(pk=sucursal_id).first()

    funcionarios_qs = Funcionario.objects.filter(activo=True)
    asistencias_hoy_qs = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__turno"
    ).filter(
        fecha=hoy,
        funcionario__activo=True
    )
    deudas_qs = Deuda.objects.filter(activa=True)

    if empresa_usuario:
        funcionarios_qs = funcionarios_qs.filter(sucursal_rel__empresa=empresa_usuario)
        asistencias_hoy_qs = asistencias_hoy_qs.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        deudas_qs = deudas_qs.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        if sucursal_seleccionada:
            funcionarios_qs = funcionarios_qs.filter(sucursal_rel=sucursal_seleccionada)
            asistencias_hoy_qs = asistencias_hoy_qs.filter(funcionario__sucursal_rel=sucursal_seleccionada)
            deudas_qs = deudas_qs.filter(funcionario__sucursal_rel=sucursal_seleccionada)
    else:
        funcionarios_qs = funcionarios_qs.none()
        asistencias_hoy_qs = asistencias_hoy_qs.none()
        deudas_qs = deudas_qs.none()

    total_funcionarios = 0
    presentes_hoy = 0
    llegadas_tarde_hoy = 0
    salidas_hoy = 0
    pendientes_hoy = 0
    trabajando_hoy = 0
    en_almuerzo_hoy = 0
    finalizados_hoy = 0
    ultimas_marcaciones = []
    funcionarios_recientes = []
    total_salario_bruto = Decimal("0.00")
    total_salario_neto = Decimal("0.00")
    total_deudas_funcionarios = Decimal("0.00")
    funcionarios_con_ips = 0
    aporte_patronal_ips_estimado = Decimal("0.00")
    salario_minimo_ips = Decimal("2899048.00")
    porcentaje_patronal_ips = Decimal("0.255")
    tasa_asistencia = 0
    tasa_puntualidad = 0
    tasa_ips = 0
    deuda_promedio = Decimal("0.00")
    estados_jornada_dashboard = []
    sucursales_metricas = []

    if perm_funcionarios:
        total_funcionarios = funcionarios_qs.count()

        funcionarios_recientes = funcionarios_qs.select_related(
            "turno",
            "sucursal_rel",
            "sucursal_rel__empresa"
        ).order_by("-creado_en")[:6]

    if perm_asistencia:
        presentes_hoy = asistencias_hoy_qs.filter(hora_entrada__isnull=False).count()
        llegadas_tarde_hoy = asistencias_hoy_qs.filter(llego_tarde=True).count()
        salidas_hoy = asistencias_hoy_qs.filter(hora_salida__isnull=False).count()

        if perm_funcionarios:
            pendientes_hoy = max(total_funcionarios - presentes_hoy, 0)

        for asistencia in asistencias_hoy_qs:
            estado = asistencia.estado_jornada
            if estado == "Trabajando":
                trabajando_hoy += 1
            elif estado == "En almuerzo":
                en_almuerzo_hoy += 1
            elif estado == "Finalizado":
                finalizados_hoy += 1

        ultimas_marcaciones = asistencias_hoy_qs.order_by("-actualizado_en")[:8]

    if perm_nomina:
        for funcionario in funcionarios_qs:
            total_salario_bruto += funcionario.salario_bruto
            total_salario_neto += funcionario.salario_neto_estimado
            if funcionario.ips:
                funcionarios_con_ips += 1

        aporte_patronal_por_funcionario = (salario_minimo_ips * porcentaje_patronal_ips).quantize(Decimal("0.01"))
        aporte_patronal_ips_estimado = (aporte_patronal_por_funcionario * Decimal(funcionarios_con_ips)).quantize(Decimal("0.01"))

    if perm_deudas:
        total_deudas_funcionarios = deudas_qs.aggregate(
            total=Sum("saldo_pendiente")
        )["total"] or Decimal("0.00")

    if total_funcionarios:
        tasa_asistencia = round((presentes_hoy / total_funcionarios) * 100)
        puntuales_hoy = max(presentes_hoy - llegadas_tarde_hoy, 0)
        tasa_puntualidad = round((puntuales_hoy / total_funcionarios) * 100)
        tasa_ips = round((funcionarios_con_ips / total_funcionarios) * 100) if perm_nomina else 0
        deuda_promedio = (total_deudas_funcionarios / Decimal(total_funcionarios)).quantize(Decimal("0.01")) if perm_deudas else Decimal("0.00")

    if perm_asistencia:
        estados_jornada_dashboard = [
            {"label": "Trabajando", "value": trabajando_hoy, "percent": round((trabajando_hoy / total_funcionarios) * 100) if total_funcionarios else 0, "class": "ok"},
            {"label": "En almuerzo", "value": en_almuerzo_hoy, "percent": round((en_almuerzo_hoy / total_funcionarios) * 100) if total_funcionarios else 0, "class": "warn"},
            {"label": "Finalizados", "value": finalizados_hoy, "percent": round((finalizados_hoy / total_funcionarios) * 100) if total_funcionarios else 0, "class": "info"},
            {"label": "Pendientes", "value": pendientes_hoy, "percent": round((pendientes_hoy / total_funcionarios) * 100) if total_funcionarios else 0, "class": "danger"},
        ]

    if empresa_usuario and perm_funcionarios:
        sucursales_para_metricas = [sucursal_seleccionada] if sucursal_seleccionada else list(sucursales_dashboard)
        for sucursal in sucursales_para_metricas:
            if not sucursal:
                continue
            total_sucursal = funcionarios_qs.filter(sucursal_rel=sucursal).count()
            presentes_sucursal = asistencias_hoy_qs.filter(funcionario__sucursal_rel=sucursal, hora_entrada__isnull=False).count() if perm_asistencia else 0
            tarde_sucursal = asistencias_hoy_qs.filter(funcionario__sucursal_rel=sucursal, llego_tarde=True).count() if perm_asistencia else 0
            pendientes_sucursal = max(total_sucursal - presentes_sucursal, 0)
            sucursales_metricas.append({
                "nombre": sucursal.nombre,
                "total": total_sucursal,
                "presentes": presentes_sucursal,
                "tarde": tarde_sucursal,
                "pendientes": pendientes_sucursal,
                "percent": round((presentes_sucursal / total_sucursal) * 100) if total_sucursal else 0,
            })

    context = {
        "titulo": "Dashboard ClockIn",
        "hoy": hoy,
        "empresa_usuario": empresa_usuario,
        "sucursales_dashboard": sucursales_dashboard,
        "sucursal_seleccionada": sucursal_seleccionada,
        "sucursal_seleccionada_id": str(sucursal_seleccionada.id) if sucursal_seleccionada else "",

        "total_funcionarios": total_funcionarios,
        "presentes_hoy": presentes_hoy,
        "llegadas_tarde_hoy": llegadas_tarde_hoy,
        "salidas_hoy": salidas_hoy,
        "pendientes_hoy": pendientes_hoy,
        "trabajando_hoy": trabajando_hoy,
        "en_almuerzo_hoy": en_almuerzo_hoy,
        "finalizados_hoy": finalizados_hoy,
        "ultimas_marcaciones": ultimas_marcaciones,
        "funcionarios_recientes": funcionarios_recientes,
        "total_salario_bruto": total_salario_bruto,
        "total_salario_neto": total_salario_neto,
        "funcionarios_con_ips": funcionarios_con_ips,
        "aporte_patronal_ips_estimado": aporte_patronal_ips_estimado,
        "total_deudas_funcionarios": total_deudas_funcionarios,
        "tasa_asistencia": tasa_asistencia,
        "tasa_puntualidad": tasa_puntualidad,
        "tasa_ips": tasa_ips,
        "deuda_promedio": deuda_promedio,
        "estados_jornada_dashboard": estados_jornada_dashboard,
        "sucursales_metricas": sucursales_metricas,

        "perm_funcionarios": perm_funcionarios,
        "perm_asistencia": perm_asistencia,
        "perm_deudas": perm_deudas,
        "perm_nomina": perm_nomina,
        "perm_icl": perm_icl,
        "perm_reportes": perm_reportes,
        "perm_dias_libres": perm_dias_libres,
        "perm_liquidacion": perm_liquidacion,
        "perm_aguinaldo": perm_aguinaldo,
        "perm_planilla_bancaria": perm_planilla_bancaria,
        "perm_banco_horas": perm_banco_horas,
        "es_admin_master": es_admin_master(request.user),
    }
    return render(request, "core/dashboard.html", context)


@login_required
def panel_global_empresas(request):
    if not es_admin_master(request.user):
        empresa = obtener_empresa_activa(request)
        if empresa:
            request.session["empresa_activa_id"] = empresa.id
        return redirect("dashboard")

    hoy = timezone.localdate()

    empresas = Empresa.objects.annotate(
        total_sucursales=Count("sucursales", distinct=True),
    ).order_by("nombre")

    tarjetas = []
    total_funcionarios = 0
    total_sucursales = 0
    proximas_vencer = 0
    pagos_pendientes = 0

    for empresa in empresas:
        funcionarios_activos = Funcionario.objects.filter(
            activo=True,
            sucursal_rel__empresa=empresa,
        ).count()
        sucursales_count = empresa.total_sucursales

        total_funcionarios += funcionarios_activos
        total_sucursales += sucursales_count

        dias_vencimiento = None
        if empresa.fecha_vencimiento_suscripcion:
            dias_vencimiento = (empresa.fecha_vencimiento_suscripcion - hoy).days
            if 0 <= dias_vencimiento <= 7:
                proximas_vencer += 1
            if dias_vencimiento < 0:
                pagos_pendientes += 1

        tarjetas.append({
            "empresa": empresa,
            "funcionarios_activos": funcionarios_activos,
            "sucursales_count": sucursales_count,
            "dias_vencimiento": dias_vencimiento,
        })

    total_empresas = empresas.count()
    empresas_activas = empresas.filter(estado=Empresa.Estados.ACTIVA, activo=True).count()
    empresas_suspendidas = empresas.filter(Q(estado=Empresa.Estados.SUSPENDIDA) | Q(activo=False)).count()

    return render(request, "core/panel_global_empresas.html", {
        "tarjetas": tarjetas,
        "total_empresas": total_empresas,
        "empresas_activas": empresas_activas,
        "empresas_suspendidas": empresas_suspendidas,
        "total_funcionarios": total_funcionarios,
        "total_sucursales": total_sucursales,
        "proximas_vencer": proximas_vencer,
        "pagos_pendientes": pagos_pendientes,
    })


@login_required
def empresa_entrar(request, pk):
    if not es_admin_master(request.user):
        messages.error(request, "Solo el administrador master puede cambiar de empresa.")
        return redirect("dashboard")

    empresa = get_object_or_404(Empresa, pk=pk)
    request.session["empresa_activa_id"] = empresa.id
    messages.success(request, f"Entraste a la empresa {empresa.nombre_visible}.")
    return redirect("dashboard")


@login_required
def empresa_activa_salir(request):
    if es_admin_master(request.user):
        request.session.pop("empresa_activa_id", None)
        return redirect("panel_global_empresas")
    return redirect("dashboard")


@login_required
def empresa_detalle_global(request, pk):
    if not es_admin_master(request.user):
        messages.error(request, "Solo el administrador master puede ver el detalle global.")
        return redirect("dashboard")

    empresa = get_object_or_404(Empresa, pk=pk)
    funcionarios_activos = Funcionario.objects.filter(activo=True, sucursal_rel__empresa=empresa).count()
    funcionarios_total = Funcionario.objects.filter(sucursal_rel__empresa=empresa).count()
    sucursales = Sucursal.objects.filter(empresa=empresa).order_by("nombre")
    usuarios = empresa.usuarios.order_by("first_name", "last_name", "username")

    metricas = {
        "funcionarios_activos": funcionarios_activos,
        "funcionarios_total": funcionarios_total,
        "sucursales": sucursales.count(),
        "turnos": Turno.objects.filter(empresa=empresa).count(),
        "nominas": NominaMensual.objects.filter(empresa=empresa).count(),
        "liquidaciones": Liquidacion.objects.filter(empresa=empresa).count(),
        "comunicaciones": ComunicacionLaboral.objects.filter(empresa=empresa).count(),
    }

    return render(request, "core/empresa_detalle_global.html", {
        "empresa": empresa,
        "metricas": metricas,
        "sucursales": sucursales,
        "usuarios": usuarios,
    })

@login_required
def empresas_lista(request):
    permiso = validar_permiso_o_redirigir(request, "empresas", "puede_ver")
    if permiso:
        return permiso
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    q = request.GET.get("q", "").strip()
    empresas = Empresa.objects.all()

    if q:
        empresas = empresas.filter(
            Q(nombre__icontains=q) |
            Q(ruc__icontains=q)
        )

    return render(request, "core/empresas_lista.html", {
        "empresas": empresas.order_by("nombre"),
        "q": q,
    })


@login_required
def empresa_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "empresas", "puede_crear")
    if permiso:
        return permiso
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    if request.method == "POST":
        form = EmpresaForm(request.POST, request.FILES)
        if form.is_valid():
            empresa = form.save()
            registrar_historial(
                request,
                "Empresas",
                "Crear",
                f"Se creó la empresa {empresa.nombre}."
            )
            messages.success(request, "Empresa creada correctamente.")
            return redirect("empresas_lista")
    else:
        form = EmpresaForm()

    return render(request, "core/empresa_form.html", {
        "form": form,
        "titulo_form": "Nueva empresa",
        "boton_texto": "Guardar empresa",
    })


@login_required
def empresa_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "empresas", "puede_editar")
    if permiso:
        return permiso
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    empresa = get_object_or_404(Empresa, pk=pk)

    if request.method == "POST":
        form = EmpresaForm(request.POST, request.FILES, instance=empresa)
        if form.is_valid():
            form.save()
            registrar_historial(
                request,
                "Empresas",
                "Editar",
                f"Se editó la empresa {empresa.nombre}."
            )
            messages.success(request, "Empresa actualizada correctamente.")
            return redirect("empresas_lista")
    else:
        form = EmpresaForm(instance=empresa)

    return render(request, "core/empresa_form.html", {
        "form": form,
        "titulo_form": f"Editar empresa: {empresa.nombre}",
        "boton_texto": "Guardar cambios",
        "empresa": empresa,
    })


@login_required
def empresa_toggle_activo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "empresas", "puede_editar")
    if permiso:
        return permiso
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    empresa = get_object_or_404(Empresa, pk=pk)
    empresa.activo = not empresa.activo
    empresa.save()

    estado = "activada" if empresa.activo else "inactivada"
    registrar_historial(
        request,
        "Empresas",
        "Cambio de estado",
        f"Empresa {empresa.nombre} fue {estado}."
    )
    messages.success(request, f"Empresa {estado} correctamente.")
    return redirect("empresas_lista")


@login_required
def sucursales_lista(request):
    permiso = validar_permiso_o_redirigir(request, "sucursales", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()
    empresa_id = request.GET.get("empresa", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    sucursales = Sucursal.objects.select_related("empresa").all()

    if not admin_master:
        if empresa_usuario:
            empresa_id = str(empresa_usuario.id)
            sucursales = sucursales.filter(empresa=empresa_usuario)
        else:
            sucursales = sucursales.none()

    if q:
        sucursales = sucursales.filter(
            Q(nombre__icontains=q) |
            Q(direccion__icontains=q) |
            Q(empresa__nombre__icontains=q)
        )

    if empresa_id:
        sucursales = sucursales.filter(empresa_id=empresa_id)

    if admin_master:
        empresas = Empresa.objects.filter(activo=True).order_by("nombre")
    else:
        empresas = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()

    return render(request, "core/sucursales_lista.html", {
        "sucursales": sucursales.order_by("empresa__nombre", "nombre"),
        "empresas": empresas,
        "empresa_id": empresa_id,
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def sucursal_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "sucursales", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = SucursalForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario

        if form.is_valid():
            sucursal = form.save(commit=False)

            if not admin_master:
                if not empresa_usuario:
                    messages.error(request, "Tu usuario no tiene empresa asignada.")
                    return redirect("sucursales_lista")

                sucursal.empresa = empresa_usuario

            sucursal.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Sucursales",
                "Crear",
                f"Se creó la sucursal {sucursal.nombre} de {sucursal.empresa.nombre}."
            )
            messages.success(request, "Sucursal creada correctamente.")
            return redirect("sucursales_lista")
    else:
        form = SucursalForm()

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario

    return render(request, "core/sucursal_form.html", {
        "form": form,
        "titulo_form": "Nueva sucursal",
        "boton_texto": "Guardar sucursal",
    })


@login_required
def sucursal_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "sucursales", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    sucursal = get_object_or_404(
        Sucursal.objects.select_related("empresa"),
        pk=pk
    )

    if not admin_master:
        if sucursal.empresa != empresa_usuario:
            messages.error(request, "No puedes editar sucursales de otra empresa.")
            return redirect("sucursales_lista")

    if request.method == "POST":
        form = SucursalForm(request.POST, instance=sucursal)

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario

        if form.is_valid():
            sucursal_editada = form.save(commit=False)

            if not admin_master:
                sucursal_editada.empresa = empresa_usuario

            sucursal_editada.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Sucursales",
                "Editar",
                f"Se editó la sucursal {sucursal_editada.nombre} de {sucursal_editada.empresa.nombre}."
            )
            messages.success(request, "Sucursal actualizada correctamente.")
            return redirect("sucursales_lista")
    else:
        form = SucursalForm(instance=sucursal)

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario

    return render(request, "core/sucursal_form.html", {
        "form": form,
        "titulo_form": f"Editar sucursal: {sucursal.nombre}",
        "boton_texto": "Guardar cambios",
        "sucursal": sucursal,
    })


@login_required
def sucursal_toggle_activo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "sucursales", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    sucursal = get_object_or_404(Sucursal.objects.select_related("empresa"), pk=pk)

    if not admin_master:
        if not empresa_usuario or sucursal.empresa != empresa_usuario:
            messages.error(request, "No puedes cambiar sucursales de otra empresa.")
            return redirect("sucursales_lista")

    sucursal.activo = not sucursal.activo
    sucursal.save()

    estado = "activada" if sucursal.activo else "inactivada"
    registrar_historial(
        request,
        "Sucursales",
        "Cambio de estado",
        f"Sucursal {sucursal.nombre} fue {estado}."
    )
    messages.success(request, f"Sucursal {estado} correctamente.")
    return redirect("sucursales_lista")


@login_required
def obtener_sucursales_por_empresa(request):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_ver")
    if permiso:
        return JsonResponse({"sucursales": []}, status=403)

    empresa_id = request.GET.get("empresa_id", "").strip()
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if not empresa_id:
        return JsonResponse({"sucursales": []})

    if not admin_master:
        if not empresa_usuario or empresa_id != str(empresa_usuario.id):
            return JsonResponse({"sucursales": []}, status=403)

    sucursal_actual_id = request.GET.get("sucursal_actual_id", "").strip()

    sucursales = Sucursal.objects.filter(
        empresa_id=empresa_id,
        activo=True
    )

    if sucursal_actual_id:
        sucursales = Sucursal.objects.filter(
            Q(empresa_id=empresa_id, activo=True) |
            Q(pk=sucursal_actual_id, empresa_id=empresa_id)
        )

    sucursales = sucursales.order_by("nombre")

    data = [
        {
            "id": s.id,
            "nombre": s.nombre,
            "activo": s.activo,
        }
        for s in sucursales
    ]
    return JsonResponse({"sucursales": data})


def _usuario_puede_operar_diarista(request, diarista):
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None
    if admin_master:
        return True
    return bool(empresa_usuario and diarista.empresa_id == empresa_usuario.id)


def _datetime_jornada(fecha, hora):
    if not fecha or not hora:
        return None
    valor = datetime.combine(fecha, hora)
    if timezone.is_naive(valor):
        return timezone.make_aware(valor, timezone.get_current_timezone())
    return valor


def _actualizar_calculos_asistencia_diarista(asistencia):
    asistencia.recalcular_minutos_trabajados()
    asistencia.minutos_atraso = 0
    if asistencia.hora_entrada and asistencia.diarista and asistencia.diarista.turno:
        turno = asistencia.diarista.turno
        hora_entrada = timezone.localtime(asistencia.hora_entrada).time()
        base = datetime.combine(asistencia.fecha, turno.hora_entrada)
        real = datetime.combine(asistencia.fecha, hora_entrada)
        tolerancia = int(turno.tolerancia_minutos or 0)
        atraso = int((real - base).total_seconds() // 60) - tolerancia
        asistencia.minutos_atraso = max(atraso, 0)
    return asistencia


@login_required
def diaristas_lista(request):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()
    empresa_id = request.GET.get("empresa", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    sector = request.GET.get("sector", "").strip()
    estado = request.GET.get("estado", "").strip()
    desde = request.GET.get("desde", "").strip()
    hasta = request.GET.get("hasta", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    diaristas = Diarista.objects.select_related("empresa", "sucursal", "turno").all()

    if not admin_master:
        if empresa_usuario:
            diaristas = diaristas.filter(empresa=empresa_usuario)
            empresa_id = str(empresa_usuario.id)
        else:
            diaristas = diaristas.none()

    if admin_master and empresa_id:
        diaristas = diaristas.filter(empresa_id=empresa_id)

    if sucursal_id:
        diaristas = diaristas.filter(sucursal_id=sucursal_id)

    if sector:
        diaristas = diaristas.filter(sector=sector)

    if estado:
        diaristas = diaristas.filter(estado=estado)

    if desde:
        diaristas = diaristas.filter(fecha_fin__gte=desde)

    if hasta:
        diaristas = diaristas.filter(fecha_inicio__lte=hasta)

    if q:
        diaristas = diaristas.filter(
            Q(nombres__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(cedula__icontains=q)
            | Q(funcion_temporal__icontains=q)
            | Q(sector__icontains=q)
            | Q(sucursal__nombre__icontains=q)
        )

    hoy = timezone.localdate()
    mes_inicio = hoy.replace(day=1)
    base_stats = Diarista.objects.select_related("empresa", "sucursal")
    if not admin_master:
        base_stats = base_stats.filter(empresa=empresa_usuario) if empresa_usuario else base_stats.none()
    elif empresa_id:
        base_stats = base_stats.filter(empresa_id=empresa_id)

    activos_actuales = base_stats.filter(
        estado=Diarista.Estados.ACTIVO,
        fecha_inicio__lte=hoy,
        fecha_fin__gte=hoy,
    ).count()
    pendientes = base_stats.filter(estado=Diarista.Estados.PENDIENTE).count()
    finalizados_mes = base_stats.filter(estado=Diarista.Estados.FINALIZADO, fecha_fin__gte=mes_inicio, fecha_fin__lte=hoy).count()
    total_estimado = sum((item.total_estimado for item in base_stats.filter(estado__in=[Diarista.Estados.ACTIVO, Diarista.Estados.PENDIENTE])), Decimal("0"))
    pagos_pendientes = PagoDiarista.objects.filter(estado__in=[PagoDiarista.Estados.BORRADOR, PagoDiarista.Estados.GENERADO])
    if not admin_master:
        pagos_pendientes = pagos_pendientes.filter(empresa=empresa_usuario) if empresa_usuario else pagos_pendientes.none()
    elif empresa_id:
        pagos_pendientes = pagos_pendientes.filter(empresa_id=empresa_id)

    if admin_master:
        empresas = Empresa.objects.filter(activo=True).order_by("nombre")
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
        if empresa_id:
            sucursales = sucursales.filter(empresa_id=empresa_id)
        sectores_qs = Diarista.objects.filter(empresa_id=empresa_id) if empresa_id else Diarista.objects.all()
    else:
        empresas = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()
        sucursales = Sucursal.objects.filter(empresa=empresa_usuario, activo=True).order_by("nombre") if empresa_usuario else Sucursal.objects.none()
        sectores_qs = Diarista.objects.filter(empresa=empresa_usuario) if empresa_usuario else Diarista.objects.none()

    sectores = sectores_qs.exclude(sector="").values_list("sector", flat=True).distinct().order_by("sector")

    return render(request, "core/diaristas_lista.html", {
        "diaristas": diaristas.order_by("-fecha_inicio", "apellidos", "nombres"),
        "q": q,
        "empresas": empresas,
        "sucursales": sucursales,
        "sectores": sectores,
        "empresa_id": empresa_id,
        "sucursal_id": sucursal_id,
        "sector_sel": sector,
        "estado_sel": estado,
        "desde": desde,
        "hasta": hasta,
        "estados": Diarista.Estados.choices,
        "activos_actuales": activos_actuales,
        "pendientes": pendientes,
        "finalizados_mes": finalizados_mes,
        "total_estimado": total_estimado,
        "pagos_pendientes": pagos_pendientes.count(),
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def diarista_detalle(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not admin_master:
        if not empresa_usuario or diarista.empresa != empresa_usuario:
            messages.error(request, "No puedes ver diaristas de otra empresa.")
            return redirect("diaristas_lista")

    asistencias = diarista.asistencias.select_related("pago").order_by("-fecha")[:20]
    pagos = diarista.pagos.order_by("-fecha_pago", "-creado_en")[:12]
    dias_trabajados = diarista.asistencias.filter(estado=AsistenciaDiarista.Estados.TRABAJADO).count()
    ausencias = diarista.asistencias.filter(estado=AsistenciaDiarista.Estados.AUSENTE).count()
    pagos_total = diarista.pagos.exclude(estado=PagoDiarista.Estados.ANULADO).aggregate(total=Sum("total_pagado"))["total"] or Decimal("0")

    return render(request, "core/diarista_detalle.html", {
        "diarista": diarista,
        "asistencias": asistencias,
        "pagos": pagos,
        "dias_trabajados": dias_trabajados,
        "ausencias": ausencias,
        "pagos_total": pagos_total,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def diarista_asistencias(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes ver asistencias de diaristas de otra empresa.")
        return redirect("diaristas_lista")

    desde = request.GET.get("desde", "").strip()
    hasta = request.GET.get("hasta", "").strip()
    estado = request.GET.get("estado", "").strip()

    asistencias = diarista.asistencias.select_related("pago").all()
    if desde:
        asistencias = asistencias.filter(fecha__gte=desde)
    if hasta:
        asistencias = asistencias.filter(fecha__lte=hasta)
    if estado:
        asistencias = asistencias.filter(estado=estado)

    trabajadas = asistencias.filter(estado=AsistenciaDiarista.Estados.TRABAJADO).count()
    ausencias = asistencias.filter(estado=AsistenciaDiarista.Estados.AUSENTE).count()
    incompletas = asistencias.filter(estado=AsistenciaDiarista.Estados.INCOMPLETO).count()
    minutos_total = asistencias.aggregate(total=Sum("minutos_trabajados"))["total"] or 0
    horas_total = f"{int(minutos_total // 60)}h {int(minutos_total % 60):02d}m"

    return render(request, "core/diarista_asistencias.html", {
        "diarista": diarista,
        "asistencias": asistencias.order_by("-fecha"),
        "desde": desde,
        "hasta": hasta,
        "estado_sel": estado,
        "estados": AsistenciaDiarista.Estados.choices,
        "trabajadas": trabajadas,
        "ausencias": ausencias,
        "incompletas": incompletas,
        "horas_total": horas_total,
    })


@login_required
def diarista_asistencia_nueva(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes registrar asistencias para otra empresa.")
        return redirect("diaristas_lista")

    if request.method == "POST":
        form = AsistenciaDiaristaForm(request.POST)
        if form.is_valid():
            asistencia = form.save(commit=False)
            asistencia.diarista = diarista
            asistencia.empresa = diarista.empresa
            asistencia.sucursal = diarista.sucursal
            asistencia.origen_marcacion = "manual"
            asistencia.hora_entrada = _datetime_jornada(asistencia.fecha, form.cleaned_data.get("hora_entrada_simple"))
            asistencia.hora_salida = _datetime_jornada(asistencia.fecha, form.cleaned_data.get("hora_salida_simple"))
            _actualizar_calculos_asistencia_diarista(asistencia)
            try:
                asistencia.save()
            except IntegrityError:
                messages.error(request, "Ya existe una jornada para esa fecha.")
                return redirect("diarista_asistencias", pk=diarista.pk)
            registrar_historial(request, "Diaristas", "Asistencia", f"Se registró asistencia de diarista {diarista.nombre_completo} para {asistencia.fecha:%d/%m/%Y}.")
            messages.success(request, "Asistencia registrada correctamente.")
            return redirect("diarista_asistencias", pk=diarista.pk)
    else:
        form = AsistenciaDiaristaForm(initial={"fecha": timezone.localdate()})

    return render(request, "core/diarista_asistencia_form.html", {
        "form": form,
        "diarista": diarista,
        "titulo_form": "Nueva jornada",
        "boton_texto": "Guardar jornada",
    })


@login_required
def diarista_asistencia_editar(request, pk, asistencia_id):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    asistencia = get_object_or_404(AsistenciaDiarista.objects.select_related("diarista", "pago"), pk=asistencia_id, diarista=diarista)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes editar asistencias de otra empresa.")
        return redirect("diaristas_lista")
    if asistencia.pago_estado in [AsistenciaDiarista.EstadosPago.INCLUIDO, AsistenciaDiarista.EstadosPago.PAGADO]:
        messages.error(request, "No puedes editar una jornada incluida en un pago.")
        return redirect("diarista_asistencias", pk=diarista.pk)

    if request.method == "POST":
        form = AsistenciaDiaristaForm(request.POST, instance=asistencia)
        if form.is_valid():
            asistencia_editada = form.save(commit=False)
            asistencia_editada.hora_entrada = _datetime_jornada(asistencia_editada.fecha, form.cleaned_data.get("hora_entrada_simple"))
            asistencia_editada.hora_salida = _datetime_jornada(asistencia_editada.fecha, form.cleaned_data.get("hora_salida_simple"))
            _actualizar_calculos_asistencia_diarista(asistencia_editada)
            asistencia_editada.save()
            registrar_historial(request, "Diaristas", "Editar asistencia", f"Se editó asistencia de diarista {diarista.nombre_completo} para {asistencia_editada.fecha:%d/%m/%Y}.")
            messages.success(request, "Asistencia actualizada correctamente.")
            return redirect("diarista_asistencias", pk=diarista.pk)
    else:
        form = AsistenciaDiaristaForm(instance=asistencia)

    return render(request, "core/diarista_asistencia_form.html", {
        "form": form,
        "diarista": diarista,
        "asistencia": asistencia,
        "titulo_form": "Editar jornada",
        "boton_texto": "Guardar cambios",
    })


@login_required
def diarista_asistencia_ausente(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes registrar ausencias para otra empresa.")
        return redirect("diaristas_lista")

    fecha_txt = request.POST.get("fecha") or request.GET.get("fecha") or timezone.localdate().isoformat()
    try:
        fecha = datetime.strptime(fecha_txt, "%Y-%m-%d").date()
    except ValueError:
        fecha = timezone.localdate()

    asistencia, creada = AsistenciaDiarista.objects.get_or_create(
        diarista=diarista,
        fecha=fecha,
        defaults={
            "empresa": diarista.empresa,
            "sucursal": diarista.sucursal,
            "estado": AsistenciaDiarista.Estados.AUSENTE,
            "origen_marcacion": "manual",
            "observacion": "Ausencia registrada manualmente.",
        }
    )
    if not creada and asistencia.pago_estado in [AsistenciaDiarista.EstadosPago.INCLUIDO, AsistenciaDiarista.EstadosPago.PAGADO]:
        messages.error(request, "No puedes modificar una jornada incluida en un pago.")
        return redirect("diarista_asistencias", pk=diarista.pk)
    if not creada:
        asistencia.hora_entrada = None
        asistencia.hora_salida = None
        asistencia.minutos_trabajados = 0
        asistencia.minutos_atraso = 0
        asistencia.estado = AsistenciaDiarista.Estados.AUSENTE
        asistencia.observacion = asistencia.observacion or "Ausencia registrada manualmente."
        asistencia.save()

    registrar_historial(request, "Diaristas", "Ausencia", f"Se registró ausencia de diarista {diarista.nombre_completo} para {fecha:%d/%m/%Y}.")
    messages.success(request, "Ausencia registrada correctamente.")
    return redirect("diarista_asistencias", pk=diarista.pk)


def _horas_desde_minutos(minutos):
    minutos = int(minutos or 0)
    return f"{minutos // 60}h {minutos % 60:02d}m"


def _jornadas_pagables_diarista(diarista, desde, hasta):
    return AsistenciaDiarista.objects.filter(
        diarista=diarista,
        empresa=diarista.empresa,
        fecha__gte=desde,
        fecha__lte=hasta,
        estado=AsistenciaDiarista.Estados.TRABAJADO,
        pago_estado__in=[AsistenciaDiarista.EstadosPago.PENDIENTE, AsistenciaDiarista.EstadosPago.REABIERTO],
        pago__isnull=True,
    ).order_by("fecha")


def _asignar_numero_pago_diarista(pago):
    if not pago.numero_comprobante:
        pago.numero_comprobante = f"DIA-{pago.fecha_pago:%Y%m%d}-{pago.pk:05d}"
        pago.save(update_fields=["numero_comprobante"])
    return pago.numero_comprobante


@login_required
def diaristas_reporte(request):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    desde = request.GET.get("desde", hoy.replace(day=1).isoformat()).strip()
    hasta = request.GET.get("hasta", hoy.isoformat()).strip()
    empresa_id = request.GET.get("empresa", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    estado = request.GET.get("estado", "").strip()
    q = request.GET.get("q", "").strip()
    export = request.GET.get("export", "").strip().lower()

    try:
        desde_fecha = datetime.strptime(desde, "%Y-%m-%d").date()
    except ValueError:
        desde_fecha = hoy.replace(day=1)
        desde = desde_fecha.isoformat()
    try:
        hasta_fecha = datetime.strptime(hasta, "%Y-%m-%d").date()
    except ValueError:
        hasta_fecha = hoy
        hasta = hasta_fecha.isoformat()
    if hasta_fecha < desde_fecha:
        desde_fecha, hasta_fecha = hasta_fecha, desde_fecha
        desde, hasta = desde_fecha.isoformat(), hasta_fecha.isoformat()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    diaristas = Diarista.objects.select_related("empresa", "sucursal", "turno").all()
    if not admin_master:
        if empresa_usuario:
            diaristas = diaristas.filter(empresa=empresa_usuario)
            empresa_id = str(empresa_usuario.id)
        else:
            diaristas = diaristas.none()
    elif empresa_id:
        diaristas = diaristas.filter(empresa_id=empresa_id)

    if sucursal_id:
        diaristas = diaristas.filter(sucursal_id=sucursal_id)
    if estado:
        diaristas = diaristas.filter(estado=estado)
    if q:
        diaristas = diaristas.filter(
            Q(nombres__icontains=q)
            | Q(apellidos__icontains=q)
            | Q(cedula__icontains=q)
            | Q(funcion_temporal__icontains=q)
            | Q(sector__icontains=q)
            | Q(sucursal__nombre__icontains=q)
        )

    asistencias_base = AsistenciaDiarista.objects.filter(
        diarista__in=diaristas,
        fecha__gte=desde_fecha,
        fecha__lte=hasta_fecha,
    )
    pagos_base = PagoDiarista.objects.filter(
        diarista__in=diaristas,
        fecha_pago__gte=desde_fecha,
        fecha_pago__lte=hasta_fecha,
    )

    total_diaristas = diaristas.count()
    jornadas_trabajadas = asistencias_base.filter(estado=AsistenciaDiarista.Estados.TRABAJADO).count()
    ausencias = asistencias_base.filter(estado=AsistenciaDiarista.Estados.AUSENTE).count()
    incompletas = asistencias_base.filter(estado=AsistenciaDiarista.Estados.INCOMPLETO).count()
    pendientes_pago = asistencias_base.filter(
        estado=AsistenciaDiarista.Estados.TRABAJADO,
        pago_estado__in=[AsistenciaDiarista.EstadosPago.PENDIENTE, AsistenciaDiarista.EstadosPago.REABIERTO],
        pago__isnull=True,
    ).count()
    minutos_total = asistencias_base.aggregate(total=Sum("minutos_trabajados"))["total"] or 0
    total_generado = pagos_base.exclude(estado=PagoDiarista.Estados.ANULADO).aggregate(total=Sum("total_pagado"))["total"] or Decimal("0")
    total_pagado = pagos_base.filter(estado=PagoDiarista.Estados.PAGADO).aggregate(total=Sum("total_pagado"))["total"] or Decimal("0")

    filas = []
    pendiente_estimado_total = Decimal("0")
    for diarista in diaristas.order_by("empresa__nombre", "sucursal__nombre", "apellidos", "nombres"):
        asistencias = asistencias_base.filter(diarista=diarista)
        pagos = pagos_base.filter(diarista=diarista).exclude(estado=PagoDiarista.Estados.ANULADO)
        trabajadas = asistencias.filter(estado=AsistenciaDiarista.Estados.TRABAJADO).count()
        ausencias_d = asistencias.filter(estado=AsistenciaDiarista.Estados.AUSENTE).count()
        incompletas_d = asistencias.filter(estado=AsistenciaDiarista.Estados.INCOMPLETO).count()
        pendientes_d = asistencias.filter(
            estado=AsistenciaDiarista.Estados.TRABAJADO,
            pago_estado__in=[AsistenciaDiarista.EstadosPago.PENDIENTE, AsistenciaDiarista.EstadosPago.REABIERTO],
            pago__isnull=True,
        ).count()
        minutos_d = asistencias.aggregate(total=Sum("minutos_trabajados"))["total"] or 0
        generado_d = pagos.aggregate(total=Sum("total_pagado"))["total"] or Decimal("0")
        pendiente_estimado = Decimal(pendientes_d) * (diarista.monto_diario_acordado or Decimal("0"))
        pendiente_estimado_total += pendiente_estimado
        filas.append({
            "diarista": diarista,
            "trabajadas": trabajadas,
            "ausencias": ausencias_d,
            "incompletas": incompletas_d,
            "pendientes_pago": pendientes_d,
            "horas": _horas_desde_minutos(minutos_d),
            "generado": generado_d,
            "pendiente_estimado": pendiente_estimado,
        })

    if export == "csv":
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="reporte_diaristas_{desde}_{hasta}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response, delimiter=";")
        writer.writerow(["Diarista", "Documento", "Empresa", "Sucursal", "Estado", "Trabajadas", "Ausencias", "Incompletas", "Pendientes de pago", "Horas", "Generado", "Pendiente estimado"])
        for fila in filas:
            d = fila["diarista"]
            writer.writerow([
                d.nombre_completo,
                d.cedula,
                d.empresa.nombre if d.empresa else "",
                d.sucursal.nombre if d.sucursal else "",
                d.get_estado_display(),
                fila["trabajadas"],
                fila["ausencias"],
                fila["incompletas"],
                fila["pendientes_pago"],
                fila["horas"],
                fila["generado"],
                fila["pendiente_estimado"],
            ])
        return response

    if admin_master:
        empresas = Empresa.objects.filter(activo=True).order_by("nombre")
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
        if empresa_id:
            sucursales = sucursales.filter(empresa_id=empresa_id)
    else:
        empresas = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()
        sucursales = Sucursal.objects.filter(empresa=empresa_usuario, activo=True).order_by("nombre") if empresa_usuario else Sucursal.objects.none()

    return render(request, "core/diaristas_reporte.html", {
        "filas": filas,
        "desde": desde,
        "hasta": hasta,
        "empresa_id": empresa_id,
        "sucursal_id": sucursal_id,
        "estado_sel": estado,
        "q": q,
        "empresas": empresas,
        "sucursales": sucursales,
        "estados": Diarista.Estados.choices,
        "total_diaristas": total_diaristas,
        "jornadas_trabajadas": jornadas_trabajadas,
        "ausencias": ausencias,
        "incompletas": incompletas,
        "pendientes_pago": pendientes_pago,
        "horas_total": _horas_desde_minutos(minutos_total),
        "total_generado": total_generado,
        "total_pagado": total_pagado,
        "pendiente_estimado_total": pendiente_estimado_total,
    })


@login_required
def diarista_pagos(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes ver pagos de diaristas de otra empresa.")
        return redirect("diaristas_lista")

    estado = request.GET.get("estado", "").strip()
    desde = request.GET.get("desde", "").strip()
    hasta = request.GET.get("hasta", "").strip()

    pagos = diarista.pagos.select_related("empresa", "sucursal").all()
    if estado:
        pagos = pagos.filter(estado=estado)
    if desde:
        pagos = pagos.filter(fecha_pago__gte=desde)
    if hasta:
        pagos = pagos.filter(fecha_pago__lte=hasta)

    pagos_validos = diarista.pagos.exclude(estado=PagoDiarista.Estados.ANULADO)
    total_pagado = pagos_validos.aggregate(total=Sum("total_pagado"))["total"] or Decimal("0")
    pendientes_jornadas = diarista.asistencias.filter(
        estado=AsistenciaDiarista.Estados.TRABAJADO,
        pago_estado__in=[AsistenciaDiarista.EstadosPago.PENDIENTE, AsistenciaDiarista.EstadosPago.REABIERTO],
        pago__isnull=True,
    ).count()

    return render(request, "core/diarista_pagos.html", {
        "diarista": diarista,
        "pagos": pagos.order_by("-fecha_pago", "-creado_en"),
        "estados": PagoDiarista.Estados.choices,
        "estado_sel": estado,
        "desde": desde,
        "hasta": hasta,
        "total_pagado": total_pagado,
        "pagos_generados": pagos_validos.count(),
        "pendientes_jornadas": pendientes_jornadas,
    })


@login_required
def diarista_pago_nuevo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes generar pagos para otra empresa.")
        return redirect("diaristas_lista")

    jornadas_preview = AsistenciaDiarista.objects.none()
    dias_preview = Decimal("0")
    total_preview = Decimal("0")

    if request.method == "POST":
        form = PagoDiaristaForm(request.POST, diarista=diarista)
        if form.is_valid():
            desde = form.cleaned_data["periodo_desde"]
            hasta = form.cleaned_data["periodo_hasta"]
            jornadas_preview = _jornadas_pagables_diarista(diarista, desde, hasta)
            dias_preview = Decimal(jornadas_preview.count())
            if dias_preview <= 0:
                form.add_error(None, "No hay jornadas trabajadas pendientes de pago en el periodo seleccionado.")
            else:
                with transaction.atomic():
                    jornadas = _jornadas_pagables_diarista(diarista, desde, hasta).select_for_update()
                    dias_calculados = Decimal(jornadas.count())
                    if dias_calculados <= 0:
                        messages.error(request, "Las jornadas seleccionadas ya fueron tomadas por otro pago.")
                        return redirect("diarista_pagos", pk=diarista.pk)

                    pago = form.save(commit=False)
                    pago.diarista = diarista
                    pago.empresa = diarista.empresa
                    pago.sucursal = diarista.sucursal
                    pago.dias_calculados = dias_calculados
                    pago.generado_por = request.user
                    pago.estado = PagoDiarista.Estados.PAGADO if form.cleaned_data.get("marcar_pagado") else PagoDiarista.Estados.GENERADO
                    pago.calcular_total()
                    if pago.total_pagado < 0:
                        form.add_error("descuentos", "Los descuentos no pueden superar el subtotal más adicionales.")
                    else:
                        pago.save()
                        _asignar_numero_pago_diarista(pago)
                        nuevo_estado_jornada = AsistenciaDiarista.EstadosPago.PAGADO if pago.estado == PagoDiarista.Estados.PAGADO else AsistenciaDiarista.EstadosPago.INCLUIDO
                        jornadas.update(pago=pago, pago_estado=nuevo_estado_jornada)
                        registrar_historial(request, "Diaristas", "Pago", f"Se generó pago {pago.numero_comprobante} para diarista {diarista.nombre_completo}.")
                        messages.success(request, "Pago de diarista generado correctamente.")
                        return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)
    else:
        form = PagoDiaristaForm(diarista=diarista, initial={"fecha_pago": timezone.localdate()})


    return render(request, "core/diarista_pago_form.html", {
        "form": form,
        "diarista": diarista,
        "jornadas_preview": jornadas_preview,
        "dias_preview": dias_preview,
        "total_preview": total_preview,
    })


@login_required
def diarista_pago_detalle(request, pk, pago_id):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes ver pagos de otra empresa.")
        return redirect("diaristas_lista")
    pago = get_object_or_404(PagoDiarista.objects.select_related("empresa", "sucursal", "diarista", "generado_por", "anulado_por"), pk=pago_id, diarista=diarista)
    jornadas = pago.jornadas.order_by("fecha")

    return render(request, "core/diarista_pago_detalle.html", {
        "diarista": diarista,
        "pago": pago,
        "jornadas": jornadas,
    })


@login_required
def diarista_pago_pdf(request, pk, pago_id):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_ver")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes generar comprobantes de otra empresa.")
        return redirect("diaristas_lista")

    pago = get_object_or_404(
        PagoDiarista.objects.select_related("empresa", "sucursal", "diarista", "generado_por", "anulado_por"),
        pk=pago_id,
        diarista=diarista,
    )
    _asignar_numero_pago_diarista(pago)
    jornadas = list(pago.jornadas.order_by("fecha"))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    normal = ParagraphStyle(
        name="DiaristaPagoNormal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#111827"),
    )
    nota = ParagraphStyle(
        name="DiaristaPagoNota",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#475569"),
    )

    elementos = []
    empresa_pdf = pago.empresa or diarista.empresa
    elementos += construir_encabezado_empresa_pdf(empresa_pdf, "COMPROBANTE DE PAGO DIARISTA")
    elementos.append(Spacer(1, 8))

    datos = [
        ["Comprobante", pago.numero_comprobante or "-"],
        ["Diarista", diarista.nombre_completo],
        ["Documento", diarista.cedula],
        ["Empresa", getattr(empresa_pdf, "nombre", "-") or "-"],
        ["Sucursal", diarista.sucursal.nombre if diarista.sucursal else "-"],
        ["Sector", diarista.sector or "-"],
        ["Función / tarea", diarista.funcion_temporal or "-"],
        ["Periodo liquidado", f"{pago.periodo_desde:%d/%m/%Y} al {pago.periodo_hasta:%d/%m/%Y}"],
        ["Fecha de pago", pago.fecha_pago.strftime("%d/%m/%Y") if pago.fecha_pago else "-"],
        ["Estado", pago.get_estado_display()],
    ]

    tabla_datos = Table(datos, colWidths=[48 * mm, 112 * mm])
    tabla_datos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 12))

    detalle = [["Fecha", "Entrada", "Salida", "Horas", "Atraso", "Estado"]]
    for jornada in jornadas:
        detalle.append([
            jornada.fecha.strftime("%d/%m/%Y"),
            timezone.localtime(jornada.hora_entrada).strftime("%H:%M") if jornada.hora_entrada else "-",
            timezone.localtime(jornada.hora_salida).strftime("%H:%M") if jornada.hora_salida else "-",
            jornada.horas_trabajadas_display,
            f"{jornada.minutos_atraso or 0} min",
            jornada.get_estado_display(),
        ])

    if len(detalle) == 1:
        detalle.append(["-", "-", "-", "-", "-", "Sin jornadas asociadas"])

    tabla_jornadas = Table(detalle, colWidths=[25 * mm, 25 * mm, 25 * mm, 28 * mm, 25 * mm, 32 * mm], repeatRows=1)
    tabla_jornadas.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 1), (4, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(Paragraph("<b>Jornadas incluidas</b>", normal))
    elementos.append(Spacer(1, 5))
    elementos.append(tabla_jornadas)
    elementos.append(Spacer(1, 12))

    liquidacion = [
        ["Concepto", "Monto"],
        [f"Días calculados ({pago.dias_calculados}) x monto diario", _gs(pago.subtotal)],
        ["Adicionales" if not pago.concepto_adicional else f"Adicionales - {pago.concepto_adicional}", _gs(pago.adicionales)],
        ["Descuentos", f"- {_gs(pago.descuentos)}"],
        ["TOTAL A COBRAR", _gs(pago.total_pagado)],
    ]
    tabla_liquidacion = Table(liquidacion, colWidths=[110 * mm, 50 * mm])
    tabla_liquidacion.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dcfce7")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#166534")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_liquidacion)

    if pago.motivo_ajuste:
        elementos.append(Spacer(1, 8))
        elementos.append(Paragraph(f"<b>Motivo de ajuste:</b> {pago.motivo_ajuste}", nota))
    if pago.observacion:
        elementos.append(Spacer(1, 5))
        elementos.append(Paragraph(f"<b>Observación:</b> {pago.observacion}", nota))
    if pago.estado == PagoDiarista.Estados.ANULADO:
        elementos.append(Spacer(1, 5))
        elementos.append(Paragraph(f"<b>Pago anulado:</b> {pago.motivo_anulacion or '-'}", nota))

    elementos.append(Spacer(1, 24))
    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma responsable", "Firma diarista"],
    ], colWidths=[80 * mm, 80 * mm])
    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="PAGO_DIARISTA",
        documento_id=pago.id,
        funcionario=None,
        titulo="Comprobante de Pago Diarista",
    )
    agregar_texto_legal_empresa_pdf(elementos, empresa_pdf)

    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="pago_diarista_{diarista.cedula}_{pago.id}.pdf"'
    response.write(pdf)
    return response


@login_required
def diarista_pago_marcar_pagado(request, pk, pago_id):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes modificar pagos de otra empresa.")
        return redirect("diaristas_lista")
    pago = get_object_or_404(PagoDiarista, pk=pago_id, diarista=diarista)
    if request.method != "POST":
        return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)
    if pago.estado == PagoDiarista.Estados.ANULADO:
        messages.error(request, "No puedes marcar como pagado un pago anulado.")
        return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)

    pago.estado = PagoDiarista.Estados.PAGADO
    pago.save(update_fields=["estado", "actualizado_en"])
    pago.jornadas.update(pago_estado=AsistenciaDiarista.EstadosPago.PAGADO)
    registrar_historial(request, "Diaristas", "Pago", f"Se marcó como pagado el pago {pago.numero_comprobante}.")
    messages.success(request, "Pago marcado como pagado.")
    return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)


@login_required
def diarista_pago_anular(request, pk, pago_id):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes anular pagos de otra empresa.")
        return redirect("diaristas_lista")
    pago = get_object_or_404(PagoDiarista, pk=pago_id, diarista=diarista)
    if request.method != "POST":
        return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)
    if pago.estado == PagoDiarista.Estados.ANULADO:
        messages.info(request, "Este pago ya estaba anulado.")
        return redirect("diarista_pago_detalle", pk=diarista.pk, pago_id=pago.pk)

    motivo = (request.POST.get("motivo_anulacion") or "Anulación administrativa.").strip()
    with transaction.atomic():
        pago.estado = PagoDiarista.Estados.ANULADO
        pago.motivo_anulacion = motivo
        pago.anulado_por = request.user
        pago.fecha_anulacion = timezone.now()
        pago.save(update_fields=["estado", "motivo_anulacion", "anulado_por", "fecha_anulacion", "actualizado_en"])
        pago.jornadas.update(pago=None, pago_estado=AsistenciaDiarista.EstadosPago.REABIERTO)

    registrar_historial(request, "Diaristas", "Anular pago", f"Se anuló pago {pago.numero_comprobante} de diarista {diarista.nombre_completo}.")
    messages.success(request, "Pago anulado y jornadas reabiertas.")
    return redirect("diarista_pagos", pk=diarista.pk)


@login_required
def diarista_cambiar_estado(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not _usuario_puede_operar_diarista(request, diarista):
        messages.error(request, "No puedes modificar diaristas de otra empresa.")
        return redirect("diaristas_lista")

    if request.method != "POST":
        return redirect("diarista_detalle", pk=diarista.pk)

    accion = (request.POST.get("accion") or "").strip().lower()
    hoy = timezone.localdate()
    estado_anterior = diarista.get_estado_display()
    activo_anterior = diarista.activo
    pagos_validos = diarista.pagos.exclude(estado=PagoDiarista.Estados.ANULADO).count()
    jornadas = diarista.asistencias.count()

    if accion == "finalizar":
        diarista.estado = Diarista.Estados.FINALIZADO
        diarista.activo = False
        if not diarista.fecha_fin or diarista.fecha_fin > hoy:
            diarista.fecha_fin = hoy
        mensaje = "Diarista finalizado correctamente. Sus asistencias y pagos históricos se conservan."
        accion_historial = "Finalizar"
    elif accion == "cancelar":
        diarista.estado = Diarista.Estados.CANCELADO
        diarista.activo = False
        if not diarista.fecha_fin or diarista.fecha_fin > hoy:
            diarista.fecha_fin = hoy
        mensaje = "Diarista cancelado correctamente. No se eliminaron asistencias ni pagos."
        accion_historial = "Cancelar"
    elif accion == "activar":
        diarista.estado = Diarista.Estados.ACTIVO
        diarista.activo = True
        if diarista.fecha_fin and diarista.fecha_fin < hoy:
            diarista.fecha_fin = hoy
        mensaje = "Diarista activado correctamente."
        accion_historial = "Activar"
    elif accion == "pendiente":
        diarista.estado = Diarista.Estados.PENDIENTE
        diarista.activo = True
        mensaje = "Diarista marcado como pendiente."
        accion_historial = "Pendiente"
    else:
        messages.error(request, "Acción de estado no válida.")
        return redirect("diarista_detalle", pk=diarista.pk)

    diarista.save(update_fields=["estado", "activo", "fecha_fin", "actualizado_en"])
    registrar_historial(
        request,
        "Diaristas",
        accion_historial,
        f"Diarista {diarista.nombre_completo} cambió de {estado_anterior} / activo={activo_anterior} "
        f"a {diarista.get_estado_display()} / activo={diarista.activo}. "
        f"Jornadas conservadas: {jornadas}. Pagos válidos conservados: {pagos_validos}."
    )
    messages.success(request, mensaje)
    return redirect("diarista_detalle", pk=diarista.pk)


@login_required
def diarista_nuevo(request):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = DiaristaForm(request.POST, empresa_activa=None if admin_master else empresa_usuario)
        if form.is_valid():
            diarista = form.save(commit=False)
            if not admin_master:
                if not empresa_usuario or diarista.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear diaristas fuera de tu empresa.")
                    return redirect("diaristas_lista")
            if diarista.sucursal and diarista.sucursal.empresa != diarista.empresa:
                messages.error(request, "La sucursal no pertenece a la empresa seleccionada.")
                return redirect("diaristas_lista")
            if diarista.turno and diarista.turno.empresa != diarista.empresa:
                messages.error(request, "El turno no pertenece a la empresa seleccionada.")
                return redirect("diaristas_lista")
            diarista.creado_por = request.user
            diarista.save()
            registrar_historial(
                request,
                "Diaristas",
                "Crear",
                f"Se creó diarista {diarista.nombre_completo} (CI: {diarista.cedula}) para {diarista.empresa.nombre}."
            )
            messages.success(request, "Diarista creado correctamente.")
            return redirect("diarista_detalle", pk=diarista.pk)
    else:
        form = DiaristaForm(empresa_activa=None if admin_master else empresa_usuario)

    return render(request, "core/diarista_form.html", {
        "form": form,
        "titulo_form": "Nuevo diarista",
        "boton_texto": "Guardar diarista",
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def diarista_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "diaristas", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    diarista = get_object_or_404(Diarista.objects.select_related("empresa", "sucursal", "turno"), pk=pk)
    if not admin_master:
        if not empresa_usuario or diarista.empresa != empresa_usuario:
            messages.error(request, "No puedes editar diaristas de otra empresa.")
            return redirect("diaristas_lista")

    if request.method == "POST":
        form = DiaristaForm(request.POST, instance=diarista, empresa_activa=None if admin_master else empresa_usuario)
        if form.is_valid():
            diarista_editado = form.save(commit=False)
            if not admin_master:
                if not empresa_usuario or diarista_editado.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover diaristas a otra empresa.")
                    return redirect("diaristas_lista")
            if diarista_editado.sucursal and diarista_editado.sucursal.empresa != diarista_editado.empresa:
                messages.error(request, "La sucursal no pertenece a la empresa seleccionada.")
                return redirect("diarista_editar", pk=diarista.pk)
            if diarista_editado.turno and diarista_editado.turno.empresa != diarista_editado.empresa:
                messages.error(request, "El turno no pertenece a la empresa seleccionada.")
                return redirect("diarista_editar", pk=diarista.pk)
            diarista_editado.save()
            registrar_historial(
                request,
                "Diaristas",
                "Editar",
                f"Se editó diarista {diarista_editado.nombre_completo} (CI: {diarista_editado.cedula})."
            )
            messages.success(request, "Diarista actualizado correctamente.")
            return redirect("diarista_detalle", pk=diarista_editado.pk)
    else:
        form = DiaristaForm(instance=diarista, empresa_activa=None if admin_master else empresa_usuario)

    return render(request, "core/diarista_form.html", {
        "form": form,
        "titulo_form": f"Editar diarista: {diarista.nombre_completo}",
        "boton_texto": "Guardar cambios",
        "diarista": diarista,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def deudas_lista(request):
    permiso = validar_permiso_o_redirigir(request, "deudas", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()
    funcionario_id = request.GET.get("funcionario", "").strip()
    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    deudas = Deuda.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa").all()

    if not admin_master:
        if empresa_usuario:
            deudas = deudas.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            deudas = deudas.none()

    if q:
        deudas = deudas.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q) |
            Q(descripcion__icontains=q) |
            Q(tipo__icontains=q)
        )

    if funcionario_id:
        deudas = deudas.filter(funcionario_id=funcionario_id)

    if admin_master:
        funcionarios = Funcionario.objects.filter(activo=True).order_by("apellido", "nombre")
    else:
        funcionarios = Funcionario.objects.filter(
            activo=True,
            sucursal_rel__empresa=empresa_usuario
        ).order_by("apellido", "nombre") if empresa_usuario else Funcionario.objects.none()

    return render(request, "core/deudas_lista.html", {
        "deudas": deudas.order_by("-fecha", "-creado_en"),
        "funcionarios": funcionarios,
        "funcionario_id": funcionario_id,
        "q": q,
        "q": q,
    })


@login_required
def deuda_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "deudas", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = DeudaForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            deuda = form.save(commit=False)

            if not admin_master:
                if not deuda.funcionario.sucursal_rel or deuda.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear deudas para otra empresa.")
                    return redirect("deudas_lista")

            deuda.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Deudas",
                "Crear",
                f"Se creó deuda para {deuda.funcionario.nombre_completo} por {deuda.saldo_pendiente}."
            )
            messages.success(request, "Deuda creada correctamente.")
            return redirect("deudas_lista")
    else:
        form = DeudaForm()
        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/deuda_form.html", {
        "form": form,
        "titulo_form": "Nueva deuda",
        "boton_texto": "Guardar deuda",
    })


@login_required
def deuda_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "deudas", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    deuda = get_object_or_404(Deuda, pk=pk)

    if not admin_master:
        if not deuda.funcionario.sucursal_rel or deuda.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes editar deudas de otra empresa.")
            return redirect("deudas_lista")

    if request.method == "POST":
        form = DeudaForm(request.POST, instance=deuda)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            deuda_editada = form.save(commit=False)

            if not admin_master:
                if not deuda_editada.funcionario.sucursal_rel or deuda_editada.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover deudas a otra empresa.")
                    return redirect("deudas_lista")

            deuda_editada.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Deudas",
                "Editar",
                f"Se editó deuda de {deuda_editada.funcionario.nombre_completo}."
            )
            messages.success(request, "Deuda actualizada correctamente.")
            return redirect("deudas_lista")
    else:
        form = DeudaForm(instance=deuda)
        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/deuda_form.html", {
        "form": form,
        "titulo_form": f"Editar deuda: {deuda.funcionario.nombre_completo}",
        "boton_texto": "Guardar cambios",
        "deuda": deuda,
    })


@login_required
def deuda_toggle_activa(request, pk):
    permiso = validar_permiso_o_redirigir(request, "deudas", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    deuda = get_object_or_404(Deuda, pk=pk)

    if not admin_master:
        if not empresa_usuario or not deuda.funcionario.sucursal_rel or deuda.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes cambiar deudas de otra empresa.")
            return redirect("deudas_lista")

    deuda.activa = not deuda.activa
    deuda.save()

    estado = "activada" if deuda.activa else "inactivada"
    registrar_historial(
        request,
        "Deudas",
        "Cambio de estado",
        f"Deuda de {deuda.funcionario.nombre_completo} fue {estado}."
    )
    messages.success(request, f"Deuda {estado} correctamente.")
    return redirect("deudas_lista")


@login_required
def funcionarios_lista(request):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()
    empresa_id = request.GET.get("empresa", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionarios = Funcionario.objects.select_related(
        "turno",
        "sucursal_rel",
        "sucursal_rel__empresa"
    ).all()

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()

    if q:
        funcionarios = funcionarios.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            _documento_busqueda_q("cedula", "tipo_documento", q) |
            Q(cargo__icontains=q) |
            Q(sector__icontains=q) |
            Q(sucursal__icontains=q) |
            Q(sucursal_rel__nombre__icontains=q) |
            Q(sucursal_rel__empresa__nombre__icontains=q) |
            Q(turno__nombre__icontains=q)
        )

    if admin_master:
        if empresa_id:
            funcionarios = funcionarios.filter(sucursal_rel__empresa_id=empresa_id)
    else:
        if empresa_usuario:
            empresa_id = str(empresa_usuario.id)

    if sucursal_id:
        funcionarios = funcionarios.filter(sucursal_rel_id=sucursal_id)

    if admin_master:
        empresas = Empresa.objects.filter(activo=True).order_by("nombre")
        sucursales = Sucursal.objects.filter(activo=True).order_by("nombre")
        if empresa_id:
            sucursales = sucursales.filter(empresa_id=empresa_id)
    else:
        empresas = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()

    context = {
        "funcionarios": funcionarios.order_by("apellido", "nombre"),
        "q": q,
        "empresas": empresas,
        "sucursales": sucursales,
        "empresa_id": empresa_id,
        "sucursal_id": sucursal_id,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    }
    return render(request, "core/funcionarios_lista.html", context)

@login_required
def funcionario_detalle(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionario = get_object_or_404(
        Funcionario.objects.select_related(
            "turno",
            "sucursal_rel",
            "sucursal_rel__empresa"
        ),
        pk=pk
    )

    if not admin_master:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes ver funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    hoy = timezone.localdate()
    anio_actual = hoy.year
    mes_actual = hoy.month

    asistencias = funcionario.asistencias.all().order_by("-fecha")[:30]
    permisos = funcionario.permisos_licencias.all().order_by("-fecha_desde")[:10]
    vacaciones = funcionario.vacaciones.all().order_by("-fecha_desde")[:10]
    movimientos_horas = funcionario.movimientos_banco_horas.all().order_by("-fecha")[:20]

    documentos = funcionario.documentos_personales.filter(activo=True)
    historial_laboral = funcionario.historial_laboral.all()
    conductas = funcionario.conductas.all()
    historial_salarial = funcionario.historial_salarial.all()

    asistencias_anio = funcionario.asistencias.filter(
        fecha__year=anio_actual,
        hora_entrada__isnull=False
    ).count()

    atrasos_anio = funcionario.asistencias.filter(
        fecha__year=anio_actual,
        llego_tarde=True
    ).count()

    permisos_anio = funcionario.permisos_licencias.filter(
        fecha_desde__year=anio_actual
    ).count()

    vacaciones_anio = funcionario.vacaciones.filter(
        fecha_desde__year=anio_actual
    ).count()

    resumen_icl_mes = calcular_icl_funcionario_mes(
        funcionario=funcionario,
        mes=mes_actual,
        anio=anio_actual
    )

    documento_form = DocumentoFuncionarioForm()
    laboral_form = HistorialLaboralFuncionarioForm()
    conducta_form = ConductaFuncionarioForm()
    salarial_form = HistorialSalarialFuncionarioForm(initial={
    "salario_anterior": f"{int(funcionario.salario_base):,}".replace(",", "."),
    "salario_nuevo": f"{int(funcionario.salario_base):,}".replace(",", "."),
    "bono_anterior": f"{int(funcionario.bono):,}".replace(",", "."),
    "bono_nuevo": f"{int(funcionario.bono):,}".replace(",", "."),
})

    return render(request, "core/funcionario_detalle.html", {
        "funcionario": funcionario,
        "asistencias": asistencias,
        "permisos": permisos,
        "vacaciones": vacaciones,
        "movimientos_horas": movimientos_horas,

        "documentos": documentos,
        "historial_laboral": historial_laboral,
        "conductas": conductas,
        "historial_salarial": historial_salarial,

        "asistencias_anio": asistencias_anio,
        "atrasos_anio": atrasos_anio,
        "permisos_anio": permisos_anio,
        "vacaciones_anio": vacaciones_anio,
        "resumen_icl_mes": resumen_icl_mes,

        "documento_form": documento_form,
        "laboral_form": laboral_form,
        "conducta_form": conducta_form,
        "salarial_form": salarial_form,

        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })

@login_required
def funcionario_documento_agregar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"), pk=pk)

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes agregar documentos a funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    if request.method == "POST":
        form = DocumentoFuncionarioForm(request.POST, request.FILES)
        if form.is_valid():
            documento = form.save(commit=False)
            documento.funcionario = funcionario
            documento.save()

            registrar_historial(
                request,
                "Funcionarios",
                "Documento digital",
                f"Se agregó documento {documento.titulo} a {funcionario.nombre_completo}."
            )

            messages.success(request, "Documento agregado correctamente.")

    return redirect("funcionario_detalle", pk=funcionario.id)


@login_required
def funcionario_historial_laboral_agregar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"), pk=pk)

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes agregar historial laboral a funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    if request.method == "POST":
        form = HistorialLaboralFuncionarioForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.funcionario = funcionario
            item.save()

            registrar_historial(
                request,
                "Funcionarios",
                "Historial laboral",
                f"Se agregó historial laboral a {funcionario.nombre_completo}: {item.titulo}."
            )

            messages.success(request, "Historial laboral agregado correctamente.")

    return redirect("funcionario_detalle", pk=funcionario.id)


@login_required
def funcionario_conducta_agregar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"), pk=pk)

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes agregar conducta a funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    if request.method == "POST":
        form = ConductaFuncionarioForm(request.POST, request.FILES)
        if form.is_valid():
            item = form.save(commit=False)
            item.funcionario = funcionario
            item.save()

            registrar_historial(
                request,
                "Funcionarios",
                "Historial de conducta",
                f"Se agregó {item.get_tipo_display()} a {funcionario.nombre_completo}: {item.titulo}."
            )

            messages.success(request, "Historial de conducta agregado correctamente.")

    return redirect("funcionario_detalle", pk=funcionario.id)


@login_required
def funcionario_historial_salarial_agregar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"), pk=pk)

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes agregar historial salarial a funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    if request.method == "POST":
        form = HistorialSalarialFuncionarioForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.funcionario = funcionario
            item.save()

            funcionario.salario_base = item.salario_nuevo
            funcionario.bono = item.bono_nuevo
            funcionario.save(update_fields=["salario_base", "bono", "actualizado_en"])

            registrar_historial(
                request,
                "Funcionarios",
                "Historial salarial",
                f"Se actualizó historial salarial de {funcionario.nombre_completo}."
            )

            messages.success(request, "Historial salarial agregado y salario actualizado correctamente.")

    return redirect("funcionario_detalle", pk=funcionario.id)

@login_required
def funcionario_nuevo(request):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = FuncionarioForm(request.POST, request.FILES, request=request, empresa_activa=None if admin_master else empresa_usuario)

        if form.is_valid():
            funcionario = form.save(commit=False)

            if not admin_master:
                if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear funcionarios fuera de tu empresa.")
                    return redirect("funcionarios_lista")

                if funcionario.turno and funcionario.turno.empresa != empresa_usuario:
                    messages.error(request, "No puedes asignar turnos de otra empresa.")
                    return redirect("funcionarios_lista")

            funcionario.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Funcionarios",
                "Crear",
                f"Se creó el funcionario {funcionario.nombre_completo} ({funcionario.documento_compacto})."
            )
            messages.success(request, "Funcionario creado correctamente.")
            return redirect("funcionarios_lista")
    else:
        form = FuncionarioForm(request=request, empresa_activa=None if admin_master else empresa_usuario)

    return render(request, "core/funcionario_form.html", {
        "form": form,
        "titulo_form": "Nuevo funcionario",
        "boton_texto": "Guardar funcionario",
    })


@login_required
def funcionario_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionario = get_object_or_404(
        Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa", "turno"),
        pk=pk
    )

    if not admin_master:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes editar funcionarios de otra empresa.")
            return redirect("funcionarios_lista")

    if request.method == "POST":
        form = FuncionarioForm(request.POST, request.FILES, instance=funcionario, request=request, empresa_activa=None if admin_master else empresa_usuario)

        if form.is_valid():
            funcionario_editado = form.save(commit=False)

            if not admin_master:
                if not funcionario_editado.sucursal_rel or funcionario_editado.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover un funcionario a otra empresa.")
                    return redirect("funcionarios_lista")

                if funcionario_editado.turno and funcionario_editado.turno.empresa != empresa_usuario:
                    messages.error(request, "No puedes asignar turnos de otra empresa.")
                    return redirect("funcionarios_lista")

            funcionario_editado.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Funcionarios",
                "Editar",
                f"Se editó el funcionario {funcionario_editado.nombre_completo} ({funcionario_editado.documento_compacto})."
            )
            messages.success(request, "Funcionario actualizado correctamente.")
            return redirect("funcionarios_lista")
    else:
        form = FuncionarioForm(instance=funcionario, request=request, empresa_activa=None if admin_master else empresa_usuario)

    return render(request, "core/funcionario_form.html", {
        "form": form,
        "titulo_form": f"Editar funcionario: {funcionario.nombre_completo}",
        "boton_texto": "Guardar cambios",
        "funcionario": funcionario,
    })


@login_required
def funcionario_toggle_activo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "funcionarios", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(
        Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"),
        pk=pk
    )

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar banco de horas de otra empresa.")
            return redirect("banco_horas_lista")
        
    funcionario.activo = not funcionario.activo
    funcionario.save()

    estado = "activado" if funcionario.activo else "inactivado"
    registrar_historial(
        request,
        "Funcionarios",
        "Cambio de estado",
        f"Funcionario {funcionario.nombre_completo} fue {estado}."
    )
    messages.success(request, f"Funcionario {estado} correctamente.")
    return redirect("funcionarios_lista")


@login_required
def turnos_lista(request):
    permiso = validar_permiso_o_redirigir(request, "turnos", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    turnos = Turno.objects.select_related("empresa").all()

    if not admin_master:
        if empresa_usuario:
            turnos = turnos.filter(empresa=empresa_usuario)
        else:
            turnos = turnos.none()

    if q:
        turnos = turnos.filter(nombre__icontains=q)

    return render(request, "core/turnos_lista.html", {
        "turnos": turnos.order_by("nombre"),
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def turno_nuevo(request):
    permiso = validar_permiso_o_redirigir(request, "turnos", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = TurnoForm(request.POST)

        if not admin_master:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()

        if form.is_valid():
            turno = form.save(commit=False)

            if admin_master:
                if not turno.empresa:
                    messages.error(request, "Debes seleccionar una empresa para el turno.")
                    return render(request, "core/turno_form.html", {
                        "form": form,
                        "titulo_form": "Nuevo turno",
                        "boton_texto": "Guardar turno",
                    })
            else:
                if not empresa_usuario:
                    messages.error(request, "Tu usuario no tiene empresa asignada.")
                    return redirect("turnos_lista")
                turno.empresa = empresa_usuario

            turno.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Turnos",
                "Crear",
                f"Se creó el turno {turno.nombre} para la empresa {turno.empresa.nombre if turno.empresa else 'Sin empresa'}."
            )
            messages.success(request, "Turno creado correctamente.")
            return redirect("turnos_lista")
    else:
        form = TurnoForm()

        if admin_master:
            form.fields["empresa"].queryset = Empresa.objects.filter(activo=True).order_by("nombre")
        else:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()
            if empresa_usuario:
                form.fields["empresa"].initial = empresa_usuario

    return render(request, "core/turno_form.html", {
        "form": form,
        "titulo_form": "Nuevo turno",
        "boton_texto": "Guardar turno",
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def turno_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "turnos", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    turno = get_object_or_404(Turno.objects.select_related("empresa"), pk=pk)

    if not admin_master:
        if turno.empresa != empresa_usuario:
            messages.error(request, "No puedes editar turnos de otra empresa.")
            return redirect("turnos_lista")

    if request.method == "POST":
        form = TurnoForm(request.POST, instance=turno)

        if not admin_master:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()

        if form.is_valid():
            turno_editado = form.save(commit=False)

            if admin_master:
                if not turno_editado.empresa:
                    messages.error(request, "Debes seleccionar una empresa para el turno.")
                    return render(request, "core/turno_form.html", {
                        "form": form,
                        "titulo_form": "Editar turno",
                        "boton_texto": "Guardar cambios",
                        "turno": turno,
                    })
            else:
                turno_editado.empresa = empresa_usuario

            turno_editado.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Turnos",
                "Editar",
                f"Se editó el turno {turno_editado.nombre}."
            )
            messages.success(request, "Turno actualizado correctamente.")
            return redirect("turnos_lista")
    else:
        form = TurnoForm(instance=turno)

        if admin_master:
            form.fields["empresa"].queryset = Empresa.objects.filter(activo=True).order_by("nombre")
        else:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()

    return render(request, "core/turno_form.html", {
        "form": form,
        "titulo_form": "Editar turno",
        "boton_texto": "Guardar cambios",
        "turno": turno,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def turno_toggle_activo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "turnos", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    turno = get_object_or_404(Turno.objects.select_related("empresa"), pk=pk)

    if not admin_master:
        if turno.empresa != empresa_usuario:
            messages.error(request, "No puedes cambiar turnos de otra empresa.")
            return redirect("turnos_lista")

    turno.activo = not turno.activo
    turno.save(update_fields=["activo"])

    registrar_historial(
        request,
        "Turnos",
        "Cambio de estado",
        f"Se cambió el estado del turno {turno.nombre} a {'Activo' if turno.activo else 'Inactivo'}."
    )
    messages.success(request, "Estado del turno actualizado correctamente.")
    return redirect("turnos_lista")

@login_required
def asistencia_marcar(request):
    permiso = validar_permiso_o_redirigir(request, "asistencia", "puede_ver")
    if permiso:
        return permiso

    fecha_filtro_str = request.GET.get("fecha", "").strip()

    try:
        fecha_filtro = datetime.strptime(fecha_filtro_str, "%Y-%m-%d").date() if fecha_filtro_str else timezone.localdate()
    except ValueError:
        fecha_filtro = timezone.localdate()

    hoy = fecha_filtro

    resultado = None

    sucursal_id = request.GET.get("sucursal", "").strip()
    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        permiso_post = validar_permiso_o_redirigir(request, "asistencia", "puede_crear")
        if permiso_post:
            return permiso_post

        form = MarcacionForm(request.POST)
        if form.is_valid():
            cedula = form.cleaned_data["cedula"].strip()

            try:
                funcionario = Funcionario.objects.select_related("turno", "sucursal_rel", "sucursal_rel__empresa").get(
                    cedula=cedula,
                    activo=True
                )
            except Funcionario.DoesNotExist:
                messages.error(request, "No se encontró un funcionario activo con esa cédula.")
                funcionario = None

            if funcionario and not admin_master:
                if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes registrar asistencia para otra empresa.")
                    funcionario = None

            if funcionario:
                if funcionario_tiene_dia_libre(funcionario, hoy):
                    messages.info(
                        request,
                        f"{funcionario.nombre_completo} tiene día libre hoy. No corresponde asistencia."
                    )
                    resultado = {
                        "tipo": "dia_libre",
                        "funcionario": funcionario,
                        "hora": timezone.localtime(),
                        "turno": funcionario.turno.nombre if funcionario.turno else "-",
                        "atraso": 0,
                        "llego_tarde": False,
                    }
                else:
                    ahora = timezone.localtime()
                    fecha_operativa = obtener_fecha_operativa_asistencia(funcionario, ahora)

                    asistencia, creada = Asistencia.objects.get_or_create(
                        funcionario=funcionario,
                        fecha=fecha_operativa
                    )

                    if not funcionario.turno:
                        messages.error(request, "El funcionario no tiene un turno asignado.")
                    else:
                        siguiente = asistencia.siguiente_marcacion

                        if siguiente == "entrada":
                            asistencia.hora_entrada = ahora
                            asistencia.calcular_atraso()

                            if asistencia.llego_tarde:
                                asistencia.observacion = f"Llegó con {asistencia.minutos_atraso} minuto(s) de atraso."
                            else:
                                asistencia.observacion = "Entrada registrada en horario."

                            asistencia.save()

                            registrar_historial(
                                request,
                                "Asistencia",
                                "Entrada",
                                f"Se registró entrada de {funcionario.nombre_completo} a las {ahora.strftime('%H:%M:%S')}."
                            )

                            resultado = {
                                "tipo": "entrada",
                                "funcionario": funcionario,
                                "hora": ahora,
                                "turno": funcionario.turno.nombre,
                                "atraso": asistencia.minutos_atraso,
                                "llego_tarde": asistencia.llego_tarde,
                            }
                            messages.success(request, "Entrada registrada correctamente.")

                        elif siguiente == "salida_almuerzo":
                            asistencia.hora_salida_almuerzo = ahora
                            asistencia.observacion = "Salida a almuerzo registrada correctamente."
                            asistencia.save()

                            registrar_historial(
                                request,
                                "Asistencia",
                                "Salida a almuerzo",
                                f"Se registró salida a almuerzo de {funcionario.nombre_completo} a las {ahora.strftime('%H:%M:%S')}."
                            )

                            resultado = {
                                "tipo": "salida_almuerzo",
                                "funcionario": funcionario,
                                "hora": ahora,
                                "turno": funcionario.turno.nombre,
                                "atraso": asistencia.minutos_atraso,
                                "llego_tarde": asistencia.llego_tarde,
                            }
                            messages.success(request, "Salida a almuerzo registrada correctamente.")

                        elif siguiente == "regreso_almuerzo":
                            asistencia.hora_regreso_almuerzo = ahora
                            if asistencia.observacion:
                                asistencia.observacion += " Regreso de almuerzo registrado correctamente."
                            else:
                                asistencia.observacion = "Regreso de almuerzo registrado correctamente."
                            asistencia.save()

                            registrar_historial(
                                request,
                                "Asistencia",
                                "Regreso de almuerzo",
                                f"Se registró regreso de almuerzo de {funcionario.nombre_completo} a las {ahora.strftime('%H:%M:%S')}."
                            )

                            resultado = {
                                "tipo": "regreso_almuerzo",
                                "funcionario": funcionario,
                                "hora": ahora,
                                "turno": funcionario.turno.nombre,
                                "atraso": asistencia.minutos_atraso,
                                "llego_tarde": asistencia.llego_tarde,
                            }
                            messages.success(request, "Regreso de almuerzo registrado correctamente.")

                        elif siguiente == "salida":
                            asistencia.hora_salida = ahora
                            if asistencia.observacion:
                                asistencia.observacion += " Salida final registrada correctamente."
                            else:
                                asistencia.observacion = "Salida final registrada correctamente."
                            asistencia.save()

                            registrar_historial(
                                request,
                                "Asistencia",
                                "Salida final",
                                f"Se registró salida final de {funcionario.nombre_completo} a las {ahora.strftime('%H:%M:%S')}."
                            )

                            resultado = {
                                "tipo": "salida",
                                "funcionario": funcionario,
                                "hora": ahora,
                                "turno": funcionario.turno.nombre,
                                "atraso": asistencia.minutos_atraso,
                                "llego_tarde": asistencia.llego_tarde,
                            }
                            messages.success(request, "Salida final registrada correctamente.")

                        else:
                            messages.warning(request, "El funcionario ya completó todas sus marcaciones del día.")
    else:
        form = MarcacionForm()

    asistencias_hoy = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "marcado_manual_por",
    ).filter(fecha=fecha_filtro)

    if not admin_master:
        if empresa_usuario:
            asistencias_hoy = asistencias_hoy.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            asistencias_hoy = asistencias_hoy.none()

    if sucursal_id:
        asistencias_hoy = asistencias_hoy.filter(funcionario__sucursal_rel_id=sucursal_id)

    if q:
        asistencias_hoy = asistencias_hoy.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q)
        )

    if admin_master:
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
    else:
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()        

    asistencias_hoy = asistencias_hoy.order_by("-hora_entrada")

    return render(request, "core/asistencia_marcar.html", {
        "form": form,
        "resultado": resultado,
        "asistencias_hoy": asistencias_hoy,
        "fecha_filtro": fecha_filtro,
        "hoy": hoy,
        "sucursales": sucursales,
        "sucursal_id": sucursal_id,
        "q": q,
    })


@login_required
def permisos_lista(request):
    permiso = validar_permiso_o_redirigir(request, "permisos", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    permisos = PermisoLicencia.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).all()

    if not admin_master:
        if empresa_usuario:
            permisos = permisos.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            permisos = permisos.none()

    if q:
        permisos = permisos.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q) |
            Q(tipo__icontains=q) |
            Q(estado__icontains=q)
        )

    return render(request, "core/permisos_lista.html", {
        "permisos": permisos,
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def permiso_nuevo(request):
    permiso_acc = validar_permiso_o_redirigir(request, "permisos", "puede_crear")
    if permiso_acc:
        return permiso_acc

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = PermisoLicenciaForm(request.POST, request.FILES)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            permiso_obj = form.save(commit=False)

            if not admin_master:
                if not permiso_obj.funcionario.sucursal_rel or permiso_obj.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear permisos para otra empresa.")
                    return redirect("permisos_lista")

            permiso_obj.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Permisos/Licencias",
                "Crear",
                f"Se creó {permiso_obj.get_tipo_display()} para {permiso_obj.funcionario.nombre_completo} del {permiso_obj.fecha_desde} al {permiso_obj.fecha_hasta}."
            )
            messages.success(request, "Permiso/licencia creado correctamente.")
            return redirect("permisos_lista")
    else:
        form = PermisoLicenciaForm()

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/permiso_form.html", {
        "form": form,
        "titulo_form": "Nuevo permiso / licencia",
        "boton_texto": "Guardar permiso",
    })


@login_required
def permiso_editar(request, pk):
    permiso_acc = validar_permiso_o_redirigir(request, "permisos", "puede_editar")
    if permiso_acc:
        return permiso_acc

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    permiso_obj = get_object_or_404(
        PermisoLicencia.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not permiso_obj.funcionario.sucursal_rel or permiso_obj.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes editar permisos de otra empresa.")
            return redirect("permisos_lista")

    if request.method == "POST":
        form = PermisoLicenciaForm(request.POST, request.FILES, instance=permiso_obj)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            permiso_editado = form.save(commit=False)

            if not admin_master:
                if not permiso_editado.funcionario.sucursal_rel or permiso_editado.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover permisos a otra empresa.")
                    return redirect("permisos_lista")

            permiso_editado.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Permisos/Licencias",
                "Editar",
                f"Se editó {permiso_editado.get_tipo_display()} de {permiso_editado.funcionario.nombre_completo}. Estado actual: {permiso_editado.get_estado_display()}."
            )
            messages.success(request, "Permiso/licencia actualizado correctamente.")
            return redirect("permisos_lista")
    else:
        form = PermisoLicenciaForm(instance=permiso_obj)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/permiso_form.html", {
        "form": form,
        "titulo_form": "Editar permiso / licencia",
        "boton_texto": "Guardar cambios",
        "permiso": permiso_obj,
    })


@login_required
def vacaciones_lista(request):
    permiso = validar_permiso_o_redirigir(request, "vacaciones", "puede_ver")
    if permiso:
        return permiso
    
    alertas_vacaciones = []

    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    vacaciones = Vacacion.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).all()

    if not admin_master:
        if empresa_usuario:
            vacaciones = vacaciones.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            vacaciones = vacaciones.none()

    if q:
        vacaciones = vacaciones.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q) |
            Q(estado__icontains=q)
        )

    if admin_master:
        funcionarios_resumen = Funcionario.objects.filter(activo=True).order_by("apellido", "nombre")
    else:
        funcionarios_resumen = Funcionario.objects.filter(
            activo=True,
            sucursal_rel__empresa=empresa_usuario
        ).order_by("apellido", "nombre") if empresa_usuario else Funcionario.objects.none()

        alertas_vacaciones = []

        for funcionario in funcionarios_resumen:
            alerta = calcular_alertas_vacaciones(funcionario)
            if alerta:
                alertas_vacaciones.append({
                    "funcionario": funcionario,
                    "alerta": alerta,
                })

    return render(request, "core/vacaciones_lista.html", {
        "vacaciones": vacaciones,
        "alertas_vacaciones": alertas_vacaciones,
        "funcionarios_resumen": funcionarios_resumen,
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def vacacion_nueva(request):
    permiso_acc = validar_permiso_o_redirigir(request, "vacaciones", "puede_crear")
    if permiso_acc:
        return permiso_acc

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = VacacionForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            vacacion = form.save(commit=False)

            if not admin_master:
                if not vacacion.funcionario.sucursal_rel or vacacion.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear vacaciones para otra empresa.")
                    return redirect("vacaciones_lista")

            vacacion.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Vacaciones",
                "Crear",
                f"Se creó vacación para {vacacion.funcionario.nombre_completo} del {vacacion.fecha_desde} al {vacacion.fecha_hasta} por {vacacion.dias_solicitados} día(s)."
            )
            messages.success(request, "Vacación registrada correctamente.")
            return redirect("vacaciones_lista")
    else:
        form = VacacionForm()

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/vacacion_form.html", {
        "form": form,
        "titulo_form": "Nueva vacación",
        "boton_texto": "Guardar vacación",
        "funcionarios_json": [
    {
        "id": f.id,
        "nombre": f.nombre_completo,
        "dias": f.saldo_vacaciones,
    }
    for f in form.fields["funcionario"].queryset
],
    })

@login_required
def vacacion_editar(request, pk):
    permiso_acc = validar_permiso_o_redirigir(request, "vacaciones", "puede_editar")
    if permiso_acc:
        return permiso_acc

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    vacacion = get_object_or_404(
        Vacacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not vacacion.funcionario.sucursal_rel or vacacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes editar vacaciones de otra empresa.")
            return redirect("vacaciones_lista")

    if request.method == "POST":
        form = VacacionForm(request.POST, instance=vacacion)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            vacacion_editada = form.save(commit=False)

            if not admin_master:
                if not vacacion_editada.funcionario.sucursal_rel or vacacion_editada.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover vacaciones a otra empresa.")
                    return redirect("vacaciones_lista")

            vacacion_editada.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Vacaciones",
                "Editar",
                f"Se editó vacación de {vacacion_editada.funcionario.nombre_completo}. Estado actual: {vacacion_editada.get_estado_display()}."
            )
            messages.success(request, "Vacación actualizada correctamente.")
            return redirect("vacaciones_lista")
    else:
        form = VacacionForm(instance=vacacion)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/vacacion_form.html", {
        "form": form,
        "titulo_form": "Editar vacación",
        "boton_texto": "Guardar cambios",
        "vacacion": vacacion,
        "funcionarios_json": [
    {
        "id": f.id,
        "nombre": f.nombre_completo,
        "dias": f.saldo_vacaciones,
    }
    for f in form.fields["funcionario"].queryset
],
    })

@login_required
def vacacion_notificacion_pdf(request, pk):
    permiso = validar_permiso_o_redirigir(request, "vacaciones", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    vacacion = get_object_or_404(
        Vacacion.objects.select_related(
            "funcionario",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa"
        ),
        pk=pk
    )

    if not admin_master:
        if not vacacion.funcionario.sucursal_rel or vacacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes generar notificación de otra empresa.")
            return redirect("vacaciones_lista")

    funcionario = vacacion.funcionario
    config = ConfiguracionGeneral.obtener()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=20 * mm,
        bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TituloVacaciones",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=12,
    ))

    styles.add(ParagraphStyle(
        name="TextoVacaciones",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=16,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#111827"),
    ))

    elementos = []

    empresa_nombre = funcionario.empresa_mostrar
    sucursal_nombre = funcionario.sucursal_mostrar
    fecha_emision = vacacion.fecha_notificacion or (vacacion.fecha_desde - timezone.timedelta(days=15))

    empresa_pdf = obtener_empresa_documento(funcionario=funcionario)
    elementos += construir_encabezado_empresa_pdf(empresa_pdf, "NOTIFICACIÓN DE VACACIONES")
    elementos.append(Spacer(1, 10))

    datos = [
        ["Empresa", empresa_nombre],
        ["Sucursal", sucursal_nombre],
        ["Fecha de emisión", fecha_emision.strftime("%d/%m/%Y")],
        ["Funcionario", funcionario.nombre_completo],
        ["Documento", funcionario.documento_compacto],
        ["Cargo", funcionario.cargo or "-"],
        ["Fecha desde", vacacion.fecha_desde.strftime("%d/%m/%Y")],
        ["Fecha hasta", vacacion.fecha_hasta.strftime("%d/%m/%Y")],
        ["Días otorgados", str(vacacion.dias_solicitados)],
    ]

    tabla = Table(datos, colWidths=[55 * mm, 105 * mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla)
    elementos.append(Spacer(1, 16))

    texto = f"""
    Por medio de la presente, se comunica formalmente al trabajador <b>{funcionario.nombre_completo}</b>,
    identificado como <b>{funcionario.documento_compacto}</b>, que hará uso de sus vacaciones anuales remuneradas
    desde el día <b>{vacacion.fecha_desde.strftime("%d/%m/%Y")}</b> hasta el día
    <b>{vacacion.fecha_hasta.strftime("%d/%m/%Y")}</b>, por un total de
    <b>{vacacion.dias_solicitados}</b> día(s).
    <br/><br/>
    Esta comunicación se realiza por escrito con la anticipación correspondiente, conforme a la normativa laboral vigente.
    Las vacaciones deberán iniciar en día lunes o en el siguiente día hábil si aquel fuese feriado.
    """

    elementos.append(Paragraph(texto, styles["TextoVacaciones"]))
    elementos.append(Spacer(1, 34))

    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma del empleador / RRHH", "Firma del funcionario"],
        ["", ""],
        ["Fecha de recepción: ____/____/______", "Aclaración: ____________________"],
    ], colWidths=[80 * mm, 80 * mm])

    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    
    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="VACACIONES",
        documento_id=vacacion.id,
        funcionario=funcionario,
        titulo="Notificación de Vacaciones",
    )

    agregar_texto_legal_empresa_pdf(elementos, empresa_pdf)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    registrar_historial(
        request,
        "Vacaciones",
        "Notificación PDF",
        f"Se generó notificación de vacaciones para {funcionario.nombre_completo} del {vacacion.fecha_desde} al {vacacion.fecha_hasta}."
    )

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="notificacion_vacaciones_{funcionario.cedula}_{vacacion.id}.pdf"'
    response.write(pdf)
    return response

def sumar_meses(fecha, meses):
    mes = fecha.month - 1 + meses
    anio = fecha.year + mes // 12
    mes = mes % 12 + 1
    dia = min(fecha.day, monthrange(anio, mes)[1])
    return date(anio, mes, dia)


def calcular_alertas_vacaciones(funcionario):
    if not funcionario.fecha_ingreso:
        return None

    hoy = timezone.localdate()
    ultimo_aniversario = date(hoy.year, funcionario.fecha_ingreso.month, funcionario.fecha_ingreso.day)

    if ultimo_aniversario > hoy:
        ultimo_aniversario = date(hoy.year - 1, funcionario.fecha_ingreso.month, funcionario.fecha_ingreso.day)

    vencimiento = sumar_meses(ultimo_aniversario, 6)
    dias_para_vencer = (vencimiento - hoy).days

    if funcionario.saldo_vacaciones <= 0:
        return None

    if dias_para_vencer < 0:
        return {
            "tipo": "vencida",
            "texto": f"Vacaciones vencidas desde {vencimiento.strftime('%d/%m/%Y')}",
            "vencimiento": vencimiento,
        }

    if dias_para_vencer <= 45:
        return {
            "tipo": "proxima",
            "texto": f"Vacaciones próximas a vencer en {dias_para_vencer} día(s)",
            "vencimiento": vencimiento,
        }

    return None

@login_required
def icl_lista(request):
    permiso = validar_permiso_o_redirigir(request, "icl", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()

    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    dias_mes = monthrange(anio, mes)[1]
    total_dias_laborales_estimados = sum(
        1 for dia in range(1, dias_mes + 1)
        if date(anio, mes, dia).weekday() != 6
    )

    funcionarios = Funcionario.objects.filter(
        activo=True
    ).select_related("turno").order_by("apellido", "nombre")

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()

    resultados = []

    for funcionario in funcionarios:
        icl_activo = icl_activo_para_funcionario(funcionario)
        asistencias = Asistencia.objects.filter(
            funcionario=funcionario,
            fecha__year=anio,
            fecha__month=mes,
            hora_entrada__isnull=False,
        )

        asistencias_count = asistencias.count()
        atrasos_count = asistencias.filter(llego_tarde=True).count()

        permisos_aprobados = PermisoLicencia.objects.filter(
            funcionario=funcionario,
            estado=PermisoLicencia.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        vacaciones_aprobadas = Vacacion.objects.filter(
            funcionario=funcionario,
            estado=Vacacion.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        dias_libres_mes = contar_dias_libres_mes(funcionario, mes, anio)
        total_dias_laborales_reales = max(total_dias_laborales_estimados - dias_libres_mes, 0)
        ausencias_estimadas = max(total_dias_laborales_reales - asistencias_count, 0)
        ausencias_justificadas = permisos_aprobados + vacaciones_aprobadas
        ausencias_no_justificadas = max(ausencias_estimadas - ausencias_justificadas, 0)

        icl = 100 - (atrasos_count * 2) - (ausencias_no_justificadas * 5)
        icl = max(0, min(100, icl))

        bono_base = funcionario.bono_aplicable
        bono_pagable_icl = Decimal("0.00") if funcionario.usa_salario_diferenciado else calcular_bono_pagable_por_icl(bono_base, icl, icl_activo)
        salario_base = funcionario.salario_base_aplicable
        salario_bruto_mes = funcionario.salario_bruto_aplicable if funcionario.usa_salario_diferenciado else (salario_base + bono_pagable_icl).quantize(Decimal("0.01"))
        deudas_mes = funcionario.descuento_deudas_mes
        salario_neto_mes = salario_bruto_mes - funcionario.descuento_ips - deudas_mes
        if salario_neto_mes < 0:
            salario_neto_mes = Decimal("0.00")
        salario_neto_mes = salario_neto_mes.quantize(Decimal("0.01"))

        resultados.append({
            "funcionario": funcionario,
            "asistencias": asistencias_count,
            "atrasos": atrasos_count,
            "ausencias_estimadas": ausencias_estimadas,
            "permisos_aprobados": permisos_aprobados,
            "vacaciones_aprobadas": vacaciones_aprobadas,
            "ausencias_no_justificadas": ausencias_no_justificadas,
            "dias_libres_mes": dias_libres_mes,
            "icl": icl,
            "icl_activo": icl_activo,
            "bono_base": bono_base,
            "bono_pagable_icl": bono_pagable_icl,
            "salario_base_mes": salario_base,
            "salario_bruto": salario_bruto_mes,
            "salario_neto": salario_neto_mes,
            "deudas_mes": deudas_mes,
        })

    resultados.sort(
        key=lambda x: (-x["icl"], x["funcionario"].apellido, x["funcionario"].nombre)
    )

    top_5 = resultados[:5]
    peores_5 = sorted(
        resultados,
        key=lambda x: (x["icl"], x["funcionario"].apellido, x["funcionario"].nombre)
    )[:5]

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]
    anios = list(range(hoy.year - 2, hoy.year + 2))

    return render(request, "core/icl_lista.html", {
        "resultados": resultados,
        "top_5": top_5,
        "peores_5": peores_5,
        "mes": mes,
        "anio": anio,
        "meses": meses,
        "anios": anios,
        "total_dias_laborales_estimados": total_dias_laborales_estimados,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
        "icl_activo_empresa": True if admin_master or not empresa_usuario else empresa_usuario.icl_activo,
    })


@login_required
def reportes(request):
    permiso = validar_permiso_o_redirigir(request, "reportes", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()

    fecha_str = request.GET.get("fecha", str(hoy))
    funcionario_id = request.GET.get("funcionario", "")
    sucursal_id = request.GET.get("sucursal", "")
    q = request.GET.get("q", "").strip()

    try:
        mes = int(request.GET.get("mes", hoy.month))
    except (TypeError, ValueError):
        mes = hoy.month

    try:
        anio = int(request.GET.get("anio", hoy.year))
    except (TypeError, ValueError):
        anio = hoy.year

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    try:
        fecha_reporte = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        fecha_reporte = hoy

    funcionarios = Funcionario.objects.filter(activo=True)

    sucursales = Sucursal.objects.filter(activo=True)

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
            sucursales = sucursales.filter(empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()
            sucursales = sucursales.none()

    if sucursal_id:
        funcionarios = funcionarios.filter(sucursal_rel_id=sucursal_id)

    if q:
        funcionarios = funcionarios.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            _documento_busqueda_q("cedula", "tipo_documento", q)
        )

    funcionarios = funcionarios.select_related(
        "turno",
        "sucursal_rel",
        "sucursal_rel__empresa"
    ).order_by("apellido", "nombre")

    asistencias_dia = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(fecha=fecha_reporte)

    if not admin_master:
        if empresa_usuario:
            asistencias_dia = asistencias_dia.filter(
                funcionario__sucursal_rel__empresa=empresa_usuario
            )
        else:
            asistencias_dia = asistencias_dia.none()

    if sucursal_id:
        asistencias_dia = asistencias_dia.filter(
            funcionario__sucursal_rel_id=sucursal_id
        )

    if q:
        asistencias_dia = asistencias_dia.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q)
        )

    if funcionario_id:
        funcionarios = funcionarios.filter(id=funcionario_id)
        asistencias_dia = asistencias_dia.filter(funcionario_id=funcionario_id)

    asistencias_dia = asistencias_dia.order_by(
        "funcionario__apellido",
        "funcionario__nombre"
    )

    funcionarios_con_turno = funcionarios.filter(turno__isnull=False)

    ids_con_asistencia = list(
        asistencias_dia.values_list("funcionario_id", flat=True).distinct()
    )

    permisos_dia = PermisoLicencia.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(
        funcionario__in=funcionarios,
        estado=PermisoLicencia.Estados.APROBADO,
        fecha_desde=fecha_reporte,
    )

    vacaciones_dia = Vacacion.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(
        funcionario__in=funcionarios,
        estado=Vacacion.Estados.APROBADO,
        fecha_desde=fecha_reporte,
    )

    ids_justificados = set(permisos_dia.values_list("funcionario_id", flat=True))
    ids_justificados.update(vacaciones_dia.values_list("funcionario_id", flat=True))

    ahora = timezone.localtime()
    ausentes_ids_inteligentes = []

    for funcionario in funcionarios_con_turno:
        if funcionario.id in ids_con_asistencia or funcionario.id in ids_justificados:
            continue

        if funcionario_tiene_dia_libre(funcionario, fecha_reporte):
            continue

        if not funcionario.turno or not funcionario.turno.hora_entrada:
            continue

        entrada_programada = timezone.make_aware(
            datetime.combine(fecha_reporte, funcionario.turno.hora_entrada)
        )

        entrada_limite = entrada_programada + timezone.timedelta(
            minutes=funcionario.turno.tolerancia_minutos or 0
        )

        if fecha_reporte < hoy:
            ausentes_ids_inteligentes.append(funcionario.id)
        elif fecha_reporte == hoy and ahora >= entrada_limite:
            ausentes_ids_inteligentes.append(funcionario.id)

    ausentes_dia = funcionarios_con_turno.filter(id__in=ausentes_ids_inteligentes)

    llegadas_tarde = asistencias_dia.filter(
        hora_entrada__isnull=False,
        llego_tarde=True
    )

    presentes_en_horario = asistencias_dia.filter(
        hora_entrada__isnull=False,
        llego_tarde=False
    )

    sin_salida = asistencias_dia.filter(
        hora_entrada__isnull=False,
        hora_salida__isnull=True
    )

    permisos_licencias_dia = []

    for item in permisos_dia:
        permisos_licencias_dia.append({
            "tipo": "Permiso / Licencia",
            "funcionario": item.funcionario,
            "obj": item,
        })

    for item in vacaciones_dia:
        permisos_licencias_dia.append({
            "tipo": "Vacación",
            "funcionario": item.funcionario,
            "obj": item,
        })

    permisos_reporte_dia = permisos_dia
    vacaciones_reporte_dia = vacaciones_dia

    presentes_dia = asistencias_dia.filter(hora_entrada__isnull=False).count()
    tardanzas_dia = llegadas_tarde.count()
    salidas_dia = asistencias_dia.filter(hora_salida__isnull=False).count()
    ausencias_dia = ausentes_dia.count()
    justificados_dia = len({item["funcionario"].id for item in permisos_licencias_dia})
    sin_salida_dia = sin_salida.count()
    programados_dia = funcionarios_con_turno.count()
    en_horario_dia = presentes_en_horario.count()

    requieren_atencion_hoy = tardanzas_dia + ausencias_dia + sin_salida_dia

    porcentaje_asistencia = 0
    if programados_dia > 0:
        porcentaje_asistencia = round((presentes_dia / programados_dia) * 100, 1)

    porcentaje_cumplimiento = 0
    if programados_dia > 0:
        porcentaje_cumplimiento = round((en_horario_dia / programados_dia) * 100, 1)

    resultados_mensuales = []
    funcionarios_para_mes = funcionarios

    dias_mes = monthrange(anio, mes)[1]

    total_dias_laborales_estimados = sum(
        1 for dia in range(1, dias_mes + 1)
        if date(anio, mes, dia).weekday() != 6
    )

    for funcionario in funcionarios_para_mes:
        asistencias_mes = Asistencia.objects.filter(
            funcionario=funcionario,
            fecha__year=anio,
            fecha__month=mes,
            hora_entrada__isnull=False,
        )

        asistencias_count = asistencias_mes.count()
        atrasos_count = asistencias_mes.filter(llego_tarde=True).count()

        dias_libres_mes = contar_dias_libres_mes(funcionario, mes, anio)

        total_dias_laborales_reales = max(
            total_dias_laborales_estimados - dias_libres_mes,
            0
        )

        ausencias_estimadas = max(
            total_dias_laborales_reales - asistencias_count,
            0
        )

        permisos_aprobados = PermisoLicencia.objects.filter(
            funcionario=funcionario,
            estado=PermisoLicencia.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        vacaciones_aprobadas = Vacacion.objects.filter(
            funcionario=funcionario,
            estado=Vacacion.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        ausencias_no_justificadas = max(
            ausencias_estimadas - (permisos_aprobados + vacaciones_aprobadas),
            0
        )

        icl = 100 - (atrasos_count * 2) - (ausencias_no_justificadas * 5)
        icl = max(0, min(100, icl))
        icl_activo = icl_activo_para_funcionario(funcionario)
        salario_base_mes = funcionario.salario_base_aplicable
        bono_base_mes = funcionario.bono_aplicable
        bono_mes = Decimal("0.00") if funcionario.usa_salario_diferenciado else calcular_bono_pagable_por_icl(bono_base_mes, icl, icl_activo)
        salario_bruto_mes = funcionario.salario_bruto_aplicable if funcionario.usa_salario_diferenciado else (salario_base_mes + bono_mes).quantize(Decimal("0.01"))
        salario_neto_mes = salario_bruto_mes - funcionario.descuento_ips - funcionario.descuento_deudas_mes
        if salario_neto_mes < 0:
            salario_neto_mes = Decimal("0.00")
        salario_neto_mes = salario_neto_mes.quantize(Decimal("0.01"))

        resultados_mensuales.append({
            "funcionario": funcionario,
            "asistencias": asistencias_count,
            "atrasos": atrasos_count,
            "ausencias": ausencias_estimadas,
            "permisos_aprobados": permisos_aprobados,
            "vacaciones_aprobadas": vacaciones_aprobadas,
            "dias_libres_mes": dias_libres_mes,
            "icl": icl,
            "icl_activo": icl_activo,
            "salario_bruto": salario_bruto_mes,
            "deudas_mes": funcionario.descuento_deudas_mes,
            "salario_neto": salario_neto_mes,
        })

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]

    anios = list(range(hoy.year - 2, hoy.year + 2))

    dias_libres_reporte_dia = []

    for funcionario in funcionarios:
        if funcionario_tiene_dia_libre(funcionario, fecha_reporte):
            dias_libres_reporte_dia.append(funcionario)

    fecha_jornada_incompleta = fecha_reporte - timezone.timedelta(days=1)

    permisos_jornada_incompleta = PermisoLicencia.objects.filter(
        funcionario__in=funcionarios,
        estado=PermisoLicencia.Estados.APROBADO,
        fecha_desde__lte=fecha_jornada_incompleta,
        fecha_hasta__gte=fecha_jornada_incompleta,
    )

    vacaciones_jornada_incompleta = Vacacion.objects.filter(
        funcionario__in=funcionarios,
        estado=Vacacion.Estados.APROBADO,
        fecha_desde__lte=fecha_jornada_incompleta,
        fecha_hasta__gte=fecha_jornada_incompleta,
    )

    ids_justificados_jornada_incompleta = set(
        permisos_jornada_incompleta.values_list("funcionario_id", flat=True)
    )
    ids_justificados_jornada_incompleta.update(
        vacaciones_jornada_incompleta.values_list("funcionario_id", flat=True)
    )

    asistencias_jornada_incompleta = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(
        fecha=fecha_jornada_incompleta,
        hora_entrada__isnull=False,
        funcionario__in=funcionarios,
    ).order_by("funcionario__apellido", "funcionario__nombre")

    jornadas_incompletas = []

    for asistencia in asistencias_jornada_incompleta:
        funcionario = asistencia.funcionario

        if funcionario.id in ids_justificados_jornada_incompleta:
            continue

        if funcionario_tiene_dia_libre(funcionario, fecha_jornada_incompleta):
            continue

        if asistencia.horas_trabajadas_segundos < 8 * 60 * 60:
            jornadas_incompletas.append(asistencia)

    jornadas_incompletas_dia = len(jornadas_incompletas)

    return render(request, "core/reportes.html", {
        "fecha_reporte": fecha_reporte,
        "fecha_jornada_incompleta": fecha_jornada_incompleta,
        "funcionarios": funcionarios,
        "funcionario_id": funcionario_id,
        "sucursal_id": sucursal_id,
        "sucursales": sucursales.order_by("nombre"),
        "q": q,

        "asistencias_dia": asistencias_dia,
        "presentes_dia": presentes_dia,
        "tardanzas_dia": tardanzas_dia,
        "salidas_dia": salidas_dia,
        "ausencias_dia": ausencias_dia,
        "justificados_dia": justificados_dia,
        "sin_salida_dia": sin_salida_dia,
        "programados_dia": programados_dia,
        "en_horario_dia": en_horario_dia,
        "porcentaje_asistencia": porcentaje_asistencia,
        "porcentaje_cumplimiento": porcentaje_cumplimiento,
        "requieren_atencion_hoy": requieren_atencion_hoy,
        "llegadas_tarde": llegadas_tarde,
        "ausentes_dia": ausentes_dia,
        "permisos_licencias_dia": permisos_licencias_dia,
        "permisos_reporte_dia": permisos_reporte_dia,
        "vacaciones_reporte_dia": vacaciones_reporte_dia,
        "dias_libres_reporte_dia": dias_libres_reporte_dia,
        "sin_salida": sin_salida,
        "presentes_en_horario": presentes_en_horario,
        "jornadas_incompletas": jornadas_incompletas,
        "jornadas_incompletas_dia": jornadas_incompletas_dia,

        "mes": mes,
        "anio": anio,
        "meses": meses,
        "anios": anios,
        "resultados_mensuales": resultados_mensuales,

        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def historial_lista(request):
    permiso = validar_permiso_o_redirigir(request, "historial", "puede_ver")
    if permiso:
        return permiso

    if not es_admin_master(request.user):
        messages.error(request, "El historial general solo está disponible para el administrador master.")
        return redirect("dashboard")

    q = request.GET.get("q", "").strip()
    empresa_usuario = obtener_empresa_activa(request)
    historial = HistorialAccion.objects.select_related("usuario", "empresa").all()

    if empresa_usuario:
        historial = historial.filter(empresa=empresa_usuario)

    if q:
        historial = historial.filter(
            Q(modulo__icontains=q) |
            Q(accion__icontains=q) |
            Q(descripcion__icontains=q) |
            Q(usuario__username__icontains=q) |
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q)
        )

    return render(request, "core/historial_lista.html", {
        "historial": historial[:300],
        "q": q,
    })

def calcular_aguinaldo_funcionario(funcionario, anio):
    nominas = NominaMensual.objects.filter(
        funcionario=funcionario,
        anio=anio
    )

    total_remuneraciones = nominas.aggregate(
        total=Sum("salario_bruto")
    )["total"] or Decimal("0.00")

    meses_computados = nominas.count()
    monto_aguinaldo = (total_remuneraciones / Decimal("12")).quantize(Decimal("0.01"))

    aguinaldo, creado = AguinaldoAnual.objects.update_or_create(
        funcionario=funcionario,
        anio=anio,
        defaults={
            "empresa": funcionario.empresa,
            "sucursal": funcionario.sucursal_rel,
            "total_remuneraciones": total_remuneraciones,
            "monto_aguinaldo": monto_aguinaldo,
            "meses_computados": meses_computados,
        }
    )

    return aguinaldo


@login_required
def aguinaldo_lista(request):
    permiso = validar_permiso_o_redirigir(request, "aguinaldo", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    anio = int(request.GET.get("anio", hoy.year))
    estado = request.GET.get("estado", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    aguinaldos = AguinaldoAnual.objects.select_related(
        "funcionario",
        "empresa",
        "sucursal",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(anio=anio)

    if not admin_master:
        if empresa_usuario:
            aguinaldos = aguinaldos.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            aguinaldos = aguinaldos.none()

    if estado:
        aguinaldos = aguinaldos.filter(estado=estado)

    if sucursal_id:
        aguinaldos = aguinaldos.filter(funcionario__sucursal_rel_id=sucursal_id)

    if q:
        aguinaldos = aguinaldos.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q)
        )

    total_general = aguinaldos.aggregate(
        total=Sum("monto_aguinaldo")
    )["total"] or Decimal("0.00")

    total_pagados = aguinaldos.filter(estado=AguinaldoAnual.Estados.PAGADO).count()
    total_pendientes = aguinaldos.filter(estado=AguinaldoAnual.Estados.PENDIENTE).count()

    anios = list(range(hoy.year - 3, hoy.year + 2))

    if admin_master:
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
    else:
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()

    return render(request, "core/aguinaldo_lista.html", {
        "aguinaldos": aguinaldos.order_by("funcionario__apellido", "funcionario__nombre"),
        "anio": anio,
        "anios": anios,
        "estado": estado,
        "sucursal_id": sucursal_id,
        "sucursales": sucursales,
        "q": q,
        "estados": AguinaldoAnual.Estados.choices,
        "total_general": total_general,
        "total_pagados": total_pagados,
        "total_pendientes": total_pendientes,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def aguinaldo_generar(request):
    permiso = validar_permiso_o_redirigir(request, "aguinaldo", "puede_crear")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    anio = int(request.GET.get("anio", hoy.year))

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionarios = Funcionario.objects.filter(activo=True).select_related(
        "sucursal_rel",
        "sucursal_rel__empresa"
    )

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()

    cantidad = 0

    for funcionario in funcionarios:
        calcular_aguinaldo_funcionario(funcionario, anio)
        cantidad += 1

    registrar_historial(
        request,
        "Aguinaldo",
        "Generar/Recalcular",
        f"Se generó o recalculó el aguinaldo del año {anio} para {cantidad} funcionario(s)."
    )

    messages.success(request, f"Aguinaldo {anio} generado correctamente para {cantidad} funcionario(s).")
    return redirect(f"/aguinaldo/?anio={anio}")


@login_required
def aguinaldo_toggle_pagado(request, pk):
    permiso = validar_permiso_o_redirigir(request, "aguinaldo", "puede_pagar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    aguinaldo = get_object_or_404(
        AguinaldoAnual.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not aguinaldo.funcionario.sucursal_rel or aguinaldo.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar aguinaldos de otra empresa.")
            return redirect("aguinaldo_lista")

    if aguinaldo.estado == AguinaldoAnual.Estados.PAGADO:
        aguinaldo.estado = AguinaldoAnual.Estados.PENDIENTE
        aguinaldo.fecha_pago = None
        accion = "revirtió a pendiente"
    else:
        aguinaldo.estado = AguinaldoAnual.Estados.PAGADO
        aguinaldo.fecha_pago = timezone.localdate()
        accion = "marcó como pagado"

    aguinaldo.save()

    registrar_historial(
        request,
        "Aguinaldo",
        "Cambio de estado",
        f"Se {accion} el aguinaldo de {aguinaldo.funcionario.nombre_completo} del año {aguinaldo.anio}."
    )

    messages.success(request, "Estado de aguinaldo actualizado correctamente.")
    return redirect(f"/aguinaldo/?anio={aguinaldo.anio}")


@login_required
def aguinaldo_pdf(request, pk):
    permiso = validar_permiso_o_redirigir(request, "aguinaldo", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    aguinaldo = get_object_or_404(
        AguinaldoAnual.objects.select_related(
            "funcionario",
            "empresa",
            "sucursal",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa",
        ),
        pk=pk
    )

    if not admin_master:
        if not aguinaldo.funcionario.sucursal_rel or aguinaldo.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes exportar aguinaldos de otra empresa.")
            return redirect("aguinaldo_lista")

    funcionario = aguinaldo.funcionario
    empresa_pdf = obtener_empresa_documento(funcionario=funcionario)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TituloAguinaldo",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=10,
    ))

    styles.add(ParagraphStyle(
        name="TextoAguinaldo",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=13,
        textColor=colors.HexColor("#111827"),
    ))

    elementos = []

    elementos += construir_encabezado_empresa_pdf(empresa_pdf, "RECIBO DE AGUINALDO")
    elementos.append(Paragraph("AGUINALDO ANUAL", styles["TituloAguinaldo"]))
    elementos.append(Spacer(1, 8))

    datos = [
        ["Funcionario", funcionario.nombre_completo],
        ["Documento", funcionario.documento_compacto],
        ["Empresa", funcionario.empresa_mostrar],
        ["Sucursal", funcionario.sucursal_mostrar],
        ["Cargo", funcionario.cargo or "-"],
        ["Año", str(aguinaldo.anio)],
        ["Meses computados", str(aguinaldo.meses_computados)],
        ["Estado", aguinaldo.get_estado_display()],
        ["Fecha de pago", aguinaldo.fecha_pago.strftime("%d/%m/%Y") if aguinaldo.fecha_pago else "-"],
    ]

    tabla_datos = Table(datos, colWidths=[55 * mm, 105 * mm])
    tabla_datos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 12))

    tabla_calc = Table([
        ["Concepto", "Monto"],
        ["Total remuneraciones computables del año", _gs(aguinaldo.total_remuneraciones)],
        ["AGUINALDO A COBRAR", _gs(aguinaldo.monto_aguinaldo)],
    ], colWidths=[115 * mm, 45 * mm])

    tabla_calc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dcfce7")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#166534")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabla_calc)
    elementos.append(Spacer(1, 14))

    texto = """
    Se deja constancia de que el presente cálculo corresponde al aguinaldo anual, calculado sobre la base
    de las remuneraciones computables registradas en el sistema para el año indicado. El aguinaldo no constituye
    salario mensual ordinario y se documenta de forma separada para fines administrativos y laborales.
    """
    elementos.append(Paragraph(texto, styles["TextoAguinaldo"]))
    elementos.append(Spacer(1, 28))

    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma responsable / RRHH", "Firma funcionario"],
        ["", ""],
        ["Aclaración: ____________________", "Aclaración: ____________________"],
    ], colWidths=[80 * mm, 80 * mm])

    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="AGUINALDO",
        documento_id=aguinaldo.id,
        funcionario=funcionario,
        titulo="Recibo de Aguinaldo",
    )

    agregar_texto_legal_empresa_pdf(elementos, empresa_pdf)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="aguinaldo_{funcionario.cedula}_{aguinaldo.anio}.pdf"'
    response.write(pdf)
    return response

@login_required
def planilla_bancaria_lista(request):
    permiso = validar_permiso_o_redirigir(request, "planilla_bancaria", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    planillas = PlanillaBancaria.objects.select_related(
        "empresa",
        "sucursal",
        "generado_por"
    )

    if not admin_master:
        if empresa_usuario:
            planillas = planillas.filter(empresa=empresa_usuario)
        else:
            planillas = planillas.none()

    return render(request, "core/planilla_bancaria_lista.html", {
        "planillas": planillas,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def planilla_bancaria_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "planilla_bancaria", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = PlanillaBancariaForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

        if form.is_valid():
            planilla = form.save(commit=False)

            if not admin_master:
                if not empresa_usuario:
                    messages.error(request, "Tu usuario no tiene empresa asignada.")
                    return redirect("planilla_bancaria_lista")

                planilla.empresa = empresa_usuario

                if planilla.sucursal and planilla.sucursal.empresa != empresa_usuario:
                    messages.error(request, "No puedes generar planillas para sucursales de otra empresa.")
                    return redirect("planilla_bancaria_lista")

            planilla.generado_por = request.user

            nominas = NominaMensual.objects.select_related(
                "funcionario",
                "funcionario__sucursal_rel",
                "funcionario__sucursal_rel__empresa",
            ).filter(
                anio=planilla.anio,
                mes=planilla.mes,
            )

            if planilla.empresa:
                nominas = nominas.filter(
                    funcionario__sucursal_rel__empresa=planilla.empresa
                )

            if planilla.sucursal:
                nominas = nominas.filter(
                    funcionario__sucursal_rel=planilla.sucursal
                )

            total_importe = Decimal("0")
            total_funcionarios = 0

            for nomina in nominas:
                total_importe += nomina.salario_neto
                total_funcionarios += 1

            planilla.total_importe = total_importe
            planilla.total_funcionarios = total_funcionarios
            planilla.estado = PlanillaBancaria.Estados.GENERADA

            planilla.save()

            registrar_historial(
                request,
                "Planilla Bancaria",
                "Generar",
                f"Se generó planilla bancaria {planilla.banco} {planilla.mes:02d}/{planilla.anio}."
            )

            messages.success(request, "Planilla bancaria generada correctamente.")
            return redirect("planilla_bancaria_lista")

    else:
        hoy = timezone.localdate()

        form = PlanillaBancariaForm(initial={
            "anio": hoy.year,
            "mes": hoy.month,
        })

        if not admin_master and empresa_usuario:
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

    return render(request, "core/planilla_bancaria_form.html", {
        "form": form,
        "titulo_form": "Nueva planilla bancaria",
        "boton_texto": "Generar planilla",
    })


@login_required
def planilla_bancaria_exportar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "planilla_bancaria", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    planilla = get_object_or_404(
        PlanillaBancaria.objects.select_related("empresa", "sucursal"),
        pk=pk
    )

    if not admin_master:
        if planilla.empresa != empresa_usuario:
            messages.error(request, "No puedes exportar planillas de otra empresa.")
            return redirect("planilla_bancaria_lista")

    nominas = NominaMensual.objects.select_related(
        "funcionario"
    ).filter(
        anio=planilla.anio,
        mes=planilla.mes,
    )

    if planilla.empresa:
        nominas = nominas.filter(
            funcionario__sucursal_rel__empresa=planilla.empresa
        )

    if planilla.sucursal:
        nominas = nominas.filter(
            funcionario__sucursal_rel=planilla.sucursal
        )

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="planilla_{planilla.banco}_{planilla.mes:02d}_{planilla.anio}.csv"'
    )

    writer = csv.writer(response)

    writer.writerow([
        "CEDULA",
        "NOMBRE",
        "BANCO",
        "TIPO_CUENTA",
        "NUMERO_CUENTA",
        "MONEDA",
        "MONTO",
        "CONCEPTO",
    ])

    for nomina in nominas:
        funcionario = nomina.funcionario

        writer.writerow([
            funcionario.documento_compacto,
            funcionario.nombre_completo,
            funcionario.banco or planilla.banco,
            funcionario.tipo_cuenta or "",
            funcionario.numero_cuenta or "",
            "PYG",
            int(nomina.salario_neto),
            f"SALARIO {planilla.mes:02d}/{planilla.anio}",
        ])

    planilla.estado = PlanillaBancaria.Estados.EXPORTADA
    planilla.save(update_fields=["estado"])

    registrar_historial(
        request,
        "Planilla Bancaria",
        "Exportar CSV",
        f"Se exportó CSV de planilla bancaria {planilla.banco}."
    )

    return response

def obtener_saldo_banco_horas(funcionario):
    ultimo = BancoHorasMovimiento.objects.filter(
        funcionario=funcionario
    ).order_by("-fecha", "-id").first()

    return ultimo.saldo_nuevo if ultimo else 0


def registrar_movimiento_banco(
    funcionario,
    tipo,
    minutos,
    observacion="",
    origen="sistema",
    user=None,
    fecha=None,
):
    saldo_anterior = obtener_saldo_banco_horas(funcionario)
    saldo_nuevo = saldo_anterior + minutos

    return BancoHorasMovimiento.objects.create(
        funcionario=funcionario,
        empresa=funcionario.empresa,
        sucursal=funcionario.sucursal_rel,
        fecha=fecha or timezone.localdate(),
        tipo=tipo,
        minutos=minutos,
        saldo_anterior=saldo_anterior,
        saldo_nuevo=saldo_nuevo,
        observacion=observacion,
        origen=origen,
        creado_por=user,
    )


def recalcular_banco_horas_funcionario(funcionario, user=None):
    BancoHorasMovimiento.objects.filter(
        funcionario=funcionario,
        origen="asistencia"
    ).delete()

    asistencias = Asistencia.objects.filter(
        funcionario=funcionario,
        hora_entrada__isnull=False,
        hora_salida__isnull=False,
    ).order_by("fecha")

    for asistencia in asistencias:
        segundos = asistencia.horas_trabajadas_segundos or 0

        # Si no hay horas reales trabajadas, no se toca el banco.
        # Las ausencias impactan en sueldo/nómina, no en banco de horas.
        if segundos <= 0:
            continue

        horas_reales = segundos / 3600
        diferencia_horas = horas_reales - 8
        diferencia_minutos = int(diferencia_horas * 60)

        bloques_30 = abs(diferencia_minutos) // 30
        minutos_finales = bloques_30 * 30

        if minutos_finales == 0:
            continue

        if diferencia_minutos > 0:
            registrar_movimiento_banco(
                funcionario=funcionario,
                tipo=BancoHorasMovimiento.Tipos.GENERADO,
                minutos=minutos_finales,
                observacion=f"Horas extras generadas automáticamente ({asistencia.fecha:%d/%m/%Y})",
                origen="asistencia",
                user=user,
                fecha=asistencia.fecha,
            )
        else:
            registrar_movimiento_banco(
                funcionario=funcionario,
                tipo=BancoHorasMovimiento.Tipos.DESCUENTO,
                minutos=-minutos_finales,
                observacion=f"Descuento automático por jornada incompleta ({asistencia.fecha:%d/%m/%Y})",
                origen="asistencia",
                user=user,
                fecha=asistencia.fecha,
            )


@login_required
def banco_horas_lista(request):
    permiso = validar_permiso_o_redirigir(request, "banco_horas", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    sucursal_id = request.GET.get("sucursal", "").strip()
    q = request.GET.get("q", "").strip()

    funcionarios = Funcionario.objects.filter(activo=True).select_related(
        "sucursal_rel",
        "sucursal_rel__empresa"
    )

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()

    if sucursal_id:
        funcionarios = funcionarios.filter(sucursal_rel_id=sucursal_id)

    if q:
        funcionarios = funcionarios.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            _documento_busqueda_q("cedula", "tipo_documento", q)
        )

    if admin_master:
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
    else:
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()

    datos = []

    for funcionario in funcionarios.order_by("apellido", "nombre"):
        saldo = obtener_saldo_banco_horas(funcionario)

        horas = abs(saldo) // 60
        minutos = abs(saldo) % 60

        saldo_texto = f"{horas}:{minutos:02d} hs"
        if saldo < 0:
            saldo_texto = f"-{saldo_texto}"

        datos.append({
            "funcionario": funcionario,
            "saldo": saldo,
            "saldo_texto": saldo_texto,
        })

    return render(request, "core/banco_horas_lista.html", {
        "datos": datos,
        "sucursales": sucursales,
        "sucursal_id": sucursal_id,
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def banco_horas_recalcular(request, funcionario_id):
    permiso = validar_permiso_o_redirigir(request, "banco_horas", "puede_editar")
    if permiso:
        return permiso

    funcionario = get_object_or_404(
        Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"),
        pk=funcionario_id
    )

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar banco de horas de otra empresa.")
            return redirect("banco_horas_lista")

    recalcular_banco_horas_funcionario(funcionario, request.user)

    registrar_historial(
        request,
        "Banco de Horas",
        "Recalcular",
        f"Se recalculó banco de horas de {funcionario.nombre_completo}."
    )

    messages.success(request, "Banco de horas recalculado correctamente.")
    return redirect("banco_horas_lista")


@login_required
def banco_horas_otorgar(request):
    permiso = validar_permiso_o_redirigir(request, "banco_horas", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = BancoHorasOtorgarForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            funcionario = form.cleaned_data["funcionario"]
            horas = form.cleaned_data["horas"]
            observacion = form.cleaned_data["observacion"]

            if not admin_master:
                if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes otorgar horas a funcionarios de otra empresa.")
                    return redirect("banco_horas_lista")

            minutos = horas * 60

            registrar_movimiento_banco(
                funcionario=funcionario,
                tipo=BancoHorasMovimiento.Tipos.HORAS_TOMADAS,
                minutos=-minutos,
                observacion=observacion or f"Otorgamiento de {horas} hora(s) libres.",
                origen="rrhh",
                user=request.user,
            )

            registrar_historial(
                request,
                "Banco de Horas",
                "Otorgar horas",
                f"Se otorgó {horas} hora(s) libres a {funcionario.nombre_completo}."
            )

            messages.success(request, "Horas descontadas correctamente del banco.")
            return redirect("banco_horas_lista")

    else:
        form = BancoHorasOtorgarForm()

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/banco_horas_otorgar.html", {
        "form": form,
    })

@login_required
def banco_horas_historial(request, funcionario_id):
    permiso = validar_permiso_o_redirigir(request, "banco_horas", "puede_ver")
    if permiso:
        return permiso

    funcionario = get_object_or_404(
        Funcionario.objects.select_related("sucursal_rel", "sucursal_rel__empresa"),
        pk=funcionario_id
    )

    empresa_usuario = obtener_empresa_activa(request)
    if empresa_usuario:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar banco de horas de otra empresa.")
            return redirect("banco_horas_lista")

    movimientos = BancoHorasMovimiento.objects.filter(
        funcionario=funcionario
    ).select_related("creado_por")

    return render(request, "core/banco_horas_historial.html", {
        "funcionario": funcionario,
        "movimientos": movimientos,
        "saldo_actual": obtener_saldo_banco_horas(funcionario),
    })

@login_required
def nomina_lista(request):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    estado = request.GET.get("estado", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    sectores_seleccionados = [
        sector.strip()
        for sector in request.GET.getlist("sector")
        if sector.strip()
    ]

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    empresa_cierre = None if admin_master else empresa_usuario

    cierre_nomina = CierreNomina.objects.filter(
        mes=mes,
        anio=anio,
        empresa=empresa_cierre,
        cerrado=True
    ).first()

    if request.method == "POST":
        permiso_post = validar_permiso_o_redirigir(request, "nomina", "puede_crear")
        if permiso_post:
            return permiso_post

        funcionarios = Funcionario.objects.filter(activo=True)

        if not admin_master:
            if empresa_usuario:
                funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
            else:
                funcionarios = funcionarios.none()

        if cierre_nomina:
            messages.error(request, "Esta nómina ya está cerrada. Debes reabrirla antes de recalcular.")
            return redirect(f"/nomina/?mes={mes}&anio={anio}")        

        funcionarios = funcionarios.order_by("apellido", "nombre")

        for funcionario in funcionarios:
            generar_nomina_funcionario(funcionario, mes, anio)

        registrar_historial(
            request,
            "Nómina",
            "Generar/Recalcular",
            f"Se generó o recalculó la nómina del período {mes:02d}/{anio}."
        )
        messages.success(request, f"Nómina de {mes:02d}/{anio} generada correctamente.")
        return redirect(f"/nomina/?mes={mes}&anio={anio}")

    nominas = NominaMensual.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).filter(
        mes=mes,
        anio=anio
    )

    if not admin_master:
        if empresa_usuario:
            nominas = nominas.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            nominas = nominas.none()

    if not nominas.exists():
        funcionarios = Funcionario.objects.filter(activo=True)

        if not admin_master:
            if empresa_usuario:
                funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
            else:
                funcionarios = funcionarios.none()

        funcionarios = funcionarios.order_by("apellido", "nombre")

        for funcionario in funcionarios:
            generar_nomina_funcionario(funcionario, mes, anio)

        nominas = NominaMensual.objects.select_related(
            "funcionario",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa"
        ).filter(
            mes=mes,
            anio=anio
        )

        if not admin_master:
            if empresa_usuario:
                nominas = nominas.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
            else:
                nominas = nominas.none()

    if estado:
        nominas = nominas.filter(estado_pago=estado)

    if sucursal_id:
        nominas = nominas.filter(funcionario__sucursal_rel_id=sucursal_id)

    if sectores_seleccionados:
        nominas = nominas.filter(funcionario__sector__in=sectores_seleccionados)

    nominas = nominas.order_by("funcionario__apellido", "funcionario__nombre")

    total_bruto = nominas.aggregate(total=Sum("salario_bruto"))["total"] or Decimal("0.00")
    total_deudas = nominas.aggregate(total=Sum("descuento_deudas"))["total"] or Decimal("0.00")
    total_neto = nominas.aggregate(total=Sum("salario_neto"))["total"] or Decimal("0.00")
    total_pagados = nominas.filter(estado_pago=NominaMensual.EstadosPago.PAGADO).count()
    total_pendientes = nominas.filter(estado_pago=NominaMensual.EstadosPago.PENDIENTE).count()

    meses = [
        (1, "Enero"), (2, "Febrero"), (3, "Marzo"), (4, "Abril"),
        (5, "Mayo"), (6, "Junio"), (7, "Julio"), (8, "Agosto"),
        (9, "Septiembre"), (10, "Octubre"), (11, "Noviembre"), (12, "Diciembre"),
    ]
    anios = list(range(hoy.year - 2, hoy.year + 2))

    if admin_master:
        sucursales = Sucursal.objects.filter(activo=True).order_by("empresa__nombre", "nombre")
    else:
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()

    sectores_qs = Funcionario.objects.filter(activo=True).exclude(sector="")
    if not admin_master:
        if empresa_usuario:
            sectores_qs = sectores_qs.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            sectores_qs = sectores_qs.none()
    if sucursal_id:
        sectores_qs = sectores_qs.filter(sucursal_rel_id=sucursal_id)
    sectores = list(
        sectores_qs.order_by("sector").values_list("sector", flat=True).distinct()
    )

    resumen_sectores = []
    sectores_en_nomina = list(
        nominas.order_by("funcionario__sector")
        .values_list("funcionario__sector", flat=True)
        .distinct()
    )
    for sector_nombre in sectores_en_nomina:
        qs_sector = nominas.filter(funcionario__sector=sector_nombre)
        resumen_sectores.append({
            "sector": sector_nombre or "Sin sector",
            "cantidad": qs_sector.count(),
            "total_haberes": qs_sector.aggregate(total=Sum("salario_bruto"))["total"] or Decimal("0.00"),
            "total_ips": qs_sector.aggregate(total=Sum("descuento_ips"))["total"] or Decimal("0.00"),
            "total_descuentos": (
                (qs_sector.aggregate(total=Sum("descuento_ips"))["total"] or Decimal("0.00"))
                + (qs_sector.aggregate(total=Sum("descuento_deudas"))["total"] or Decimal("0.00"))
            ),
            "total_deudas": qs_sector.aggregate(total=Sum("descuento_deudas"))["total"] or Decimal("0.00"),
            "total_ajustes": Decimal("0.00"),
            "total_neto": qs_sector.aggregate(total=Sum("salario_neto"))["total"] or Decimal("0.00"),
        })

    return render(request, "core/nomina_lista.html", {
        "nominas": nominas,
        "mes": mes,
        "anio": anio,
        "estado": estado,
        "sucursal_id": sucursal_id,
        "sectores": sectores,
        "sectores_seleccionados": sectores_seleccionados,
        "resumen_sectores": resumen_sectores,
        "meses": meses,
        "anios": anios,
        "sucursales": sucursales,
        "cierre_nomina": cierre_nomina,
        "total_bruto": total_bruto,
        "total_deudas": total_deudas,
        "total_neto": total_neto,
        "total_pagados": total_pagados,
        "total_pendientes": total_pendientes,
        "estados_pago": NominaMensual.EstadosPago.choices,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def nomina_toggle_pagado(request, pk):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_pagar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    nomina = get_object_or_404(
        NominaMensual.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not nomina.funcionario.sucursal_rel or nomina.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes cambiar nóminas de otra empresa.")
            return redirect(f"/nomina/?mes={nomina.mes}&anio={nomina.anio}")

    if nomina.estado_pago == NominaMensual.EstadosPago.PAGADO:
        nomina.estado_pago = NominaMensual.EstadosPago.PENDIENTE
        nomina.fecha_pago = None
        accion = "revirtió a pendiente"
    else:
        nomina.estado_pago = NominaMensual.EstadosPago.PAGADO
        nomina.fecha_pago = timezone.localdate()
        accion = "marcó como pagada"

    nomina.save()

    registrar_historial(
        request,
        "Nómina",
        "Cambio de estado",
        f"Se {accion} la nómina de {nomina.funcionario.nombre_completo} del período {nomina.mes:02d}/{nomina.anio}."
    )
    messages.success(request, "Estado de nómina actualizado correctamente.")
    return redirect(f"/nomina/?mes={nomina.mes}&anio={nomina.anio}")

def _gs(valor):
    return f"Gs. {Decimal(valor or 0):,.0f}".replace(",", ".")

@login_required
def nomina_cerrar_periodo(request):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_pagar")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    empresa_cierre = None if admin_master else empresa_usuario

    cierre, creado = CierreNomina.objects.get_or_create(
        mes=mes,
        anio=anio,
        empresa=empresa_cierre,
        defaults={
            "cerrado": True,
            "cerrado_por": request.user,
            "cerrado_en": timezone.now(),
            "observacion": "Cierre manual de nómina.",
        }
    )

    if not creado:
        cierre.cerrado = True
        cierre.cerrado_por = request.user
        cierre.cerrado_en = timezone.now()
        cierre.save()

    registrar_historial(
        request,
        "Nómina",
        "Cerrar período",
        f"Se cerró la nómina del período {mes:02d}/{anio}."
    )

    messages.success(request, f"Nómina {mes:02d}/{anio} cerrada correctamente.")
    return redirect(f"/nomina/?mes={mes}&anio={anio}")


@login_required
def nomina_reabrir_periodo(request):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_pagar")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    empresa_cierre = None if admin_master else empresa_usuario

    cierre = CierreNomina.objects.filter(
        mes=mes,
        anio=anio,
        empresa=empresa_cierre,
        cerrado=True
    ).first()

    if cierre:
        cierre.cerrado = False
        cierre.save()

        registrar_historial(
            request,
            "Nómina",
            "Reabrir período",
            f"Se reabrió la nómina del período {mes:02d}/{anio}."
        )

        messages.success(request, f"Nómina {mes:02d}/{anio} reabierta correctamente.")
    else:
        messages.warning(request, "No existe un cierre activo para este período.")

    return redirect(f"/nomina/?mes={mes}&anio={anio}")

def _nomina_permitida(request, nomina):
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if admin_master:
        return True

    return (
        empresa_usuario
        and nomina.funcionario.sucursal_rel
        and nomina.funcionario.sucursal_rel.empresa == empresa_usuario
    )


def _filename_seguro(valor):
    texto = re.sub(r"[^A-Za-z0-9_-]+", "_", str(valor or "sin_sector")).strip("_")
    return texto or "sin_sector"


def _nominas_filtradas_para_exportar(request, mes, anio, sucursal_id="", sectores=None, estado=""):
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None
    sectores = [s for s in (sectores or []) if s]

    nominas = NominaMensual.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).filter(
        mes=mes,
        anio=anio,
    )

    if not admin_master:
        if empresa_usuario:
            nominas = nominas.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            nominas = nominas.none()

    if sucursal_id:
        nominas = nominas.filter(funcionario__sucursal_rel_id=sucursal_id)

    if sectores:
        nominas = nominas.filter(funcionario__sector__in=sectores)

    if estado:
        nominas = nominas.filter(estado_pago=estado)

    return nominas.order_by(
        "funcionario__sector",
        "funcionario__apellido",
        "funcionario__nombre",
    )


def _sectores_desde_nominas(nominas):
    return [
        sector or ""
        for sector in nominas.order_by("funcionario__sector")
        .values_list("funcionario__sector", flat=True)
        .distinct()
    ]


def _totales_nomina_queryset(nominas):
    total_bruto = nominas.aggregate(total=Sum("salario_bruto"))["total"] or Decimal("0.00")
    total_ips = nominas.aggregate(total=Sum("descuento_ips"))["total"] or Decimal("0.00")
    total_deudas = nominas.aggregate(total=Sum("descuento_deudas"))["total"] or Decimal("0.00")
    total_neto = nominas.aggregate(total=Sum("salario_neto"))["total"] or Decimal("0.00")
    return {
        "cantidad": nominas.count(),
        "total_haberes": total_bruto,
        "total_ips": total_ips,
        "total_deudas": total_deudas,
        "total_descuentos": total_ips + total_deudas,
        "total_ajustes": Decimal("0.00"),
        "total_neto": total_neto,
    }


def _pdf_nomina_sector_bytes(request, nominas, sector_nombre, mes, anio):
    primera = nominas.first()
    empresa_pdf = obtener_empresa_documento(funcionario=primera.funcionario) if primera else None
    sucursal_nombre = primera.funcionario.sucursal_mostrar if primera else "-"
    totales = _totales_nomina_queryset(nominas)
    sector_label = sector_nombre or "Sin sector"

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    elementos = []

    if empresa_pdf:
        elementos += construir_encabezado_empresa_pdf(empresa_pdf, "NOMINA POR SECTOR")
    else:
        elementos.append(Paragraph("NOMINA POR SECTOR", styles["Heading1"]))

    resumen = Table([
        ["Sector", sector_label],
        ["Sucursal", sucursal_nombre],
        ["Periodo", f"{mes:02d}/{anio}"],
        ["Fecha de generacion", timezone.localtime().strftime("%d/%m/%Y %H:%M")],
        ["Cantidad de funcionarios", str(totales["cantidad"])],
        ["Total haberes", _gs(totales["total_haberes"])],
        ["Total IPS", _gs(totales["total_ips"])],
        ["Total otros descuentos", _gs(totales["total_deudas"])],
        ["Total ajustes", _gs(totales["total_ajustes"])],
        ["Total neto a pagar", _gs(totales["total_neto"])],
    ], colWidths=[60 * mm, 110 * mm])
    resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))
    elementos.append(Spacer(1, 8))
    elementos.append(resumen)
    elementos.append(Spacer(1, 10))

    data = [["Funcionario", "CI", "Cargo", "Modalidad", "Haberes", "IPS", "Otros desc.", "Ajustes", "Neto"]]
    for n in nominas:
        modalidad = getattr(n.funcionario, "modalidad_salarial_display", "Salario base + bono")
        data.append([
            n.funcionario.nombre_completo,
            n.funcionario.documento_compacto,
            n.funcionario.cargo or "-",
            modalidad,
            _gs(n.salario_bruto),
            _gs(n.descuento_ips),
            _gs(n.descuento_deudas),
            _gs(0),
            _gs(n.salario_neto),
        ])

    tabla = Table(data, colWidths=[34 * mm, 18 * mm, 24 * mm, 26 * mm, 22 * mm, 18 * mm, 22 * mm, 18 * mm, 22 * mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.2),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cbd5e1")),
        ("ALIGN", (4, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


@login_required
def nomina_extracto_pdf(request, pk):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_ver")
    if permiso:
        return permiso

    nomina = get_object_or_404(
        NominaMensual.objects.select_related(
            "funcionario",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa"
        ),
        pk=pk
    )

    if not _nomina_permitida(request, nomina):
        messages.error(request, "No puedes exportar nóminas de otra empresa.")
        return redirect("nomina_lista")

    funcionario = nomina.funcionario
    config = ConfiguracionGeneral.obtener()
    icl_activo_nomina = icl_activo_para_funcionario(funcionario)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TituloNomina",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    ))

    elementos = []

    empresa_pdf = obtener_empresa_documento(funcionario=nomina.funcionario)

    elementos += construir_encabezado_empresa_pdf(empresa_pdf, "EXTRACTO DE NÓMINA")
    elementos.append(Spacer(1, 8))

    datos = [
        ["Funcionario", funcionario.nombre_completo],
        ["Documento", funcionario.documento_compacto],
        ["Empresa", funcionario.empresa_mostrar],
        ["Sucursal", funcionario.sucursal_mostrar],
        ["Cargo", funcionario.cargo or "-"],
        ["Período", f"{nomina.mes:02d}/{nomina.anio}"],
        ["Estado", nomina.get_estado_pago_display()],
        ["Fecha de pago", nomina.fecha_pago.strftime("%d/%m/%Y") if nomina.fecha_pago else "-"],
        ["Modalidad de cobro", nomina.modalidad_cobro or "-"],
        ["Banco", nomina.banco or "-"],
        ["Cuenta", f"{nomina.tipo_cuenta or '-'} / {nomina.numero_cuenta or '-'}"],
    ]

    tabla_datos = Table(datos, colWidths=[55 * mm, 105 * mm])
    tabla_datos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 12))

    tabla_liquidacion = Table([
        ["Concepto", "Monto"],
        ["Salario base", _gs(nomina.salario_base)],
        ["Bono base configurado", _gs(nomina.bono_base)],
        ["Bono pagado según ICL" if icl_activo_nomina else "Bono íntegro sin impacto ICL", _gs(nomina.bono_icl)],
        ["Salario bruto", _gs(nomina.salario_bruto)],
        ["IPS", f"- {_gs(nomina.descuento_ips)}"],
        ["Deudas", f"- {_gs(nomina.descuento_deudas)}"],
        ["NETO FINAL A COBRAR", _gs(nomina.salario_neto)],
    ], colWidths=[110 * mm, 50 * mm])

    tabla_liquidacion.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dcfce7")),
        ("TEXTCOLOR", (0, -1), (-1, -1), colors.HexColor("#166534")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabla_liquidacion)
    elementos.append(Spacer(1, 26))

    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma responsable", "Firma funcionario"],
    ], colWidths=[80 * mm, 80 * mm])

    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="NOMINA",
        documento_id=nomina.id,
        funcionario=funcionario,
        titulo="Extracto de Nómina",
    )

    agregar_texto_legal_empresa_pdf(elementos, empresa_pdf)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="extracto_nomina_{funcionario.cedula}_{nomina.mes:02d}_{nomina.anio}.pdf"'
    response.write(pdf)
    return response


@login_required
def nomina_sucursal_pdf(request):
    permiso = validar_permiso_o_redirigir(request, "nomina", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    sucursal_id = request.GET.get("sucursal", "").strip()

    if not sucursal_id:
        messages.error(request, "Debes seleccionar una sucursal para generar el extracto general.")
        return redirect(f"/nomina/?mes={mes}&anio={anio}")

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    sucursal = get_object_or_404(Sucursal.objects.select_related("empresa"), pk=sucursal_id)

    if not admin_master and sucursal.empresa != empresa_usuario:
        messages.error(request, "No puedes exportar nóminas de otra empresa.")
        return redirect("nomina_lista")

    nominas = NominaMensual.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).filter(
        mes=mes,
        anio=anio,
        funcionario__sucursal_rel=sucursal
    ).order_by("funcionario__apellido", "funcionario__nombre")

    total_bruto = nominas.aggregate(total=Sum("salario_bruto"))["total"] or Decimal("0.00")
    total_ips = nominas.aggregate(total=Sum("descuento_ips"))["total"] or Decimal("0.00")
    total_deudas = nominas.aggregate(total=Sum("descuento_deudas"))["total"] or Decimal("0.00")
    total_neto = nominas.aggregate(total=Sum("salario_neto"))["total"] or Decimal("0.00")
    total_bono = nominas.aggregate(total=Sum("bono_icl"))["total"] or Decimal("0.00")
    icl_activo_sucursal = bool(getattr(sucursal.empresa, "icl_activo", True))

    config = ConfiguracionGeneral.obtener()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="TituloSucursal",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=16,
        leading=20,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    ))

    elementos = []

    elementos.append(Paragraph(config.nombre_sistema or "ClockIn", styles["TituloSucursal"]))
    elementos.append(Paragraph("EXTRACTO GENERAL DE NÓMINA POR SUCURSAL", styles["TituloSucursal"]))
    elementos.append(Spacer(1, 8))

    resumen = Table([
        ["Empresa", sucursal.empresa.nombre],
        ["Sucursal", sucursal.nombre],
        ["Período", f"{mes:02d}/{anio}"],
        ["Cantidad de funcionarios", str(nominas.count())],
        ["Total bruto", _gs(total_bruto)],
        ["Total bono ICL" if icl_activo_sucursal else "Total bono íntegro", _gs(total_bono)],
        ["Total IPS", _gs(total_ips)],
        ["Total deudas", _gs(total_deudas)],
        ["Total neto general", _gs(total_neto)],
    ], colWidths=[65 * mm, 105 * mm])

    resumen.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
    ]))
    elementos.append(resumen)
    elementos.append(Spacer(1, 12))

    data = [["Funcionario", "CI", "Bruto", "IPS", "Deudas", "Neto", "Estado"]]

    for n in nominas:
        data.append([
            n.funcionario.nombre_completo,
            n.funcionario.documento_compacto,
            _gs(n.salario_bruto),
            _gs(n.descuento_ips),
            _gs(n.descuento_deudas),
            _gs(n.salario_neto),
            n.get_estado_pago_display(),
        ])

    tabla = Table(data, colWidths=[45 * mm, 22 * mm, 24 * mm, 22 * mm, 24 * mm, 26 * mm, 24 * mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("ALIGN", (2, 1), (5, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabla)

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="nomina_sucursal_{sucursal.id}_{mes:02d}_{anio}.pdf"'
    response.write(pdf)
    return response


@login_required
def nomina_sectores_pdf(request):
    if not tiene_permiso(request.user, "nomina", "puede_exportar"):
        messages.error(request, "No tienes permiso para exportar nominas.")
        return redirect("nomina_lista")

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    sucursal_id = request.GET.get("sucursal", "").strip()
    estado = request.GET.get("estado", "").strip()
    sectores = request.GET.getlist("sector")

    nominas = _nominas_filtradas_para_exportar(request, mes, anio, sucursal_id, sectores, estado)
    sectores_exportar = sectores or _sectores_desde_nominas(nominas)

    if not sectores_exportar:
        messages.error(request, "No hay sectores con nomina para exportar.")
        return redirect(f"/nomina/?mes={mes}&anio={anio}")

    if len(sectores_exportar) == 1:
        sector = sectores_exportar[0]
        nominas_sector = nominas.filter(funcionario__sector=sector)
        pdf = _pdf_nomina_sector_bytes(request, nominas_sector, sector, mes, anio)
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="nomina_{_filename_seguro(sector)}_{mes:02d}_{anio}.pdf"'
        response.write(pdf)
        registrar_historial(request, "Nomina", "Exportar PDF por sector", f"Se exporto PDF de nomina del sector {sector or 'Sin sector'} {mes:02d}/{anio}.")
        return response

    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as zip_file:
        for sector in sectores_exportar:
            nominas_sector = nominas.filter(funcionario__sector=sector)
            if not nominas_sector.exists():
                continue
            pdf = _pdf_nomina_sector_bytes(request, nominas_sector, sector, mes, anio)
            zip_file.writestr(f"{_filename_seguro(sector)}.pdf", pdf)

    registrar_historial(request, "Nomina", "Exportar ZIP PDF por sector", f"Se exporto ZIP PDF por sectores {mes:02d}/{anio}.")
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="nomina_{mes:02d}_{anio}_por_sectores_pdf.zip"'
    return response


def _xlsx_nomina_sector_bytes(nominas, sector_nombre, mes, anio):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    sector_label = sector_nombre or "Sin sector"
    primera = nominas.first()
    empresa_nombre = primera.funcionario.empresa_mostrar if primera else "-"
    sucursal_nombre = primera.funcionario.sucursal_mostrar if primera else "-"
    totales = _totales_nomina_queryset(nominas)

    wb = Workbook()
    ws = wb.active
    ws.title = "Nomina"

    rows = [
        ["NOMINA POR SECTOR"],
        ["Empresa", empresa_nombre],
        ["Sucursal", sucursal_nombre],
        ["Sector", sector_label],
        ["Periodo", f"{mes:02d}/{anio}"],
        ["Fecha generacion", timezone.localtime().strftime("%d/%m/%Y %H:%M")],
        [],
        ["Funcionario", "Documento", "Cargo", "Modalidad salarial", "Total haberes", "IPS", "Otros descuentos", "Ajustes", "Salario neto"],
    ]

    for row in rows:
        ws.append(row)

    for n in nominas:
        modalidad = getattr(n.funcionario, "modalidad_salarial_display", "Salario base + bono")
        ws.append([
            n.funcionario.nombre_completo,
            n.funcionario.documento_compacto,
            n.funcionario.cargo or "-",
            modalidad,
            float(n.salario_bruto or 0),
            float(n.descuento_ips or 0),
            float(n.descuento_deudas or 0),
            0,
            float(n.salario_neto or 0),
        ])

    ws.append([])
    ws.append(["Resumen", "", "", "", float(totales["total_haberes"]), float(totales["total_ips"]), float(totales["total_deudas"]), float(totales["total_ajustes"]), float(totales["total_neto"])])
    ws.append(["Cantidad funcionarios", totales["cantidad"]])

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[8]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws["A1"].font = Font(bold=True, size=16)
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["D"].width = 24

    for row in ws.iter_rows(min_row=9, min_col=5, max_col=9):
        for cell in row:
            cell.number_format = '#,##0'

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@login_required
def nomina_sectores_excel(request):
    if not tiene_permiso(request.user, "nomina", "puede_exportar"):
        messages.error(request, "No tienes permiso para exportar nominas.")
        return redirect("nomina_lista")

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))
    sucursal_id = request.GET.get("sucursal", "").strip()
    estado = request.GET.get("estado", "").strip()
    sectores = request.GET.getlist("sector")

    nominas = _nominas_filtradas_para_exportar(request, mes, anio, sucursal_id, sectores, estado)
    sectores_exportar = sectores or _sectores_desde_nominas(nominas)

    if not sectores_exportar:
        messages.error(request, "No hay sectores con nomina para exportar.")
        return redirect(f"/nomina/?mes={mes}&anio={anio}")

    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    if len(sectores_exportar) == 1:
        sector = sectores_exportar[0]
        nominas_sector = nominas.filter(funcionario__sector=sector)
        xlsx = _xlsx_nomina_sector_bytes(nominas_sector, sector, mes, anio)
        registrar_historial(request, "Nomina", "Exportar Excel por sector", f"Se exporto Excel de nomina del sector {sector or 'Sin sector'} {mes:02d}/{anio}.")
        response = HttpResponse(xlsx, content_type=content_type)
        response["Content-Disposition"] = f'attachment; filename="nomina_{_filename_seguro(sector)}_{mes:02d}_{anio}.xlsx"'
        return response

    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as zip_file:
        for sector in sectores_exportar:
            nominas_sector = nominas.filter(funcionario__sector=sector)
            if not nominas_sector.exists():
                continue
            xlsx = _xlsx_nomina_sector_bytes(nominas_sector, sector, mes, anio)
            zip_file.writestr(f"{_filename_seguro(sector)}.xlsx", xlsx)

    registrar_historial(request, "Nomina", "Exportar ZIP Excel por sector", f"Se exporto ZIP Excel por sectores {mes:02d}/{anio}.")
    response = HttpResponse(buffer.getvalue(), content_type="application/zip")
    response["Content-Disposition"] = f'attachment; filename="nomina_{mes:02d}_{anio}_por_sectores_excel.zip"'
    return response


@login_required
def configuracion_general(request):
    permiso = validar_permiso_o_redirigir(request, "configuracion", "puede_ver")
    if permiso:
        return permiso
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    config = ConfiguracionGeneral.obtener()

    mapa_hex_a_tema = {
        "#2563eb": ConfiguracionGeneral.TEMA_AZUL,
        "#16a34a": ConfiguracionGeneral.TEMA_VERDE,
        "#dc2626": ConfiguracionGeneral.TEMA_ROJO,
        "#ea580c": ConfiguracionGeneral.TEMA_NARANJA,
        "#7c3aed": ConfiguracionGeneral.TEMA_MORADO,
        "#0891b2": ConfiguracionGeneral.TEMA_TURQUESA,
        "#475569": ConfiguracionGeneral.TEMA_GRIS,
    }

    valores_validos = {item[0] for item in ConfiguracionGeneral.TEMAS_CHOICES}

    if config.color_primario in mapa_hex_a_tema:
        config.color_primario = mapa_hex_a_tema[config.color_primario]
        config.save(update_fields=["color_primario"])
    elif config.color_primario not in valores_validos:
        config.color_primario = ConfiguracionGeneral.TEMA_AZUL
        config.save(update_fields=["color_primario"])

    if request.method == "POST":
        permiso_post = validar_permiso_o_redirigir(request, "configuracion", "puede_editar")
        if permiso_post:
            return permiso_post

        post_data = request.POST.copy()

        for campo in [
            "bancos_personalizados",
            "cargos_personalizados",
            "sectores_personalizados",
        ]:
            valor = post_data.get(campo, "")
            items = []
            for linea in valor.splitlines():
                item = linea.strip()
                if item and item not in items:
                    items.append(item)
            post_data[campo] = "\n".join(items)

        form = ConfiguracionGeneralForm(post_data, instance=config)

        if form.is_valid():
            config = form.save()

            funcionarios_actualizables = Funcionario.objects.all()
            empresa_usuario = obtener_empresa_activa(request)
            if empresa_usuario:
                funcionarios_actualizables = funcionarios_actualizables.filter(sucursal_rel__empresa=empresa_usuario)

            funcionarios_actualizables.update(
                salario_base=config.salario_base_default,
                porcentaje_limite_deuda=config.porcentaje_limite_deuda_default,
            )

            registrar_historial(
                request,
                "Configuraciones",
                "Editar",
                f"Configuración PRO Plus actualizada. "
                f"Salario base: {config.salario_base_default}, "
                f"Límite deuda: {config.porcentaje_limite_deuda_default}%, "
                f"Tolerancia: {config.tolerancia_minutos_default} min, "
                f"Lectura biométrica: {config.biometrico_segundos_lectura}s, "
                f"Tema: {config.color_primario}."
            )

            messages.success(request, "Configuración PRO Plus actualizada correctamente.")
            return redirect("configuracion_general")

        messages.error(request, "No se pudo guardar la configuración. Revisa los campos marcados.")

    else:
        form = ConfiguracionGeneralForm(instance=config)

    return render(request, "core/configuracion_general.html", {
        "form": form,
        "config": config,
    })


@login_required
def suscripcion_panel(request):
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    suscripcion = SuscripcionSistema.obtener()
    pago_form = PagoSuscripcionSistemaForm(initial={"fecha_pago": timezone.localdate()})
    pagos = suscripcion.pagos.select_related("registrado_por")[:12]

    if request.method == "POST":
        form = SuscripcionSistemaForm(request.POST, instance=suscripcion)
        if form.is_valid():
            suscripcion = form.save(commit=False)
            suscripcion.actualizado_por = request.user
            suscripcion.save()

            registrar_historial(
                request,
                "Suscripciones",
                "Actualizar",
                f"Se actualizo la suscripcion. Proximo pago: {suscripcion.fecha_proximo_pago}."
            )

            messages.success(request, "Suscripcion actualizada correctamente.")
            return redirect("suscripcion_panel")
            messages.error(request, "No se pudo guardar la suscripcion. Revisa los campos marcados.")
    else:
        form = SuscripcionSistemaForm(instance=suscripcion)

    return render(request, "core/suscripcion_panel.html", {
        "form": form,
        "pago_form": pago_form,
        "pagos": pagos,
        "suscripcion": suscripcion,
    })


@login_required
def suscripcion_registrar_pago(request):
    bloqueo = _bloquear_si_no_admin_master(request)
    if bloqueo:
        return bloqueo

    suscripcion = SuscripcionSistema.obtener()

    if request.method != "POST":
        return redirect("suscripcion_panel")

    form = PagoSuscripcionSistemaForm(request.POST)
    if not form.is_valid():
        messages.error(request, "No se pudo registrar el pago. Revisa los campos marcados.")
        return redirect("suscripcion_panel")

    pago = form.save(commit=False)
    pago.suscripcion = suscripcion
    pago.registrado_por = request.user
    pago.save()

    base_fecha = suscripcion.fecha_proximo_pago
    if base_fecha < pago.fecha_pago:
        base_fecha = pago.fecha_pago

    suscripcion.fecha_ultimo_pago = pago.fecha_pago
    suscripcion.fecha_proximo_pago = sumar_meses(base_fecha, int(pago.meses_cubiertos or 1))
    suscripcion.estado = SuscripcionSistema.Estados.ACTIVA
    suscripcion.actualizado_por = request.user
    suscripcion.save(update_fields=[
        "fecha_ultimo_pago",
        "fecha_proximo_pago",
        "estado",
        "actualizado_por",
        "actualizado_en",
    ])

    registrar_historial(
        request,
        "Suscripciones",
        "Registrar pago",
        f"Se registro pago de suscripcion por {pago.meses_cubiertos} mes(es). Proximo pago: {suscripcion.fecha_proximo_pago}."
    )

    messages.success(request, "Pago registrado y proximo vencimiento actualizado.")
    return redirect("suscripcion_panel")


@login_required
def suscripcion_bloqueada(request):
    suscripcion = SuscripcionSistema.obtener()

    if es_admin_master(request.user):
        return redirect("suscripcion_panel")

    if not suscripcion.bloqueada:
        return redirect("dashboard")

    return render(request, "core/suscripcion_bloqueada.html", {
        "suscripcion": suscripcion,
    })

from decimal import Decimal, ROUND_HALF_UP

def calcular_liquidacion_funcionario_legacy(
    funcionario,
    tipo_salida,
    fecha_salida,
    dias_trabajados_pendientes=None,
    vacaciones_causadas_pendientes_dias=None,
    preaviso_dias_otorgados=0,
    preaviso_cumplido=False,
    descontar_preaviso=False,
    otros_descuentos=Decimal("0"),
):
    from decimal import Decimal
    from datetime import date
    from calendar import monthrange
    from django.apps import apps

    def D(valor):
        try:
            return Decimal(str(valor or 0))
        except Exception:
            return Decimal("0")

    salario_base = D(getattr(funcionario, "salario_base", 0))
    bono_base = D(getattr(funcionario, "bono", 0))
    salario_total = salario_base + bono_base

    salario_diario = salario_total / Decimal("30") if salario_total > 0 else Decimal("0")

    fecha_ingreso = getattr(funcionario, "fecha_ingreso", None)

    antiguedad_anios = 0
    antiguedad_meses = 0
    antiguedad_dias = 0

    if fecha_ingreso and fecha_salida:
        antiguedad_anios = fecha_salida.year - fecha_ingreso.year
        antiguedad_meses = fecha_salida.month - fecha_ingreso.month
        antiguedad_dias = fecha_salida.day - fecha_ingreso.day

        if antiguedad_dias < 0:
            antiguedad_meses -= 1
            mes_anterior = fecha_salida.month - 1 or 12
            anio_mes_anterior = fecha_salida.year if fecha_salida.month > 1 else fecha_salida.year - 1
            antiguedad_dias += monthrange(anio_mes_anterior, mes_anterior)[1]

        if antiguedad_meses < 0:
            antiguedad_anios -= 1
            antiguedad_meses += 12

        if antiguedad_anios < 0:
            antiguedad_anios = 0
            antiguedad_meses = 0
            antiguedad_dias = 0

    if dias_trabajados_pendientes is None:
        dias_trabajados_pendientes = fecha_salida.day if fecha_salida else 0

    try:
        dias_trabajados_pendientes = int(dias_trabajados_pendientes or 0)
    except Exception:
        dias_trabajados_pendientes = 0

    # Salario pendiente del mes trabajado
    # Ejemplo: salario 2.899.048 / 30 * 5 días = 483.175
    dias_trabajados_pendientes = int(dias_trabajados_pendientes or 0)

    salario_diario = salario_base / Decimal("30")

    salario_pendiente_monto = (
        salario_diario * Decimal(dias_trabajados_pendientes)
    ).quantize(Decimal("1"), rounding=ROUND_HALF_UP)

    ausencias_descuento = 0

    try:
        Asistencia = apps.get_model("core", "Asistencia")

        fecha_inicio_mes = date(fecha_salida.year, fecha_salida.month, 1)

        asistencias_qs = Asistencia.objects.filter(
            funcionario=funcionario,
            fecha__gte=fecha_inicio_mes,
            fecha__lte=fecha_salida,
        )

        fechas_asistidas = set(asistencias_qs.values_list("fecha", flat=True))

        dia_actual = fecha_inicio_mes
        while dia_actual <= fecha_salida:
            if dia_actual.weekday() != 6:
                if dia_actual not in fechas_asistidas:
                    ausencias_descuento += 1

            dia_actual = dia_actual.replace(day=dia_actual.day + 1) if dia_actual.day < monthrange(dia_actual.year, dia_actual.month)[1] else None

            if dia_actual is None:
                break

    except Exception:
        ausencias_descuento = 0

    descuento_ausencias = salario_diario * Decimal(ausencias_descuento)

    if descuento_ausencias > 0:
        salario_pendiente_monto -= descuento_ausencias

    if salario_pendiente_monto < 0:
        salario_pendiente_monto = Decimal("0")

    if vacaciones_causadas_pendientes_dias is None:
        vacaciones_causadas_pendientes_dias = 0

    try:
        vacaciones_causadas_pendientes_dias = int(vacaciones_causadas_pendientes_dias or 0)
    except Exception:
        vacaciones_causadas_pendientes_dias = 0

    vacaciones_causadas_monto = salario_diario * Decimal(vacaciones_causadas_pendientes_dias)

    tipo_normalizado = str(tipo_salida or "").lower()

    if tipo_normalizado == "renuncia":
        vacaciones_proporcionales_dias = 0
        vacaciones_proporcionales_monto = Decimal("0")
    else:
        meses_trabajados_anio = fecha_salida.month if fecha_salida else 0

        if antiguedad_anios < 5:
            dias_vacaciones_anual = 12
        elif antiguedad_anios < 10:
            dias_vacaciones_anual = 18
        else:
            dias_vacaciones_anual = 30

        vacaciones_proporcionales_dias = round((dias_vacaciones_anual / 12) * meses_trabajados_anio)
        vacaciones_proporcionales_monto = salario_diario * Decimal(vacaciones_proporcionales_dias)

    aguinaldo_proporcional_monto = Decimal("0")

    if fecha_salida:
        meses_aguinaldo = fecha_salida.month
        aguinaldo_proporcional_monto = (salario_total * Decimal(meses_aguinaldo)) / Decimal("12")

    preaviso_dias_corresponde = 0
    preaviso_monto = Decimal("0")

    if tipo_normalizado in ["despido_sin_causa", "despido_sin_justa_causa"]:
        if antiguedad_anios < 1:
            preaviso_dias_corresponde = 30
        elif antiguedad_anios < 5:
            preaviso_dias_corresponde = 45
        elif antiguedad_anios < 10:
            preaviso_dias_corresponde = 60
        else:
            preaviso_dias_corresponde = 90

        if not preaviso_cumplido:
            dias_preaviso_pagar = max(preaviso_dias_corresponde - int(preaviso_dias_otorgados or 0), 0)
            preaviso_monto = salario_diario * Decimal(dias_preaviso_pagar)

    if descontar_preaviso and tipo_normalizado == "renuncia":
        preaviso_monto = Decimal("0")

    indemnizacion_dias = 0
    indemnizacion_monto = Decimal("0")

    if tipo_normalizado in ["despido_sin_causa", "despido_sin_justa_causa"]:
        indemnizacion_dias = antiguedad_anios * 15
        indemnizacion_monto = salario_diario * Decimal(indemnizacion_dias)

    ips_monto = Decimal("0")

    if getattr(funcionario, "ips", False):
        ips_monto = salario_pendiente_monto * Decimal("0.09")

    deudas_monto = Decimal("0")

    try:
        Deuda = apps.get_model("core", "Deuda")

        deudas_qs = Deuda.objects.filter(funcionario=funcionario)

        if hasattr(Deuda, "activa"):
            deudas_qs = deudas_qs.filter(activa=True)

        for deuda in deudas_qs:
            deudas_monto += D(getattr(deuda, "saldo_pendiente", None) or getattr(deuda, "monto", 0))

    except Exception:
        deudas_monto = Decimal("0")

    otros_descuentos = D(otros_descuentos)

    total_haberes = (
        salario_pendiente_monto
        + vacaciones_causadas_monto
        + vacaciones_proporcionales_monto
        + aguinaldo_proporcional_monto
        + preaviso_monto
        + indemnizacion_monto
    )

    total_descuentos = ips_monto + deudas_monto + otros_descuentos

    total_liquidacion = total_haberes - total_descuentos

    if total_liquidacion < 0:
        total_liquidacion = Decimal("0")

    requiere_revision_juridica = False
    alerta_revision = ""

    if tipo_normalizado in ["despido_justa_causa", "despido_por_justa_causa", "periodo_prueba", "abandono"]:
        requiere_revision_juridica = True
        alerta_revision = "Este tipo de salida requiere revisión jurídica antes de confirmar la liquidación."

    return {
        "antiguedad_anios": antiguedad_anios,
        "antiguedad_meses": antiguedad_meses,
        "antiguedad_dias": antiguedad_dias,

        "salario_base_snapshot": salario_base,
        "bono_base_snapshot": bono_base,

        "dias_trabajados_pendientes": dias_trabajados_pendientes,
        "salario_pendiente_monto": salario_pendiente_monto,

        "ausencias_descuento": ausencias_descuento,
        "descuento_ausencias": descuento_ausencias,

        "vacaciones_causadas_pendientes_dias": vacaciones_causadas_pendientes_dias,
        "vacaciones_causadas_monto": vacaciones_causadas_monto,

        "vacaciones_proporcionales_dias": vacaciones_proporcionales_dias,
        "vacaciones_proporcionales_monto": vacaciones_proporcionales_monto,

        "aguinaldo_proporcional_monto": aguinaldo_proporcional_monto,

        "preaviso_dias_corresponde": preaviso_dias_corresponde,
        "preaviso_dias_otorgados": preaviso_dias_otorgados,
        "preaviso_cumplido": preaviso_cumplido,
        "preaviso_monto": preaviso_monto,

        "indemnizacion_dias": indemnizacion_dias,
        "indemnizacion_monto": indemnizacion_monto,

        "ips_monto": ips_monto,
        "deudas_monto": deudas_monto,
        "otros_descuentos": otros_descuentos,

        "total_haberes": total_haberes,
        "total_descuentos": total_descuentos,
        "total_liquidacion": total_liquidacion,

        "requiere_revision_juridica": requiere_revision_juridica,
        "alerta_revision": alerta_revision,
    }

@login_required
def liquidaciones_lista(request):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidaciones = Liquidacion.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa"
    ).all()

    if not admin_master:
        if empresa_usuario:
            liquidaciones = liquidaciones.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            liquidaciones = liquidaciones.none()

    if q:
        liquidaciones = liquidaciones.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q) |
            Q(tipo_salida__icontains=q) |
            Q(estado__icontains=q)
        )

    return render(request, "core/liquidaciones_lista.html", {
        "liquidaciones": liquidaciones,
        "q": q,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })

def _liq_dias_salario_auto(fecha_salida):
    if not fecha_salida:
        return 0
    return int(fecha_salida.day)


def _liq_vacaciones_pendientes_auto(funcionario):
    if not funcionario:
        return 0

    try:
        saldo = funcionario.saldo_vacaciones
        return max(int(saldo or 0), 0)
    except Exception:
        return 0


def _liq_int_auto(valor, default):
    try:
        valor_int = int(valor or 0)
    except Exception:
        valor_int = 0

    if valor_int <= 0:
        return default

    return valor_int


def _liq_alertas_funcionario(funcionario):
    alertas = []

    try:
        if funcionario.total_deuda_activa and funcionario.total_deuda_activa > 0:
            alertas.append(f"Tiene deuda activa: Gs. {int(funcionario.total_deuda_activa):,}".replace(",", "."))
    except Exception:
        pass

    try:
        saldo_vac = funcionario.saldo_vacaciones
        if saldo_vac and saldo_vac > 0:
            alertas.append(f"Tiene {saldo_vac} día(s) de vacaciones pendientes.")
    except Exception:
        pass

    try:
        if funcionario.conductas.exists():
            alertas.append("Tiene historial de conducta registrado.")
    except Exception:
        pass

    try:
        if not funcionario.fecha_ingreso:
            alertas.append("No tiene fecha de ingreso cargada.")
    except Exception:
        pass

    return alertas


def _liq_alertas_proteccion_guardadas(liquidacion):
    resumen = {
        "ausencias_descuento": liquidacion.ausencias_descuento,
        "preaviso_dias_corresponde": liquidacion.preaviso_dias_corresponde,
        "preaviso_dias_otorgados": liquidacion.preaviso_dias_otorgados,
        "preaviso_dias_faltantes": max(
            int(liquidacion.preaviso_dias_corresponde or 0) - int(liquidacion.preaviso_dias_otorgados or 0),
            0,
        ),
        "deudas_monto": liquidacion.deudas_monto,
        "otros_descuentos": liquidacion.otros_descuentos,
        "vacaciones_causadas_pendientes_dias": liquidacion.vacaciones_causadas_pendientes_dias,
        "indemnizacion_suma_anio_por_fraccion": (
            liquidacion.tipo_salida == Liquidacion.TiposSalida.DESPIDO_SIN_JUSTA_CAUSA
            and (
                liquidacion.antiguedad_meses > 6
                or (liquidacion.antiguedad_meses == 6 and liquidacion.antiguedad_dias > 0)
            )
        ),
    }
    return construir_alertas_proteccion_liquidacion(resumen, liquidacion.tipo_salida)


def _liquidacion_empresa(liquidacion):
    return liquidacion.empresa or getattr(liquidacion.funcionario, "empresa", None)


def _usuario_puede_ajustar_liquidacion(request, liquidacion):
    empresa = _liquidacion_empresa(liquidacion)
    return bool(
        empresa
        and getattr(empresa, "permite_ajuste_manual_liquidacion", False)
        and liquidacion.estado == Liquidacion.Estados.BORRADOR
        and tiene_permiso(request.user, "liquidacion", "puede_ajustar")
    )


def _recalcular_totales_liquidacion_con_ajustes(liquidacion):
    ajustes = liquidacion.ajustes_manuales.filter(
        estado=AjusteManualLiquidacion.Estados.ACTIVO
    )

    ajuste_haberes = Decimal("0.00")
    ajuste_descuentos = Decimal("0.00")

    for ajuste in ajustes:
        if ajuste.tipo == AjusteManualLiquidacion.Tipos.HABER:
            ajuste_haberes += Decimal(ajuste.diferencia or 0)
        elif ajuste.tipo == AjusteManualLiquidacion.Tipos.DESCUENTO:
            ajuste_descuentos += Decimal(ajuste.diferencia or 0)

    liquidacion.total_haberes = (
        Decimal(liquidacion.total_haberes_automatico or 0) + ajuste_haberes
    ).quantize(Decimal("0.01"))
    liquidacion.total_descuentos = (
        Decimal(liquidacion.total_descuentos_automatico or 0) + ajuste_descuentos
    ).quantize(Decimal("0.01"))
    liquidacion.total_liquidacion = (
        liquidacion.total_haberes - liquidacion.total_descuentos
    ).quantize(Decimal("0.01"))

    if liquidacion.total_haberes < 0:
        liquidacion.total_haberes = Decimal("0.00")
    if liquidacion.total_descuentos < 0:
        liquidacion.total_descuentos = Decimal("0.00")
    if liquidacion.total_liquidacion < 0:
        liquidacion.total_liquidacion = Decimal("0.00")

    liquidacion.save(update_fields=[
        "total_haberes",
        "total_descuentos",
        "total_liquidacion",
        "actualizado_en",
    ])


@login_required
def liquidacion_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    resumen = None

    if request.method == "POST":
        form = LiquidacionForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            liquidacion = form.save(commit=False)

            if not admin_master:
                if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear liquidaciones para otra empresa.")
                    return redirect("liquidaciones_lista")

            dias_auto = _liq_dias_salario_auto(liquidacion.fecha_salida)
            vacaciones_auto = _liq_vacaciones_pendientes_auto(liquidacion.funcionario)

            liquidacion.dias_trabajados_pendientes = _liq_int_auto(
                liquidacion.dias_trabajados_pendientes,
                dias_auto
            )

            liquidacion.vacaciones_causadas_pendientes_dias = _liq_int_auto(
                liquidacion.vacaciones_causadas_pendientes_dias,
                vacaciones_auto
            )

            if not liquidacion.fecha_calculo:
                liquidacion.fecha_calculo = timezone.localdate()

            resumen = calcular_liquidacion_funcionario(
                funcionario=liquidacion.funcionario,
                tipo_salida=liquidacion.tipo_salida,
                fecha_salida=liquidacion.fecha_salida,
                dias_trabajados_pendientes=liquidacion.dias_trabajados_pendientes,
                vacaciones_causadas_pendientes_dias=liquidacion.vacaciones_causadas_pendientes_dias,
                preaviso_dias_otorgados=liquidacion.preaviso_dias_otorgados,
                preaviso_cumplido=liquidacion.preaviso_cumplido,
                descontar_preaviso=liquidacion.descontar_preaviso,
                otros_descuentos=liquidacion.otros_descuentos,
            )

            for campo, valor in resumen.items():
                setattr(liquidacion, campo, valor)

            liquidacion.save()

            registrar_historial(
                request,
                "Liquidaciones",
                "Crear",
                f"Se creó liquidación para {liquidacion.funcionario.nombre_completo} - {liquidacion.get_tipo_salida_display()}."
            )

            messages.success(request, "Liquidación generada correctamente.")
            return redirect("liquidacion_detalle", pk=liquidacion.pk)

    else:
        form = LiquidacionForm(initial={"fecha_calculo": timezone.localdate()})

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/liquidacion_form.html", {
        "form": form,
        "resumen": resumen,
        "titulo_form": "Nueva liquidación",
    })


@login_required
def liquidacion_preview(request):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_crear")
    if permiso:
        return JsonResponse({"ok": False, "error": "Sin permiso."}, status=403)

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionario_id = request.GET.get("funcionario")
    tipo_salida = request.GET.get("tipo_salida")
    fecha_salida = request.GET.get("fecha_salida")

    dias_trabajados_pendientes = request.GET.get("dias_trabajados_pendientes")
    vacaciones_causadas_pendientes_dias = request.GET.get("vacaciones_causadas_pendientes_dias")
    preaviso_dias_otorgados = request.GET.get("preaviso_dias_otorgados", "0")
    preaviso_cumplido = request.GET.get("preaviso_cumplido") == "true"
    descontar_preaviso = request.GET.get("descontar_preaviso") == "true"
    otros_descuentos = request.GET.get("otros_descuentos", "0")

    if not funcionario_id or not tipo_salida or not fecha_salida:
        return JsonResponse({"ok": False, "error": "Faltan datos para calcular."})

    try:
        funcionario = Funcionario.objects.select_related(
            "sucursal_rel",
            "sucursal_rel__empresa",
            "turno"
        ).get(pk=funcionario_id, activo=True)
    except Funcionario.DoesNotExist:
        return JsonResponse({"ok": False, "error": "Funcionario no encontrado."})

    if not admin_master:
        if not funcionario.sucursal_rel or funcionario.sucursal_rel.empresa != empresa_usuario:
            return JsonResponse({"ok": False, "error": "No puedes calcular liquidación para otra empresa."}, status=403)

    try:
        fecha_salida_obj = datetime.strptime(fecha_salida, "%Y-%m-%d").date()
    except ValueError:
        return JsonResponse({"ok": False, "error": "Fecha de salida inválida."})

    try:
        dias_auto = _liq_dias_salario_auto(fecha_salida_obj)
        vacaciones_auto = _liq_vacaciones_pendientes_auto(funcionario)

        dias_trabajados_pendientes = _liq_int_auto(
            dias_trabajados_pendientes,
            dias_auto
        )

        vacaciones_causadas_pendientes_dias = _liq_int_auto(
            vacaciones_causadas_pendientes_dias,
            vacaciones_auto
        )

        preaviso_dias_otorgados = int(preaviso_dias_otorgados or 0)
        otros_descuentos = Decimal(str(otros_descuentos or 0))
    except (ValueError, InvalidOperation):
        return JsonResponse({"ok": False, "error": "Hay valores numéricos inválidos."})

    resumen = calcular_liquidacion_funcionario(
        funcionario=funcionario,
        tipo_salida=tipo_salida,
        fecha_salida=fecha_salida_obj,
        dias_trabajados_pendientes=dias_trabajados_pendientes,
        vacaciones_causadas_pendientes_dias=vacaciones_causadas_pendientes_dias,
        preaviso_dias_otorgados=preaviso_dias_otorgados,
        preaviso_cumplido=preaviso_cumplido,
        descontar_preaviso=descontar_preaviso,
        otros_descuentos=otros_descuentos,
    )

    liquidaciones_previas = Liquidacion.objects.filter(
        funcionario=funcionario
    ).order_by("-fecha_salida")[:5]

    alertas = _liq_alertas_funcionario(funcionario)

    return JsonResponse({
        "ok": True,
        "funcionario": funcionario.nombre_completo,
        "cedula": funcionario.cedula,
        "tipo_documento": funcionario.documento_etiqueta,
        "documento_formateado": funcionario.documento_formateado,
        "documento_compacto": funcionario.documento_compacto,
        "cargo": funcionario.cargo or "-",
        "sector": funcionario.sector or "-",
        "sucursal": funcionario.sucursal_mostrar,
        "empresa": funcionario.empresa_mostrar,
        "fecha_ingreso": funcionario.fecha_ingreso.strftime("%d/%m/%Y") if funcionario.fecha_ingreso else "-",
        "tipo_salida": tipo_salida,

        "auto": {
            "dias_trabajados_pendientes": dias_auto,
            "vacaciones_pendientes": vacaciones_auto,
            "fecha_calculo": timezone.localdate().strftime("%Y-%m-%d"),
        },

        "alertas": alertas,

        "liquidaciones_previas": [
            {
                "fecha": l.fecha_salida.strftime("%d/%m/%Y"),
                "tipo": l.get_tipo_salida_display(),
                "total": str(l.total_liquidacion),
                "estado": l.get_estado_display(),
            }
            for l in liquidaciones_previas
        ],

        "antiguedad": {
            "anios": resumen["antiguedad_anios"],
            "meses": resumen["antiguedad_meses"],
            "dias": resumen["antiguedad_dias"],
        },

        "modalidad_salarial_snapshot": resumen["modalidad_salarial_snapshot"],
        "salario_base_snapshot": str(resumen["salario_base_snapshot"]),
        "bono_base_snapshot": str(resumen["bono_base_snapshot"]),
        "salario_diferenciado_snapshot": str(resumen["salario_diferenciado_snapshot"]),
        "salario_bruto_aplicable_snapshot": str(resumen["salario_bruto_aplicable_snapshot"]),
        "porcentaje_ips_snapshot": str(resumen["porcentaje_ips_snapshot"]),
        "descuento_ips_calculado_snapshot": str(resumen["descuento_ips_calculado_snapshot"]),
        "dias_trabajados_pendientes": resumen["dias_trabajados_pendientes"],
        "salario_pendiente_monto": str(resumen["salario_pendiente_monto"]),

        "ausencias_descuento": resumen.get("ausencias_descuento", 0),
        "descuento_ausencias": str(resumen.get("descuento_ausencias", 0)),

        "vacaciones_causadas_pendientes_dias": resumen["vacaciones_causadas_pendientes_dias"],
        "vacaciones_causadas_monto": str(resumen["vacaciones_causadas_monto"]),
        "vacaciones_proporcionales_dias": resumen["vacaciones_proporcionales_dias"],
        "vacaciones_proporcionales_monto": str(resumen["vacaciones_proporcionales_monto"]),
        "aguinaldo_proporcional_monto": str(resumen["aguinaldo_proporcional_monto"]),
        "preaviso_dias_corresponde": resumen["preaviso_dias_corresponde"],
        "preaviso_dias_otorgados": resumen["preaviso_dias_otorgados"],
        "preaviso_dias_faltantes": resumen["preaviso_dias_faltantes"],
        "preaviso_cumplido": resumen["preaviso_cumplido"],
        "preaviso_monto": str(resumen["preaviso_monto"]),
        "indemnizacion_dias": resumen["indemnizacion_dias"],
        "indemnizacion_monto": str(resumen["indemnizacion_monto"]),
        "indemnizacion_anios_computables": resumen["indemnizacion_anios_computables"],
        "indemnizacion_suma_anio_por_fraccion": resumen["indemnizacion_suma_anio_por_fraccion"],
        "salario_indemnizacion_base": str(resumen["salario_indemnizacion_base"]),
        "salario_diario_base": str(resumen["salario_diario_base"]),
        "ips_monto": str(resumen["ips_monto"]),
        "deudas_monto": str(resumen["deudas_monto"]),
        "otros_descuentos": str(resumen["otros_descuentos"]),
        "total_haberes": str(resumen["total_haberes"]),
        "total_descuentos": str(resumen["total_descuentos"]),
        "total_liquidacion": str(resumen["total_liquidacion"]),
        "requiere_revision_juridica": resumen["requiere_revision_juridica"],
        "alerta_revision": resumen["alerta_revision"],
        "alertas_proteccion": resumen["alertas_proteccion"],
    })


@login_required
def liquidacion_detalle(request, pk):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes ver liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    alertas_proteccion = _liq_alertas_proteccion_guardadas(liquidacion)
    ajustes_manuales = liquidacion.ajustes_manuales.select_related(
        "creado_por",
        "anulado_por",
    ).all()
    ajustes_activos = [
        ajuste for ajuste in ajustes_manuales
        if ajuste.estado == AjusteManualLiquidacion.Estados.ACTIVO
    ]
    ajuste_haberes = sum(
        (Decimal(ajuste.diferencia or 0) for ajuste in ajustes_activos if ajuste.tipo == AjusteManualLiquidacion.Tipos.HABER),
        Decimal("0.00"),
    )
    ajuste_descuentos = sum(
        (Decimal(ajuste.diferencia or 0) for ajuste in ajustes_activos if ajuste.tipo == AjusteManualLiquidacion.Tipos.DESCUENTO),
        Decimal("0.00"),
    )

    return render(request, "core/liquidacion_detalle.html", {
        "liquidacion": liquidacion,
        "ajustes_manuales": ajustes_manuales,
        "ajuste_haberes": ajuste_haberes,
        "ajuste_descuentos": ajuste_descuentos,
        "puede_ajustar_liquidacion": _usuario_puede_ajustar_liquidacion(request, liquidacion),
        "alertas_proteccion": alertas_proteccion,
        "preaviso_dias_faltantes": max(
            int(liquidacion.preaviso_dias_corresponde or 0) - int(liquidacion.preaviso_dias_otorgados or 0),
            0,
        ),
    })


@login_required
def liquidacion_ajuste_manual(request, pk):
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes ajustar liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    empresa_liquidacion = _liquidacion_empresa(liquidacion)

    if not empresa_liquidacion or not getattr(empresa_liquidacion, "permite_ajuste_manual_liquidacion", False):
        messages.error(request, "La empresa no permite ajustes manuales de liquidaciones.")
        return redirect("liquidacion_detalle", pk=pk)

    if liquidacion.estado != Liquidacion.Estados.BORRADOR:
        messages.error(request, "Solo se pueden ajustar liquidaciones en estado borrador.")
        return redirect("liquidacion_detalle", pk=pk)

    if not tiene_permiso(request.user, "liquidacion", "puede_ajustar"):
        messages.error(request, "No tienes permiso para realizar ajustes manuales de liquidaciones.")
        return redirect("liquidacion_detalle", pk=pk)

    if request.method == "POST":
        form = AjusteManualLiquidacionForm(request.POST)
        if form.is_valid():
            ajuste = form.save(commit=False)
            ajuste.liquidacion = liquidacion
            ajuste.empresa = empresa_liquidacion
            ajuste.funcionario = liquidacion.funcionario
            ajuste.creado_por = request.user
            ajuste.save()

            _recalcular_totales_liquidacion_con_ajustes(liquidacion)

            registrar_historial(
                request,
                "Liquidaciones",
                "Ajuste manual",
                f"Se creó ajuste manual '{ajuste.concepto}' para {liquidacion.funcionario.nombre_completo}. Diferencia: {ajuste.diferencia}."
            )
            messages.success(request, "Ajuste manual aplicado correctamente.")
            return redirect("liquidacion_detalle", pk=pk)
    else:
        form = AjusteManualLiquidacionForm()

    return render(request, "core/liquidacion_ajuste_form.html", {
        "form": form,
        "liquidacion": liquidacion,
    })


@login_required
def liquidacion_ajuste_anular(request, pk, ajuste_id):
    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )
    ajuste = get_object_or_404(
        AjusteManualLiquidacion.objects.select_related("liquidacion"),
        pk=ajuste_id,
        liquidacion=liquidacion,
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes anular ajustes de otra empresa.")
            return redirect("liquidaciones_lista")

    if not _usuario_puede_ajustar_liquidacion(request, liquidacion):
        messages.error(request, "No tienes permiso o la liquidación no permite ajustes.")
        return redirect("liquidacion_detalle", pk=pk)

    if request.method == "POST":
        motivo = (request.POST.get("motivo_anulacion") or "").strip()
        if len(motivo) < 10:
            messages.error(request, "Debes indicar un motivo de anulación de al menos 10 caracteres.")
            return redirect("liquidacion_detalle", pk=pk)

        ajuste.estado = AjusteManualLiquidacion.Estados.ANULADO
        ajuste.motivo_anulacion = motivo
        ajuste.anulado_por = request.user
        ajuste.anulado_en = timezone.now()
        ajuste.save(update_fields=[
            "estado",
            "motivo_anulacion",
            "anulado_por",
            "anulado_en",
        ])

        _recalcular_totales_liquidacion_con_ajustes(liquidacion)

        registrar_historial(
            request,
            "Liquidaciones",
            "Anular ajuste",
            f"Se anuló ajuste manual '{ajuste.concepto}' de {liquidacion.funcionario.nombre_completo}. Motivo: {motivo}"
        )
        messages.success(request, "Ajuste manual anulado correctamente.")

    return redirect("liquidacion_detalle", pk=pk)


@login_required
def liquidacion_confirmar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_confirmar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes confirmar liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    if liquidacion.estado == Liquidacion.Estados.ANULADA:
        messages.error(request, "No puedes confirmar una liquidación anulada.")
        return redirect("liquidacion_detalle", pk=pk)

    liquidacion.estado = Liquidacion.Estados.CONFIRMADA
    liquidacion.save(update_fields=["estado", "actualizado_en"])

    registrar_historial(
        request,
        "Liquidaciones",
        "Confirmar",
        f"Se confirmó la liquidación de {liquidacion.funcionario.nombre_completo}."
    )

    messages.success(request, "Liquidación confirmada correctamente.")
    return redirect("liquidacion_detalle", pk=pk)


@login_required
def liquidacion_marcar_pagada(request, pk):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_pagar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    if liquidacion.estado == Liquidacion.Estados.ANULADA:
        messages.error(request, "No puedes marcar como pagada una liquidación anulada.")
        return redirect("liquidacion_detalle", pk=pk)

    liquidacion.estado = Liquidacion.Estados.PAGADA
    liquidacion.save(update_fields=["estado", "actualizado_en"])

    registrar_historial(
        request,
        "Liquidaciones",
        "Pagar",
        f"Se marcó como pagada la liquidación de {liquidacion.funcionario.nombre_completo}."
    )

    messages.success(request, "Liquidación marcada como pagada.")
    return redirect("liquidacion_detalle", pk=pk)


@login_required
def liquidacion_anular(request, pk):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_anular")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related("funcionario", "funcionario__sucursal_rel", "funcionario__sucursal_rel__empresa"),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    if liquidacion.estado == Liquidacion.Estados.PAGADA:
        messages.error(request, "No puedes anular una liquidación ya pagada.")
        return redirect("liquidacion_detalle", pk=pk)

    liquidacion.estado = Liquidacion.Estados.ANULADA
    liquidacion.save(update_fields=["estado", "actualizado_en"])

    registrar_historial(
        request,
        "Liquidaciones",
        "Anular",
        f"Se anuló la liquidación de {liquidacion.funcionario.nombre_completo}."
    )

    messages.success(request, "Liquidación anulada correctamente.")
    return redirect("liquidacion_detalle", pk=pk)


@login_required
def liquidacion_pdf(request, pk):
    permiso = validar_permiso_o_redirigir(request, "liquidacion", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    liquidacion = get_object_or_404(
        Liquidacion.objects.select_related(
            "funcionario",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa"
        ),
        pk=pk
    )

    if not admin_master:
        if not liquidacion.funcionario.sucursal_rel or liquidacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes operar liquidaciones de otra empresa.")
            return redirect("liquidaciones_lista")

    funcionario = liquidacion.funcionario
    config = ConfiguracionGeneral.obtener()
    alertas_proteccion = _liq_alertas_proteccion_guardadas(liquidacion)
    ajustes_manuales_pdf = liquidacion.ajustes_manuales.filter(
        estado=AjusteManualLiquidacion.Estados.ACTIVO
    ).select_related("creado_por")
    preaviso_dias_faltantes = max(
        int(liquidacion.preaviso_dias_corresponde or 0) - int(liquidacion.preaviso_dias_otorgados or 0),
        0,
    )

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="TituloClockIn",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=8,
    ))

    styles.add(ParagraphStyle(
        name="SubtituloClockIn",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#475569"),
        spaceAfter=12,
    ))

    styles.add(ParagraphStyle(
        name="SeccionClockIn",
        parent=styles["Heading3"],
        fontName="Helvetica-Bold",
        fontSize=11,
        leading=14,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#1d4ed8"),
        spaceAfter=6,
    ))

    styles.add(ParagraphStyle(
        name="TextoClockIn",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#111827"),
    ))

    styles.add(ParagraphStyle(
        name="TextoBoldClockIn",
        parent=styles["Normal"],
        fontName="Helvetica-Bold",
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#111827"),
    ))

    elementos = []

    nombre_sistema = config.nombre_sistema if config and config.nombre_sistema else "ClockIn"
    subtitulo = config.subtitulo_sistema if config and config.subtitulo_sistema else "Sistema Web RRHH"

    empresa_pdf = obtener_empresa_documento(funcionario=liquidacion.funcionario)

    elementos += construir_encabezado_empresa_pdf(
        empresa_pdf,
        "LIQUIDACIÓN LABORAL"
    )

    elementos.append(
        Paragraph(
            "LIQUIDACIÓN FINAL",
            styles["TituloClockIn"]
        )
    )

    elementos.append(Spacer(1, 4))

    datos_superiores = [
        ["Funcionario", funcionario.nombre_completo],
        ["Documento", funcionario.documento_compacto],
        ["Tipo de salida", liquidacion.get_tipo_salida_display()],
        ["Estado", liquidacion.get_estado_display()],
        ["Fecha de salida", liquidacion.fecha_salida.strftime("%d/%m/%Y") if liquidacion.fecha_salida else "-"],
        ["Fecha de cálculo", liquidacion.fecha_calculo.strftime("%d/%m/%Y") if liquidacion.fecha_calculo else "-"],
        ["Antigüedad", f"{liquidacion.antiguedad_anios} año(s), {liquidacion.antiguedad_meses} mes(es), {liquidacion.antiguedad_dias} día(s)"],
    ]

    tabla_datos = Table(
        datos_superiores,
        colWidths=[55 * mm, 105 * mm]
    )

    tabla_datos.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabla_datos)
    elementos.append(Spacer(1, 12))

    modalidad_salarial_pdf = (
        "Salario diferenciado"
        if liquidacion.modalidad_salarial_snapshot == "diferenciado"
        else "Salario base + bono"
    )

    tabla_base_salarial = Table([
        ["Base salarial historica", ""],
        ["Modalidad salarial", modalidad_salarial_pdf],
        ["Salario base snapshot", f"Gs. {liquidacion.salario_base_snapshot:,.0f}".replace(",", ".")],
        ["Bono base snapshot", f"Gs. {liquidacion.bono_base_snapshot:,.0f}".replace(",", ".")],
        ["Salario diferenciado snapshot", f"Gs. {liquidacion.salario_diferenciado_snapshot:,.0f}".replace(",", ".")],
        ["Salario bruto aplicable", f"Gs. {liquidacion.salario_bruto_aplicable_snapshot:,.0f}".replace(",", ".")],
        ["IPS aplicado", f"{liquidacion.porcentaje_ips_snapshot}% / Gs. {liquidacion.descuento_ips_calculado_snapshot:,.0f}".replace(",", ".")],
    ], colWidths=[70 * mm, 90 * mm])

    tabla_base_salarial.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2fe")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#075985")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#f0f9ff")),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#bae6fd")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabla_base_salarial)
    elementos.append(Spacer(1, 12))

    elementos.append(
        Paragraph(
            "1. HABERES",
            styles["SeccionClockIn"]
        )
    )

    tabla_haberes = Table([
        ["Concepto", "Detalle", "Monto"],
        ["Salario pendiente bruto", f"{liquidacion.dias_trabajados_pendientes} día(s)", f"Gs. {(liquidacion.salario_pendiente_monto + liquidacion.descuento_ausencias):,.0f}".replace(",", ".")],
        ["Ausencias descontadas", f"{liquidacion.ausencias_descuento} día(s)", f"- Gs. {liquidacion.descuento_ausencias:,.0f}".replace(",", ".")],
        ["Salario pendiente neto", "-", f"Gs. {liquidacion.salario_pendiente_monto:,.0f}".replace(",", ".")],
        ["Vacaciones causadas pendientes", f"{liquidacion.vacaciones_causadas_pendientes_dias} día(s)", f"Gs. {liquidacion.vacaciones_causadas_monto:,.0f}".replace(",", ".")],
        ["Vacaciones proporcionales", f"{liquidacion.vacaciones_proporcionales_dias} día(s)", f"Gs. {liquidacion.vacaciones_proporcionales_monto:,.0f}".replace(",", ".")],
        ["Aguinaldo proporcional", "-", f"Gs. {liquidacion.aguinaldo_proporcional_monto:,.0f}".replace(",", ".")],
        ["Preaviso", f"{liquidacion.preaviso_dias_corresponde} día(s) corresponde / {liquidacion.preaviso_dias_otorgados} día(s) otorgado(s)", f"Gs. {liquidacion.preaviso_monto:,.0f}".replace(",", ".")],
        ["Indemnización", f"{liquidacion.indemnizacion_dias} día(s)", f"Gs. {liquidacion.indemnizacion_monto:,.0f}".replace(",", ".")],
        ["Preaviso pendiente", f"{preaviso_dias_faltantes} dia(s) faltante(s)", "-"],
        ["TOTAL HABERES", "", f"Gs. {liquidacion.total_haberes:,.0f}".replace(",", ".")],
    ], colWidths=[70 * mm, 65 * mm, 25 * mm])

    tabla_haberes.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e3a8a")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, -2), "Helvetica"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#eff6ff")),
        ("ALIGN", (2, 0), (2, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabla_haberes)
    elementos.append(Spacer(1, 12))

    elementos.append(
        Paragraph(
            "2. DESCUENTOS",
            styles["SeccionClockIn"]
        )
    )

    tabla_desc = Table([
        ["Concepto", "Monto"],
        ["IPS", f"Gs. {liquidacion.ips_monto:,.0f}".replace(",", ".")],
        ["Deudas", f"Gs. {liquidacion.deudas_monto:,.0f}".replace(",", ".")],
        ["Otros descuentos", f"Gs. {liquidacion.otros_descuentos:,.0f}".replace(",", ".")],
        ["TOTAL DESCUENTOS", f"Gs. {liquidacion.total_descuentos:,.0f}".replace(",", ".")],
    ], colWidths=[135 * mm, 25 * mm])

    tabla_desc.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fee2e2")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#991b1b")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fef2f2")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#fecaca")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))

    elementos.append(tabla_desc)
    elementos.append(Spacer(1, 12))

    elementos.append(
        Paragraph(
            "3. TOTAL FINAL",
            styles["SeccionClockIn"]
        )
    )

    tabla_total = Table([
        ["TOTAL LIQUIDACIÓN", f"Gs. {liquidacion.total_liquidacion:,.0f}".replace(",", ".")]
    ], colWidths=[135 * mm, 25 * mm])

    tabla_total.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#dcfce7")),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#166534")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#86efac")),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elementos.append(tabla_total)
    elementos.append(Spacer(1, 14))

    if ajustes_manuales_pdf.exists():
        elementos.append(
            Paragraph(
                "4. AJUSTES MANUALES",
                styles["SeccionClockIn"]
            )
        )

        data_ajustes = [[
            "Tipo",
            "Concepto",
            "Automatico",
            "Ajustado",
            "Diferencia",
            "Usuario",
        ]]

        for ajuste in ajustes_manuales_pdf:
            data_ajustes.append([
                ajuste.get_tipo_display(),
                ajuste.concepto,
                f"Gs. {ajuste.importe_anterior:,.0f}".replace(",", "."),
                f"Gs. {ajuste.importe_nuevo:,.0f}".replace(",", "."),
                f"Gs. {ajuste.diferencia:,.0f}".replace(",", "."),
                str(ajuste.creado_por or "-"),
            ])

        data_ajustes.append([
            "Total",
            "Resultado final",
            f"Gs. {liquidacion.total_liquidacion_automatico:,.0f}".replace(",", "."),
            "",
            "",
            f"Gs. {liquidacion.total_liquidacion:,.0f}".replace(",", "."),
        ])

        tabla_ajustes = Table(data_ajustes, colWidths=[22 * mm, 45 * mm, 25 * mm, 25 * mm, 25 * mm, 28 * mm])
        tabla_ajustes.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fef3c7")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#92400e")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fffbeb")),
            ("ALIGN", (2, 1), (5, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#fde68a")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))

        elementos.append(tabla_ajustes)
        elementos.append(Spacer(1, 12))

    if liquidacion.motivo_observacion:
        elementos.append(
            Paragraph(
                "4. OBSERVACIÓN",
                styles["SeccionClockIn"]
            )
        )

        elementos.append(
            Paragraph(
                liquidacion.motivo_observacion.replace("\n", "<br/>"),
                styles["TextoClockIn"]
            )
        )

        elementos.append(Spacer(1, 12))

    if liquidacion.requiere_revision_juridica and liquidacion.alerta_revision:
        elementos.append(
            Paragraph(
                "5. ALERTA DE REVISIÓN",
                styles["SeccionClockIn"]
            )
        )

        elementos.append(
            Paragraph(
                liquidacion.alerta_revision,
                styles["TextoBoldClockIn"]
            )
        )

        elementos.append(Spacer(1, 12))

    if alertas_proteccion:
        elementos.append(
            Paragraph(
                "ALERTAS DE PROTECCION PARA LA EMPRESA",
                styles["SeccionClockIn"]
            )
        )

        for alerta in alertas_proteccion:
            elementos.append(
                Paragraph(
                    f"- {alerta}",
                    styles["TextoClockIn"]
                )
            )

        elementos.append(Spacer(1, 12))

    elementos.append(Spacer(1, 22))

    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma del empleador / responsable", "Firma del funcionario"],
        ["", ""],
        ["Aclaración: ____________________", "Aclaración: ____________________"],
    ], colWidths=[80 * mm, 80 * mm])

    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica"),
        ("FONTNAME", (0, 3), (-1, 3), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="LIQUIDACION",
        documento_id=liquidacion.id,
        funcionario=funcionario,
        titulo="Liquidación Laboral",
    )

    agregar_texto_legal_empresa_pdf(
        elementos,
        empresa_pdf
    )

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        f'inline; filename="liquidacion_{funcionario.cedula}_{liquidacion.id}.pdf"'
    )

    response.write(pdf)

    return response

def generar_texto_comunicacion_laboral(comunicacion):
    funcionario = comunicacion.funcionario
    empresa_nombre = funcionario.empresa_mostrar
    fecha_emision = comunicacion.fecha_emision.strftime("%d/%m/%Y") if comunicacion.fecha_emision else "-"
    fecha_ref = comunicacion.fecha_referencia.strftime("%d/%m/%Y") if comunicacion.fecha_referencia else "la fecha indicada"
    detalle = comunicacion.detalle_hecho or "los hechos detallados por la empresa"

    tipo = comunicacion.tipo

    textos = {
        ComunicacionLaboral.Tipos.AMONESTACION: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica formalmente al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, que se deja constancia
de una <b>amonestación disciplinaria</b> vinculada a los siguientes hechos:
<br/><br/>
<b>{detalle}</b>
<br/><br/>
La presente comunicación se realiza a los efectos de documentar la situación ocurrida, advertir al trabajador
sobre la necesidad de cumplir estrictamente con sus obligaciones laborales, reglamentos internos, horarios,
instrucciones de trabajo y normas de conducta aplicables.
<br/><br/>
Se deja constancia de que la reiteración de hechos similares podrá dar lugar a nuevas medidas disciplinarias,
conforme a la gravedad del caso, los antecedentes existentes y la normativa laboral vigente.
""",

        ComunicacionLaboral.Tipos.PREAVISO: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica formalmente al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, el otorgamiento de
<b>preaviso</b> conforme a la normativa laboral vigente.
<br/><br/>
La presente comunicación se emite en fecha <b>{fecha_emision}</b>, tomando como referencia
<b>{fecha_ref}</b>.
<br/><br/>
El plazo de preaviso deberá ser determinado conforme a la antigüedad del trabajador y a las disposiciones
aplicables del Código del Trabajo. Se deja constancia de que este documento tiene por finalidad comunicar
formalmente la decisión empresarial y preservar respaldo documental suficiente.
<br/><br/>
Detalle adicional:
<br/>
<b>{detalle}</b>
""",

        ComunicacionLaboral.Tipos.ABANDONO: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica formalmente al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, que se ha registrado una
situación compatible con <b>abandono de trabajo</b> o inasistencia injustificada, según los antecedentes
obrantes en la empresa.
<br/><br/>
Hechos registrados:
<br/>
<b>{detalle}</b>
<br/><br/>
Se intima al trabajador a justificar de manera inmediata y documentada su ausencia o interrupción de tareas.
La falta de justificación suficiente podrá ser considerada para la adopción de las medidas laborales que
correspondan conforme al Código del Trabajo y demás normas aplicables.
<br/><br/>
La presente comunicación se emite a efectos de dejar constancia formal y permitir el ejercicio del derecho
a formular aclaraciones o descargos.
""",

        ComunicacionLaboral.Tipos.PERMISO: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, la recepción, autorización,
rechazo o registro administrativo de un <b>permiso laboral</b>, conforme a los datos consignados por RRHH.
<br/><br/>
Detalle del permiso:
<br/>
<b>{detalle}</b>
<br/><br/>
El trabajador deberá cumplir con las condiciones, fechas, horarios y documentación respaldatoria indicadas
por la empresa. En caso de no presentar los justificativos correspondientes, la situación podrá ser tratada
como ausencia injustificada.
""",

        ComunicacionLaboral.Tipos.AUSENCIA: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica formalmente al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, que se ha registrado una
<b>ausencia injustificada</b> o una falta de comunicación previa respecto a su inasistencia.
<br/><br/>
Detalle registrado:
<br/>
<b>{detalle}</b>
<br/><br/>
Se solicita al trabajador presentar la justificación correspondiente en el plazo más breve posible. La falta
de justificación documentada podrá generar constancia disciplinaria y ser considerada como antecedente
laboral, conforme a la normativa vigente y a los reglamentos internos aplicables.
""",

        ComunicacionLaboral.Tipos.SUSPENSION: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, la aplicación o inicio del
procedimiento relacionado con una <b>suspensión disciplinaria</b>, según los hechos detallados.
<br/><br/>
Hechos:
<br/>
<b>{detalle}</b>
<br/><br/>
La medida deberá ser proporcional a la falta imputada, respetar el derecho de defensa del trabajador y
documentarse debidamente. En caso de corresponder, el trabajador podrá presentar su descargo o explicación
ante RRHH dentro del plazo otorgado por la empresa.
""",

        ComunicacionLaboral.Tipos.CITACION_DESCARGO: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> cita formalmente al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, a presentar su
<b>descargo</b> respecto a los hechos comunicados.
<br/><br/>
Hechos objeto de descargo:
<br/>
<b>{detalle}</b>
<br/><br/>
La finalidad de esta comunicación es garantizar que el trabajador pueda exponer su versión, presentar
documentos o aclaraciones, y ejercer su derecho de defensa antes de que la empresa adopte una decisión
disciplinaria o administrativa definitiva.
""",

        ComunicacionLaboral.Tipos.CAMBIO_CARGO_SECTOR: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> comunica al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>, una modificación,
reasignación o cambio administrativo relacionado con su cargo, sector, sucursal, funciones u organización
interna del trabajo.
<br/><br/>
Detalle:
<br/>
<b>{detalle}</b>
<br/><br/>
La presente comunicación se emite para dejar constancia formal de la medida administrativa, sin perjuicio
de los derechos laborales que correspondan al trabajador conforme a la normativa vigente.
""",

        ComunicacionLaboral.Tipos.MEMORANDUM: f"""
Por medio de la presente, la empresa <b>{empresa_nombre}</b> emite memorándum interno dirigido al trabajador
<b>{funcionario.nombre_completo}</b>, identificado como <b>{funcionario.documento_compacto}</b>.
<br/><br/>
Asunto:
<br/>
<b>{comunicacion.asunto or comunicacion.titulo}</b>
<br/><br/>
Detalle:
<br/>
<b>{detalle}</b>
<br/><br/>
Se deja constancia de la presente comunicación para fines administrativos, organizativos y documentales.
""",
    }

    return textos.get(tipo, detalle)


@login_required
def comunicaciones_lista(request):
    permiso = validar_permiso_o_redirigir(request, "comunicaciones", "puede_ver")
    if permiso:
        return permiso

    q = request.GET.get("q", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    estado = request.GET.get("estado", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    comunicaciones = ComunicacionLaboral.objects.select_related(
        "funcionario",
        "empresa",
        "sucursal",
        "generado_por",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).all()

    if not admin_master:
        if empresa_usuario:
            comunicaciones = comunicaciones.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            comunicaciones = comunicaciones.none()

    if q:
        comunicaciones = comunicaciones.filter(
            Q(funcionario__nombre__icontains=q) |
            Q(funcionario__apellido__icontains=q) |
            _documento_busqueda_q("funcionario__cedula", "funcionario__tipo_documento", q) |
            Q(titulo__icontains=q) |
            Q(asunto__icontains=q) |
            Q(detalle_hecho__icontains=q)
        )

    if tipo:
        comunicaciones = comunicaciones.filter(tipo=tipo)

    if estado:
        comunicaciones = comunicaciones.filter(estado=estado)

    return render(request, "core/comunicaciones_lista.html", {
        "comunicaciones": comunicaciones,
        "q": q,
        "tipo": tipo,
        "estado": estado,
        "tipos": ComunicacionLaboral.Tipos.choices,
        "estados": ComunicacionLaboral.Estados.choices,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def comunicacion_nueva(request):
    permiso = validar_permiso_o_redirigir(request, "comunicaciones", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = ComunicacionLaboralForm(request.POST, request.FILES)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            comunicacion = form.save(commit=False)

            if not admin_master:
                if not comunicacion.funcionario.sucursal_rel or comunicacion.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear comunicaciones para otra empresa.")
                    return redirect("comunicaciones_lista")

            comunicacion.empresa = comunicacion.funcionario.empresa
            comunicacion.sucursal = comunicacion.funcionario.sucursal_rel
            comunicacion.generado_por = request.user

            if not comunicacion.contenido:
                comunicacion.contenido = generar_texto_comunicacion_laboral(comunicacion)

            comunicacion.save()

            registrar_historial(
                request,
                "Comunicaciones",
                "Crear",
                f"Se creó comunicación {comunicacion.get_tipo_display()} para {comunicacion.funcionario.nombre_completo}."
            )

            messages.success(request, "Comunicación creada correctamente.")
            return redirect("comunicacion_detalle", pk=comunicacion.pk)
    else:
        form = ComunicacionLaboralForm(initial={"fecha_emision": timezone.localdate()})

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/comunicacion_form.html", {
        "form": form,
        "titulo_form": "Nueva comunicación",
        "boton_texto": "Guardar comunicación",
    })


@login_required
def comunicacion_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "comunicaciones", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    comunicacion = get_object_or_404(
        ComunicacionLaboral.objects.select_related(
            "funcionario",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa"
        ),
        pk=pk
    )

    if not admin_master:
        if not comunicacion.funcionario.sucursal_rel or comunicacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes editar comunicaciones de otra empresa.")
            return redirect("comunicaciones_lista")

    if request.method == "POST":
        form = ComunicacionLaboralForm(request.POST, request.FILES, instance=comunicacion)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

        if form.is_valid():
            comunicacion = form.save(commit=False)

            if not admin_master:
                if not comunicacion.funcionario.sucursal_rel or comunicacion.funcionario.sucursal_rel.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover comunicaciones a otra empresa.")
                    return redirect("comunicaciones_lista")

            comunicacion.empresa = comunicacion.funcionario.empresa
            comunicacion.sucursal = comunicacion.funcionario.sucursal_rel
            comunicacion.save()

            registrar_historial(
                request,
                "Comunicaciones",
                "Editar",
                f"Se editó comunicación de {comunicacion.funcionario.nombre_completo}."
            )

            messages.success(request, "Comunicación actualizada correctamente.")
            return redirect("comunicacion_detalle", pk=comunicacion.pk)
    else:
        form = ComunicacionLaboralForm(instance=comunicacion)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")

    return render(request, "core/comunicacion_form.html", {
        "form": form,
        "titulo_form": "Editar comunicación",
        "boton_texto": "Guardar cambios",
        "comunicacion": comunicacion,
    })


@login_required
def comunicacion_detalle(request, pk):
    permiso = validar_permiso_o_redirigir(request, "comunicaciones", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    comunicacion = get_object_or_404(
        ComunicacionLaboral.objects.select_related("funcionario", "empresa", "sucursal", "generado_por"),
        pk=pk
    )

    if not admin_master:
        if not comunicacion.funcionario.sucursal_rel or comunicacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes ver comunicaciones de otra empresa.")
            return redirect("comunicaciones_lista")

    return render(request, "core/comunicacion_detalle.html", {
        "comunicacion": comunicacion,
    })


@login_required
def comunicacion_pdf(request, pk):
    permiso = validar_permiso_o_redirigir(request, "comunicaciones", "puede_ver")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    comunicacion = get_object_or_404(
        ComunicacionLaboral.objects.select_related(
            "funcionario",
            "empresa",
            "sucursal",
            "generado_por",
            "funcionario__sucursal_rel",
            "funcionario__sucursal_rel__empresa",
        ),
        pk=pk
    )

    if not admin_master:
        if not comunicacion.funcionario.sucursal_rel or comunicacion.funcionario.sucursal_rel.empresa != empresa_usuario:
            messages.error(request, "No puedes generar PDF de otra empresa.")
            return redirect("comunicaciones_lista")

    funcionario = comunicacion.funcionario
    empresa_pdf = obtener_empresa_documento(funcionario=funcionario)

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=20 * mm,
        rightMargin=20 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
    )

    styles = getSampleStyleSheet()

    styles.add(ParagraphStyle(
        name="TituloComunicacion",
        parent=styles["Heading1"],
        fontName="Helvetica-Bold",
        fontSize=15,
        leading=19,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#111827"),
        spaceAfter=10,
    ))

    styles.add(ParagraphStyle(
        name="TextoComunicacion",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=15,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#111827"),
    ))

    elementos = []

    elementos += construir_encabezado_empresa_pdf(
        empresa_pdf,
        "COMUNICACIÓN LABORAL"
    )

    elementos.append(
        Paragraph(
            comunicacion.titulo.upper(),
            styles["TituloComunicacion"]
        )
    )

    elementos.append(Spacer(1, 8))

    datos = [
        ["Funcionario", funcionario.nombre_completo],
        ["Documento", funcionario.documento_compacto],
        ["Cargo", funcionario.cargo or "-"],
        ["Sector", funcionario.sector or "-"],
        ["Sucursal", funcionario.sucursal_mostrar],
        ["Tipo de comunicación", comunicacion.get_tipo_display()],
        ["Fecha de emisión", comunicacion.fecha_emision.strftime("%d/%m/%Y")],
        ["Fecha de referencia", comunicacion.fecha_referencia.strftime("%d/%m/%Y") if comunicacion.fecha_referencia else "-"],
        ["Estado", comunicacion.get_estado_display()],
    ]

    tabla = Table(
        datos,
        colWidths=[55 * mm, 105 * mm]
    )

    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eff6ff")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1d4ed8")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))

    elementos.append(tabla)
    elementos.append(Spacer(1, 14))

    contenido = comunicacion.contenido or generar_texto_comunicacion_laboral(comunicacion)
    contenido = contenido.replace("\n", "<br/>")

    elementos.append(
        Paragraph(
            contenido,
            styles["TextoComunicacion"]
        )
    )

    elementos.append(Spacer(1, 30))

    firmas = Table([
        ["_______________________________", "_______________________________"],
        ["Firma del empleador / RRHH", "Firma del funcionario"],
        ["", ""],
        ["Aclaración: ____________________", "Aclaración: ____________________"],
        ["Fecha: ____/____/______", "Fecha: ____/____/______"],
    ], colWidths=[80 * mm, 80 * mm])

    firmas.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))

    elementos.append(firmas)

    agregar_firma_qr_documento_pdf(
        elementos=elementos,
        request=request,
        empresa=empresa_pdf,
        tipo_documento="COMUNICACION",
        documento_id=comunicacion.id,
        funcionario=funcionario,
        titulo=comunicacion.titulo,
    )

    agregar_texto_legal_empresa_pdf(
        elementos,
        empresa_pdf
    )

    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    registrar_historial(
        request,
        "Comunicaciones",
        "PDF",
        f"Se generó PDF de comunicación para {funcionario.nombre_completo}."
    )

    response = HttpResponse(content_type="application/pdf")

    response["Content-Disposition"] = (
        f'inline; filename="comunicacion_{funcionario.cedula}_{comunicacion.id}.pdf"'
    )

    response.write(pdf)

    return response

@login_required
def dias_libres_lista(request):
    permiso = validar_permiso_o_redirigir(request, "dias_libres", "puede_ver")
    if permiso:
        return permiso

    empresa_id = request.GET.get("empresa", "").strip()
    sucursal_id = request.GET.get("sucursal", "").strip()
    sector = request.GET.get("sector", "").strip()
    q = request.GET.get("q", "").strip()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionarios_activos = Funcionario.objects.filter(activo=True).select_related(
        "turno",
        "sucursal_rel",
        "sucursal_rel__empresa"
    )

    if not admin_master:
        if empresa_usuario:
            empresa_id = str(empresa_usuario.id)
            funcionarios_activos = funcionarios_activos.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios_activos = funcionarios_activos.none()

    if empresa_id:
        funcionarios_activos = funcionarios_activos.filter(sucursal_rel__empresa_id=empresa_id)

    if sucursal_id:
        funcionarios_activos = funcionarios_activos.filter(sucursal_rel_id=sucursal_id)

    if sector:
        funcionarios_activos = funcionarios_activos.filter(sector=sector)

    if q:
        funcionarios_activos = funcionarios_activos.filter(
            Q(nombre__icontains=q) |
            Q(apellido__icontains=q) |
            _documento_busqueda_q("cedula", "tipo_documento", q) |
            Q(sector__icontains=q)
        )

    if request.method == "POST":
        permiso_post = validar_permiso_o_redirigir(request, "dias_libres", "puede_crear")
        if permiso_post:
            return permiso_post

        dias_campos = [
            "lunes",
            "martes",
            "miercoles",
            "jueves",
            "viernes",
            "sabado",
            "domingo",
        ]

        mapa_dias = {
            "lunes": 0,
            "martes": 1,
            "miercoles": 2,
            "jueves": 3,
            "viernes": 4,
            "sabado": 5,
            "domingo": 6,
        }

        for funcionario in funcionarios_activos:
            planilla, creada = PlanillaSemanalFuncionario.objects.get_or_create(
                funcionario=funcionario
            )

            for campo in dias_campos:
                valor_turno = request.POST.get(f"{campo}_{funcionario.id}", "").strip()

                if valor_turno == "":
                    setattr(planilla, campo, None)
                else:
                    try:
                        turno = Turno.objects.get(pk=valor_turno, activo=True)

                        if not admin_master and empresa_usuario:
                            if turno.empresa != empresa_usuario:
                                continue

                        setattr(planilla, campo, turno)
                    except Turno.DoesNotExist:
                        setattr(planilla, campo, None)

            planilla.save()

            # Compatibilidad con el sistema viejo de DiaLibre:
            # La tabla DiaLibre solo permite 1 día libre activo por funcionario.
            # Por eso tomamos el PRIMER día vacío como día libre principal.
            dias_libres_detectados = []

            for campo, numero_dia in mapa_dias.items():
                if getattr(planilla, campo) is None:
                    dias_libres_detectados.append(numero_dia)

            # Eliminamos duplicados activos antiguos para evitar IntegrityError.
            dias_libres_activos = DiaLibre.objects.filter(
                funcionario=funcionario,
                activo=True
            ).order_by("id")

            dia_libre_principal = dias_libres_detectados[0] if dias_libres_detectados else None

            if dia_libre_principal is not None:
                dia_libre_obj = dias_libres_activos.first()

                if dia_libre_obj:
                    dia_libre_obj.empresa = funcionario.empresa
                    dia_libre_obj.sucursal = funcionario.sucursal_rel
                    dia_libre_obj.sector = funcionario.sector or ""
                    dia_libre_obj.dia_semana = dia_libre_principal
                    dia_libre_obj.fecha_inicio = timezone.localdate()
                    dia_libre_obj.activo = True
                    dia_libre_obj.observacion = "Generado automáticamente desde planilla semanal."
                    dia_libre_obj.save()

                    dias_libres_activos.exclude(pk=dia_libre_obj.pk).delete()
                else:
                    DiaLibre.objects.create(
                        funcionario=funcionario,
                        empresa=funcionario.empresa,
                        sucursal=funcionario.sucursal_rel,
                        sector=funcionario.sector or "",
                        dia_semana=dia_libre_principal,
                        fecha_inicio=timezone.localdate(),
                        activo=True,
                        observacion="Generado automáticamente desde planilla semanal."
                    )
            else:
                dias_libres_activos.delete()

        registrar_historial(
            request,
            "Días Libres",
            "Planilla semanal",
            f"Se actualizó la planilla semanal de turnos. Sector: {sector or 'Todos'}."
        )

        messages.success(request, "Planilla semanal actualizada correctamente.")
        return redirect(f"/dias-libres/?empresa={empresa_id}&sucursal={sucursal_id}&sector={sector}&q={q}")

    if admin_master:
        empresas = Empresa.objects.filter(activo=True).order_by("nombre")
        sucursales = Sucursal.objects.filter(activo=True).order_by("nombre")
        turnos = Turno.objects.filter(activo=True).select_related("empresa").order_by("nombre")

        if empresa_id:
            sucursales = sucursales.filter(empresa_id=empresa_id)
            turnos = turnos.filter(empresa_id=empresa_id)
    else:
        empresas = Empresa.objects.filter(id=empresa_usuario.id) if empresa_usuario else Empresa.objects.none()
        sucursales = Sucursal.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Sucursal.objects.none()
        turnos = Turno.objects.filter(
            activo=True,
            empresa=empresa_usuario
        ).order_by("nombre") if empresa_usuario else Turno.objects.none()

    sectores_qs = Funcionario.objects.filter(activo=True).exclude(sector="")

    if not admin_master and empresa_usuario:
        sectores_qs = sectores_qs.filter(sucursal_rel__empresa=empresa_usuario)

    if empresa_id:
        sectores_qs = sectores_qs.filter(sucursal_rel__empresa_id=empresa_id)

    if sucursal_id:
        sectores_qs = sectores_qs.filter(sucursal_rel_id=sucursal_id)

    sectores = sectores_qs.values_list("sector", flat=True).distinct().order_by("sector")

    total_funcionarios = funcionarios_activos.count()

    funcionarios_rapidos = []

    total_con_planilla = 0
    total_pendientes = 0

    for funcionario in funcionarios_activos.order_by("apellido", "nombre"):
        planilla, creada = PlanillaSemanalFuncionario.objects.get_or_create(
            funcionario=funcionario
        )

        tiene_algo = any([
            planilla.lunes,
            planilla.martes,
            planilla.miercoles,
            planilla.jueves,
            planilla.viernes,
            planilla.sabado,
            planilla.domingo,
        ])

        if tiene_algo:
            total_con_planilla += 1
        else:
            total_pendientes += 1

        funcionarios_rapidos.append({
            "funcionario": funcionario,
            "planilla": planilla,
        })

    return render(request, "core/dias_libres_lista.html", {
        "funcionarios_rapidos": funcionarios_rapidos,
        "turnos": turnos,
        "empresas": empresas,
        "sucursales": sucursales,
        "sectores": sectores,
        "empresa_id": empresa_id,
        "sucursal_id": sucursal_id,
        "sector": sector,
        "q": q,
        "total_funcionarios": total_funcionarios,
        "total_asignados": total_con_planilla,
        "total_pendientes": total_pendientes,
        "empresa_usuario": empresa_usuario,
        "es_admin_master": admin_master,
    })


@login_required
def dia_libre_nuevo(request):
    permiso = validar_permiso_o_redirigir(request, "dias_libres", "puede_crear")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    if request.method == "POST":
        form = DiaLibreForm(request.POST)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

        if form.is_valid():
            dia_libre = form.save(commit=False)

            if not admin_master:
                if dia_libre.empresa != empresa_usuario:
                    messages.error(request, "No puedes crear días libres para otra empresa.")
                    return redirect("dias_libres_lista")

            dia_libre.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Días Libres",
                "Crear",
                f"Se asignó día libre {dia_libre.get_dia_semana_display()} a {dia_libre.funcionario.nombre_completo}."
            )
            messages.success(request, "Día libre asignado correctamente.")
            return redirect("dias_libres_lista")
    else:
        form = DiaLibreForm()

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["empresa"].initial = empresa_usuario
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

    return render(request, "core/dia_libre_form.html", {
        "form": form,
        "titulo_form": "Nuevo día libre",
        "boton_texto": "Guardar día libre",
    })


@login_required
def dia_libre_editar(request, pk):
    permiso = validar_permiso_o_redirigir(request, "dias_libres", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    dia_libre = get_object_or_404(DiaLibre, pk=pk)

    if not admin_master:
        if dia_libre.empresa != empresa_usuario:
            messages.error(request, "No puedes editar días libres de otra empresa.")
            return redirect("dias_libres_lista")

    if request.method == "POST":
        form = DiaLibreForm(request.POST, instance=dia_libre)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

        if form.is_valid():
            dia_libre_editado = form.save(commit=False)

            if not admin_master:
                if dia_libre_editado.empresa != empresa_usuario:
                    messages.error(request, "No puedes mover días libres a otra empresa.")
                    return redirect("dias_libres_lista")

            dia_libre_editado.save()
            form.save_m2m()

            registrar_historial(
                request,
                "Días Libres",
                "Editar",
                f"Se editó día libre de {dia_libre_editado.funcionario.nombre_completo}."
            )
            messages.success(request, "Día libre actualizado correctamente.")
            return redirect("dias_libres_lista")
    else:
        form = DiaLibreForm(instance=dia_libre)

        if not admin_master and empresa_usuario:
            form.fields["funcionario"].queryset = Funcionario.objects.filter(
                activo=True,
                sucursal_rel__empresa=empresa_usuario
            ).order_by("apellido", "nombre")
            form.fields["empresa"].queryset = Empresa.objects.filter(id=empresa_usuario.id)
            form.fields["sucursal"].queryset = Sucursal.objects.filter(
                activo=True,
                empresa=empresa_usuario
            ).order_by("nombre")

    return render(request, "core/dia_libre_form.html", {
        "form": form,
        "titulo_form": "Editar día libre",
        "boton_texto": "Actualizar día libre",
    })


@login_required
def dia_libre_toggle_activo(request, pk):
    permiso = validar_permiso_o_redirigir(request, "dias_libres", "puede_editar")
    if permiso:
        return permiso

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    dia_libre = get_object_or_404(DiaLibre, pk=pk)

    if not admin_master:
        if not empresa_usuario or dia_libre.empresa != empresa_usuario:
            messages.error(request, "No puedes cambiar dias libres de otra empresa.")
            return redirect("dias_libres_lista")

    dia_libre.activo = not dia_libre.activo
    dia_libre.save(update_fields=["activo", "actualizado_en"])

    registrar_historial(
        request,
        "Días Libres",
        "Estado",
        f"Se cambió estado de día libre de {dia_libre.funcionario.nombre_completo} a {'Activo' if dia_libre.activo else 'Inactivo'}."
    )

    messages.success(request, "Estado del día libre actualizado correctamente.")
    return redirect("dias_libres_lista")

@login_required
def asistencia_eliminar(request, pk):
    if not es_admin_master(request.user):
        messages.error(request, "Solo el administrador global puede eliminar asistencias.")
        return redirect("asistencia_marcar")

    empresa_usuario = obtener_empresa_activa(request)
    asistencias_qs = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__sucursal_rel",
    )
    if empresa_usuario:
        asistencias_qs = asistencias_qs.filter(
            funcionario__sucursal_rel__empresa=empresa_usuario
        )

    asistencia = get_object_or_404(
        asistencias_qs,
        pk=pk
    )

    funcionario = asistencia.funcionario
    fecha = asistencia.fecha

    if request.method == "POST":
        motivo = request.POST.get("motivo", "").strip()

        if not motivo:
            messages.error(request, "Debes indicar el motivo de eliminación.")
            return redirect("asistencia_marcar")

        registrar_historial(
            request,
            "Asistencia",
            "Eliminar",
            f"Se eliminó asistencia de {funcionario.nombre_completo} "
            f"del día {fecha.strftime('%d/%m/%Y')}. Motivo: {motivo}"
        )

        asistencia.delete()

        messages.success(
            request,
            f"Asistencia de {funcionario.nombre_completo} eliminada correctamente."
        )
        return redirect("asistencia_marcar")

    return render(request, "core/asistencia_eliminar.html", {
        "asistencia": asistencia,
    })
from reportlab.lib.pagesizes import landscape

@login_required
def reporte_diario_pdf(request):
    permiso = validar_permiso_o_redirigir(request, "reportes", "puede_ver")
    if permiso:
        return permiso

    fecha_str = request.GET.get("fecha", str(timezone.localdate()))

    try:
        fecha_reporte = datetime.strptime(fecha_str, "%Y-%m-%d").date()
    except ValueError:
        fecha_reporte = timezone.localdate()

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    asistencias = Asistencia.objects.select_related(
        "funcionario",
        "funcionario__turno",
        "funcionario__sucursal_rel",
        "funcionario__sucursal_rel__empresa",
    ).filter(fecha=fecha_reporte)

    if not admin_master:
        if empresa_usuario:
            asistencias = asistencias.filter(funcionario__sucursal_rel__empresa=empresa_usuario)
        else:
            asistencias = asistencias.none()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    elementos = []
    empresa_pdf = empresa_usuario if empresa_usuario else None

    elementos += construir_encabezado_empresa_pdf(
        empresa_pdf,
        f"REPORTE DIARIO - {fecha_reporte.strftime('%d/%m/%Y')}"
    )

    data = [[
        "Funcionario", "CI", "Sucursal", "Turno", "Entrada", "Salida", "Atraso", "Estado"
    ]]

    for a in asistencias.order_by("funcionario__apellido", "funcionario__nombre"):
        data.append([
            a.funcionario.nombre_completo,
            a.funcionario.documento_compacto,
            a.funcionario.sucursal_mostrar,
            a.funcionario.turno.nombre if a.funcionario.turno else "-",
            a.hora_entrada.strftime("%H:%M") if a.hora_entrada else "-",
            a.hora_salida.strftime("%H:%M") if a.hora_salida else "-",
            f"{a.minutos_atraso} min",
            a.estado_jornada,
        ])

    tabla = Table(data, colWidths=[55*mm, 24*mm, 40*mm, 38*mm, 22*mm, 22*mm, 22*mm, 34*mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="reporte_diario_{fecha_reporte}.pdf"'
    response.write(pdf)
    return response


@login_required
def reporte_mensual_pdf(request):
    permiso = validar_permiso_o_redirigir(request, "reportes", "puede_ver")
    if permiso:
        return permiso

    hoy = timezone.localdate()
    mes = int(request.GET.get("mes", hoy.month))
    anio = int(request.GET.get("anio", hoy.year))

    empresa_usuario = obtener_empresa_activa(request)
    admin_master = es_admin_master(request.user) and empresa_usuario is None

    funcionarios = Funcionario.objects.filter(activo=True).select_related(
        "turno",
        "sucursal_rel",
        "sucursal_rel__empresa"
    )

    if not admin_master:
        if empresa_usuario:
            funcionarios = funcionarios.filter(sucursal_rel__empresa=empresa_usuario)
        else:
            funcionarios = funcionarios.none()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    elementos = []
    empresa_pdf = empresa_usuario if empresa_usuario else None

    elementos += construir_encabezado_empresa_pdf(
        empresa_pdf,
        f"REPORTE MENSUAL - {mes:02d}/{anio}"
    )

    data = [[
        "Funcionario", "CI", "Sucursal", "Cargo", "Asistencias", "Atrasos",
        "Días libres", "Ausencias", "Permisos", "Vacaciones", "ICL"
    ]]

    dias_mes = monthrange(anio, mes)[1]
    total_dias_laborales_estimados = sum(
        1 for dia in range(1, dias_mes + 1)
        if date(anio, mes, dia).weekday() != 6
    )

    for funcionario in funcionarios.order_by("apellido", "nombre"):
        asistencias_mes = Asistencia.objects.filter(
            funcionario=funcionario,
            fecha__year=anio,
            fecha__month=mes,
            hora_entrada__isnull=False,
        )

        asistencias_count = asistencias_mes.count()
        atrasos_count = asistencias_mes.filter(llego_tarde=True).count()
        dias_libres_mes = contar_dias_libres_mes(funcionario, mes, anio)

        permisos_aprobados = PermisoLicencia.objects.filter(
            funcionario=funcionario,
            estado=PermisoLicencia.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        vacaciones_aprobadas = Vacacion.objects.filter(
            funcionario=funcionario,
            estado=Vacacion.Estados.APROBADO,
            fecha_desde__year=anio,
            fecha_desde__month=mes,
        ).count()

        total_reales = max(total_dias_laborales_estimados - dias_libres_mes, 0)
        ausencias = max(total_reales - asistencias_count, 0)
        ausencias_no_justificadas = max(ausencias - permisos_aprobados - vacaciones_aprobadas, 0)

        icl = 100 - (atrasos_count * 2) - (ausencias_no_justificadas * 5)
        icl = max(0, min(100, icl))
        icl_activo = icl_activo_para_funcionario(funcionario)

        data.append([
            funcionario.nombre_completo,
            funcionario.documento_compacto,
            funcionario.sucursal_mostrar,
            funcionario.cargo or "-",
            asistencias_count,
            atrasos_count,
            dias_libres_mes,
            ausencias,
            permisos_aprobados,
            vacaciones_aprobadas,
            f"{icl}%" if icl_activo else f"{icl}% (sin impacto)",
        ])

    tabla = Table(data, colWidths=[45*mm, 21*mm, 34*mm, 32*mm, 22*mm, 20*mm, 22*mm, 22*mm, 22*mm, 24*mm, 18*mm])
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dbeafe")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 6.5),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))

    elementos.append(tabla)
    doc.build(elementos)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="reporte_mensual_{mes:02d}_{anio}.pdf"'
    response.write(pdf)
    return response


